#!/usr/bin/env python3
"""Append F-061 v2 spectral_mass to an existing research workbook.

Requires per-compartment ``ACD_D1_{harmonic,inharmonic,subbass}`` (and
the matching D0 / r columns). Workbooks that lack them are refused —
v1 pooled-count backfill is not applied silently.

Writes ``spectral_mass`` and ``spectral_mass_count`` immediately right of
``EWSD_score_acoustic_balanced``, applies the same blue data bars as a
fresh Stage 3 export, and saves alongside as ``<name>_massfilled.xlsx``.
Never overwrites the source workbook.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

_REPO = Path(__file__).resolve().parents[1]
if str(_REPO) not in sys.path:
    sys.path.insert(0, str(_REPO))

from tools.spectral_mass import (
    REQUIRED_D0_COLUMNS,
    REQUIRED_D1_COLUMNS,
    REQUIRED_R_COLUMNS,
    add_spectral_mass_column,
    apply_spectral_mass_data_bar,
    place_spectral_mass_right_of_ewsd,
)

SHEET = "Spectral_Density_Metrics"


def massfilled_path(src: Path) -> Path:
    return src.with_name(f"{src.stem}_massfilled{src.suffix}")


def backfill_spectral_mass(src: Path, dest: Path | None = None) -> Path:
    src = Path(src)
    if not src.is_file():
        raise FileNotFoundError(f"workbook not found: {src}")
    dest = Path(dest) if dest is not None else massfilled_path(src)
    if dest.exists():
        raise FileExistsError(f"refusing to overwrite existing file: {dest}")
    frame = pd.read_excel(src, sheet_name=SHEET)
    missing = [
        col
        for col in REQUIRED_D0_COLUMNS + REQUIRED_D1_COLUMNS + REQUIRED_R_COLUMNS
        if col not in frame.columns
    ]
    if missing:
        raise ValueError(
            "F-061 v2 backfill requires per-compartment ACD columns "
            f"{list(REQUIRED_D1_COLUMNS)}; missing: {missing}. "
            "Re-export Stage 3 (or re-analyse from audio) so "
            "ACD_D1_harmonic / ACD_D1_inharmonic / ACD_D1_subbass are present. "
            "Refusing to apply the v1 pooled-count formula."
        )
    frame = add_spectral_mass_column(frame)
    frame = place_spectral_mass_right_of_ewsd(frame)

    wb = load_workbook(src)
    idx = wb.sheetnames.index(SHEET)
    del wb[SHEET]
    ws = wb.create_sheet(SHEET, idx)
    for row in dataframe_to_rows(frame, index=False, header=True):
        ws.append(row)
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    apply_spectral_mass_data_bar(ws, headers)
    wb.save(dest)
    return dest


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", help="existing research workbook (.xlsx)")
    args = parser.parse_args()
    dest = backfill_spectral_mass(Path(args.workbook))
    print(f"wrote {dest}")


if __name__ == "__main__":
    main()
