"""Compare two research workbooks for numerical invisibility.

Every numeric cell on a shared sheet/column must match (NaN == NaN).
New sheets, new columns, and non-volatile text cells must be explicitly
whitelisted. Provenance / clock fields are reported but do not fail the
diff (they change when HEAD or export time changes).
"""
from __future__ import annotations

import argparse
import math
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

VOLATILE_COLUMNS = frozenset(
    {
        "code_commit",
        "git_commit",
        "git_status_reason",
        "code_dirty",
        "analysis_version",
        "generated",
        "generated_utc",
        "export_generated",
        "run_id",
    }
)
VOLATILE_HEADER_FRAGMENTS = ("timestamp", "generated_at", "git_describe")
_ISO_STAMP = re.compile(r"\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}")
_GIT_IDENT = re.compile(r"(^|-)g?[0-9a-f]{7,40}(-dirty)?$", re.I)


@dataclass
class DiffReport:
    numeric_mismatches: list[str] = field(default_factory=list)
    extra_sheets: list[str] = field(default_factory=list)
    missing_sheets: list[str] = field(default_factory=list)
    extra_columns: list[str] = field(default_factory=list)
    missing_columns: list[str] = field(default_factory=list)
    text_mismatches: list[str] = field(default_factory=list)
    volatile_notes: list[str] = field(default_factory=list)

    def ok(self, *, allowed_sheets: Iterable[str], allowed_columns: Iterable[str]) -> bool:
        allow_s = set(allowed_sheets)
        allow_c = set(allowed_columns)
        extra_s = [s for s in self.extra_sheets if s not in allow_s]
        extra_c = [c for c in self.extra_columns if _column_key(c) not in allow_c]
        text = [t for t in self.text_mismatches if not _text_whitelisted(t, allow_s, allow_c)]
        return (
            not self.numeric_mismatches
            and not extra_s
            and not self.missing_sheets
            and not extra_c
            and not self.missing_columns
            and not text
        )

    def summary(self) -> str:
        lines = [
            f"numeric_mismatches: {len(self.numeric_mismatches)}",
            f"extra_sheets: {self.extra_sheets or '[]'}",
            f"missing_sheets: {self.missing_sheets or '[]'}",
            f"extra_columns: {self.extra_columns or '[]'}",
            f"missing_columns: {self.missing_columns or '[]'}",
            f"text_mismatches: {len(self.text_mismatches)}",
            f"volatile_notes: {len(self.volatile_notes)}",
        ]
        for block, rows in (
            ("NUMERIC", self.numeric_mismatches[:20]),
            ("TEXT", self.text_mismatches[:20]),
            ("VOLATILE", self.volatile_notes[:10]),
        ):
            if rows:
                lines.append(f"-- {block} --")
                lines.extend(rows)
        return "\n".join(lines)


def _column_key(item: str) -> str:
    if ":" in item:
        return item.split(":", 1)[1]
    return item


def _text_whitelisted(item: str, allow_s: set[str], allow_c: set[str]) -> bool:
    if any(item.startswith(f"{s}!") for s in allow_s):
        return True
    for col in allow_c:
        if f"!{col}!" in f"!{item}!" or item.endswith(f":{col}") or f" {col} " in item:
            return True
    return False


def _looks_like_identity(va, vb) -> bool:
    for val in (va, vb):
        if not isinstance(val, str):
            continue
        s = val.strip()
        if _ISO_STAMP.search(s) or _GIT_IDENT.search(s) or "dirty" in s.lower():
            return True
    return False


def _is_volatile(header: str | None) -> bool:
    if not header:
        return False
    name = str(header).strip()
    if name in VOLATILE_COLUMNS:
        return True
    low = name.lower()
    return any(frag in low for frag in VOLATILE_HEADER_FRAGMENTS)


def _as_number(value):
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return float("nan")
        return float(value)
    if isinstance(value, str):
        # Do not coerce strings: git short hashes can be all-digit (e.g. 0968082).
        return None
    return None


def diff_workbooks(
    baseline: Path,
    candidate: Path,
    *,
    allowed_sheets: Iterable[str] = (),
    allowed_columns: Iterable[str] = (),
) -> DiffReport:
    report = DiffReport()
    wb_a = load_workbook(baseline, data_only=True)
    wb_b = load_workbook(candidate, data_only=True)
    names_a = set(wb_a.sheetnames)
    names_b = set(wb_b.sheetnames)
    report.missing_sheets = sorted(names_a - names_b)
    report.extra_sheets = sorted(names_b - names_a)

    for sheet in wb_a.sheetnames:
        if sheet not in names_b:
            continue
        ws_a = wb_a[sheet]
        ws_b = wb_b[sheet]
        headers_a = [ws_a.cell(1, c).value for c in range(1, ws_a.max_column + 1)]
        headers_b = [ws_b.cell(1, c).value for c in range(1, ws_b.max_column + 1)]
        if sheet == "README" or (ws_a.max_column == 1 and ws_b.max_column == 1):
            _diff_single_column_sheet(sheet, ws_a, ws_b, report)
            continue
        set_a = {h for h in headers_a if h}
        set_b = {h for h in headers_b if h}
        for h in sorted(set_a - set_b):
            report.missing_columns.append(f"{sheet}:{h}")
        for h in sorted(set_b - set_a):
            report.extra_columns.append(f"{sheet}:{h}")
        shared = [h for h in headers_a if h in set_b]
        col_a = {h: i + 1 for i, h in enumerate(headers_a)}
        col_b = {h: i + 1 for i, h in enumerate(headers_b)}
        max_row = max(ws_a.max_row, ws_b.max_row)
        for header in shared:
            ca, cb = col_a[header], col_b[header]
            volatile = _is_volatile(str(header))
            for r in range(2, max_row + 1):
                va = ws_a.cell(r, ca).value if r <= ws_a.max_row else None
                vb = ws_b.cell(r, cb).value if r <= ws_b.max_row else None
                na, nb = _as_number(va), _as_number(vb)
                loc = f"{sheet}!{header}!R{r}"
                if na is not None or nb is not None:
                    if na is None or nb is None:
                        if volatile:
                            report.volatile_notes.append(f"{loc}: {va!r} -> {vb!r}")
                        else:
                            report.numeric_mismatches.append(f"{loc}: {va!r} -> {vb!r}")
                        continue
                    if math.isnan(na) and math.isnan(nb):
                        continue
                    if na != nb:
                        if volatile:
                            report.volatile_notes.append(f"{loc}: {na} -> {nb}")
                        else:
                            report.numeric_mismatches.append(f"{loc}: {na} -> {nb}")
                    continue
                if va != vb:
                    if volatile or _looks_like_identity(va, vb):
                        report.volatile_notes.append(f"{loc}: {va!r} -> {vb!r}")
                    else:
                        report.text_mismatches.append(f"{loc}: {va!r} -> {vb!r}")
    return report


def _diff_single_column_sheet(sheet, ws_a, ws_b, report: DiffReport) -> None:
    max_row = max(ws_a.max_row, ws_b.max_row)
    for r in range(1, max_row + 1):
        va = ws_a.cell(r, 1).value if r <= ws_a.max_row else None
        vb = ws_b.cell(r, 1).value if r <= ws_b.max_row else None
        if va == vb:
            continue
        loc = f"{sheet}!A{r}"
        na, nb = _as_number(va), _as_number(vb)
        if na is not None and nb is not None:
            if math.isnan(na) and math.isnan(nb):
                continue
            if na != nb:
                report.numeric_mismatches.append(f"{loc}: {na} -> {nb}")
            continue
        text = f"{va!r} -> {vb!r}"
        low = f"{va} {vb}".lower()
        if any(tok in low for tok in ("commit", "generated", "dirty", "version")) or (
            isinstance(va, str)
            and isinstance(vb, str)
            and _ISO_STAMP.search(va)
            and _ISO_STAMP.search(vb)
        ):
            report.volatile_notes.append(f"{loc}: {text}")
        else:
            report.text_mismatches.append(f"{loc}: {text}")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("baseline")
    parser.add_argument("candidate")
    parser.add_argument("--allow-sheet", action="append", default=[])
    parser.add_argument("--allow-column", action="append", default=[])
    args = parser.parse_args()
    report = diff_workbooks(
        Path(args.baseline),
        Path(args.candidate),
        allowed_sheets=args.allow_sheet,
        allowed_columns=args.allow_column,
    )
    print(report.summary())
    if report.ok(allowed_sheets=args.allow_sheet, allowed_columns=args.allow_column):
        print("DIFF_OK")
        return 0
    print("DIFF_FAIL")
    return 1


if __name__ == "__main__":
    sys.exit(main())
