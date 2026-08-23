"""F-061 v2 spectral mass — derived Stage 3 column (no signal processing).

v1 pooled D0 across compartments, so inharmonic and sub-bass entities
entered the count at entity weight (30% of counted entities on the cello
corpus against ~2% of energy; I+S contribution to the count up to 15%).
Author requirement: sub-bass and inharmonic content must never count
more than its real (energy) weight.

v2 (compartment-proportional count)::

    count_k = sqrt(D0_k * D1_k)          per compartment k in {H, I, S}
    count   = sum_k r_k * count_k        r_k = E_k / sum(E)
    lambda  = E_total / count
    spectral_mass = count * lambda ** 0.15

Within-compartment dominance protection is retained (the F#4 property);
across compartments each pile is capped at its energy share.
MASS_COUNT_BLEND applies within compartment; MASS_LEVEL_EXPONENT is
unchanged.

Inputs are existing ACD per-compartment exports. NaN in any required
input, or ACD_status != "ok", yields NaN (never 0.0).
"""
from __future__ import annotations

from typing import Mapping, Sequence

import numpy as np
import pandas as pd

MASS_COUNT_BLEND: float = 0.5
MASS_LEVEL_EXPONENT: float = 0.15

SPECTRAL_MASS_FORMULA_ID: str = "F-061"
SPECTRAL_MASS_FORMULA_VERSION: str = "2.0"

SPECTRAL_MASS_COLUMN: str = "spectral_mass"
SPECTRAL_MASS_COUNT_COLUMN: str = "spectral_mass_count"
SPECTRAL_MASS_COUNT_BLEND_COLUMN: str = "spectral_mass_count_blend"
SPECTRAL_MASS_LEVEL_EXPONENT_COLUMN: str = "spectral_mass_level_exponent"
SPECTRAL_MASS_FORMULA_ID_COLUMN: str = "spectral_mass_formula_id"
SPECTRAL_MASS_FORMULA_VERSION_COLUMN: str = "spectral_mass_formula_version"

EWSD_ACOUSTIC_BALANCED_COLUMN: str = "EWSD_score_acoustic_balanced"

SPECTRAL_MASS_DATA_BAR_COLOR: str = "FF4472C4"

COMPARTMENT_KEYS: tuple[str, ...] = ("harmonic", "inharmonic", "subbass")
REQUIRED_D1_COLUMNS: tuple[str, ...] = (
    "ACD_D1_harmonic",
    "ACD_D1_inharmonic",
    "ACD_D1_subbass",
)
REQUIRED_D0_COLUMNS: tuple[str, ...] = (
    "ACD_D0_harmonic",
    "ACD_D0_inharmonic",
    "ACD_D0_subbass",
)
REQUIRED_R_COLUMNS: tuple[str, ...] = (
    "ACD_r_harmonic",
    "ACD_r_inharmonic",
    "ACD_r_subbass",
)

_PLACEMENT_COLUMNS: Sequence[str] = (
    SPECTRAL_MASS_COLUMN,
    SPECTRAL_MASS_COUNT_COLUMN,
    SPECTRAL_MASS_COUNT_BLEND_COLUMN,
    SPECTRAL_MASS_LEVEL_EXPONENT_COLUMN,
    SPECTRAL_MASS_FORMULA_ID_COLUMN,
    SPECTRAL_MASS_FORMULA_VERSION_COLUMN,
)


def _finite_positive(value: float) -> bool:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return False
    return bool(np.isfinite(number) and number > 0.0)


def compartment_count(d0: float, d1: float) -> float:
    """Within-compartment presence/share blend: (D0*D1)**MASS_COUNT_BLEND."""
    if not _finite_positive(d0) or not _finite_positive(d1):
        return float("nan")
    return float((float(d0) * float(d1)) ** MASS_COUNT_BLEND)


def compute_spectral_mass_v1(
    d0: float,
    d1: float,
    lam: float,
    *,
    status: str = "ok",
) -> tuple[float, float]:
    """F-061 v1 pooled count. Kept for v1↔v2 comparison only."""
    if str(status).strip() != "ok":
        return float("nan"), float("nan")
    count = compartment_count(d0, d1)
    if not _finite_positive(count) or not _finite_positive(lam):
        return float("nan"), float("nan")
    return float(count * (float(lam) ** MASS_LEVEL_EXPONENT)), count


def compute_spectral_mass(
    d0: Mapping[str, float],
    d1: Mapping[str, float],
    r: Mapping[str, float],
    e_total: float,
    *,
    status: str = "ok",
    lam: float | None = None,
) -> tuple[float, float]:
    """F-061 v2: return ``(spectral_mass, spectral_mass_count)`` or NaNs.

    ``lam`` is derived as ``E_total / count`` unless an explicit value is
    passed (extensivity tests that hold level fixed).
    """
    if str(status).strip() != "ok":
        return float("nan"), float("nan")
    count = 0.0
    any_term = False
    for key in COMPARTMENT_KEYS:
        rk = r.get(key, float("nan"))
        try:
            rk_f = float(rk)
        except (TypeError, ValueError):
            return float("nan"), float("nan")
        if not np.isfinite(rk_f):
            return float("nan"), float("nan")
        if rk_f == 0.0:
            continue
        if rk_f < 0.0:
            return float("nan"), float("nan")
        ck = compartment_count(d0.get(key, float("nan")), d1.get(key, float("nan")))
        if not np.isfinite(ck):
            return float("nan"), float("nan")
        count += rk_f * ck
        any_term = True
    if not any_term or not _finite_positive(count):
        return float("nan"), float("nan")
    if lam is None:
        if not _finite_positive(e_total):
            return float("nan"), float("nan")
        lam_f = float(e_total) / count
    else:
        try:
            lam_f = float(lam)
        except (TypeError, ValueError):
            return float("nan"), float("nan")
        if not _finite_positive(lam_f):
            return float("nan"), float("nan")
    return float(count * (lam_f ** MASS_LEVEL_EXPONENT)), float(count)


def compartment_count_contribution(
    d0: Mapping[str, float],
    d1: Mapping[str, float],
    r: Mapping[str, float],
    key: str,
) -> float:
    """Return ``r_k * count_k`` for one compartment, or NaN."""
    try:
        rk = float(r.get(key, float("nan")))
    except (TypeError, ValueError):
        return float("nan")
    if not np.isfinite(rk):
        return float("nan")
    if rk == 0.0:
        return 0.0
    ck = compartment_count(d0.get(key, float("nan")), d1.get(key, float("nan")))
    if not np.isfinite(ck):
        return float("nan")
    return float(rk * ck)


def place_spectral_mass_right_of_ewsd(df: pd.DataFrame) -> pd.DataFrame:
    """``spectral_mass`` then ``spectral_mass_count`` sit immediately right of EWSD."""
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return df
    if EWSD_ACOUSTIC_BALANCED_COLUMN not in df.columns:
        return df
    wanted = [c for c in _PLACEMENT_COLUMNS if c in df.columns]
    if not wanted:
        return df
    remaining = [c for c in df.columns if c not in wanted]
    idx = remaining.index(EWSD_ACOUSTIC_BALANCED_COLUMN)
    ordered = remaining[: idx + 1] + wanted + remaining[idx + 1 :]
    return df.loc[:, ordered]


def _row_e_total(row: pd.Series) -> float:
    if "ACD_energy_total" in row.index:
        try:
            energy = float(row["ACD_energy_total"])
        except (TypeError, ValueError):
            energy = float("nan")
        if np.isfinite(energy) and energy > 0.0:
            return energy
    try:
        score = float(row.get("ACD_score", float("nan")))
        lam = float(row.get("ACD_magnitude_per_component", float("nan")))
    except (TypeError, ValueError):
        return float("nan")
    if np.isfinite(score) and np.isfinite(lam) and score > 0.0 and lam > 0.0:
        return float(score * lam)
    return float("nan")


def add_spectral_mass_column(df: pd.DataFrame) -> pd.DataFrame:
    """F-061 v2: compartment-weighted count times bounded level."""
    out = df.copy()
    mass = pd.Series(np.nan, index=out.index, dtype=float)
    count = pd.Series(np.nan, index=out.index, dtype=float)
    if "ACD_status" in out.columns:
        status_ok = out["ACD_status"].map(
            lambda v: str(v).strip() == "ok" if pd.notna(v) else False
        )
    else:
        status_ok = pd.Series(False, index=out.index)

    have_inputs = all(
        col in out.columns
        for col in REQUIRED_D0_COLUMNS + REQUIRED_D1_COLUMNS + REQUIRED_R_COLUMNS
    )
    if have_inputs:
        for idx in out.index:
            if not bool(status_ok.loc[idx]):
                continue
            row = out.loc[idx]
            d0 = {k: row[f"ACD_D0_{k}"] for k in COMPARTMENT_KEYS}
            d1 = {k: row[f"ACD_D1_{k}"] for k in COMPARTMENT_KEYS}
            r = {k: row[f"ACD_r_{k}"] for k in COMPARTMENT_KEYS}
            mass_i, count_i = compute_spectral_mass(
                d0, d1, r, _row_e_total(row), status="ok"
            )
            mass.loc[idx] = mass_i
            count.loc[idx] = count_i

    out[SPECTRAL_MASS_COLUMN] = mass
    out[SPECTRAL_MASS_COUNT_COLUMN] = count
    out[SPECTRAL_MASS_COUNT_BLEND_COLUMN] = MASS_COUNT_BLEND
    out[SPECTRAL_MASS_LEVEL_EXPONENT_COLUMN] = MASS_LEVEL_EXPONENT
    out[SPECTRAL_MASS_FORMULA_ID_COLUMN] = SPECTRAL_MASS_FORMULA_ID
    out[SPECTRAL_MASS_FORMULA_VERSION_COLUMN] = SPECTRAL_MASS_FORMULA_VERSION
    return place_spectral_mass_right_of_ewsd(out)


def spectral_mass_data_bar_rule():
    """Same DataBarRule shape as the red EWSD bars; blue fill."""
    from openpyxl.formatting.rule import DataBarRule

    return DataBarRule(
        start_type="min",
        end_type="max",
        color=SPECTRAL_MASS_DATA_BAR_COLOR,
        showValue=True,
        minLength=0,
        maxLength=100,
    )


def apply_spectral_mass_data_bar(ws, headers=None) -> None:
    """Attach the blue data-bar rule to ``spectral_mass`` on ``ws``."""
    from openpyxl.utils import get_column_letter

    if headers is None:
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    last = ws.max_row
    if last < 2:
        return
    try:
        ci = list(headers).index(SPECTRAL_MASS_COLUMN) + 1
    except ValueError:
        return
    letter = get_column_letter(ci)
    ws.conditional_formatting.add(f"{letter}2:{letter}{last}", spectral_mass_data_bar_rule())
