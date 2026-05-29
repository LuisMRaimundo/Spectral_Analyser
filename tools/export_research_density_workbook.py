#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Post-process ``compiled_density_metrics.xlsx`` into a reduced research workbook.

Does **not** alter compilation logic or the source workbook; read-only input.

**Output Excel behaviour:** the writer uses **worksheet-level AutoFilter** and
**frozen header rows** on data sheets only. It does **not** create formal Excel
**Table** / ListObject parts (no ``xl/tables/table*.xml``), so Microsoft Excel
should open the file without repair prompts. **README** and **Dashboard** sheets
are not auto-filtered. DataFrame column names are sanitised before export (no
blank headers; duplicate names receive ``_2``, ``_3``, …).

**Metadata CLI:** optional ``--instrument``, ``--dynamic``, and
``--force-metadata`` (see ``README.md`` and the research workbook README sheet).

**Spectral_Density_Metrics:** merges ``Legacy_Compatibility`` from the compiled
workbook; adds ``density_weighted_sum_cdm_mean`` =
(``density_weighted_sum`` + ``Combined Density Metric``) / 2; applies soft column
highlights on those three columns (research file only). Normative semantics:
``docs/DENSITY_EXPORT_SCHEMA.md`` §R.
"""

from __future__ import annotations

import argparse
import hashlib
import os
import re
import subprocess
import sys
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Mapping, Optional, Sequence, Tuple

_REPO_ROOT = Path(__file__).resolve().parents[1]
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from metadata_sanitizer import (
    apply_publication_clean_research_metadata_fields,
    drop_publication_noise_columns_from_dataframe,
    format_utc_publication_timestamp,
    publication_clean_drop_known_sparse_columns,
    publication_clean_export_enabled,
    publication_research_canonical_density_columns,
)
from constants import BODY_DENSITY_MAX_HZ, FULL_SPECTRUM_MAX_HZ

SCRIPT_NAME = "export_research_density_workbook.py"
SCRIPT_VERSION = "1.1.2"
TIER_STRATEGY_LABEL = "90_tier_granular"
TIER_DEPENDENT_LABEL = "tier_dependent_see_Analysis_Settings_By_Note"
UNKNOWN_NOT_PARSEABLE = "unknown_not_parseable"
FIXED_FFT_MODE_LABEL = "fixed_fft_mode"
FREQ_MAG_RECOVERY_PARTIAL = "partially_unavailable_in_compiled_source"


@dataclass
class ResearchExportMetadata:
    """Optional CLI overrides for Instrument / Dynamic columns."""

    instrument: Optional[str] = None
    dynamic: Optional[str] = None
    force_metadata: bool = False


CHART_AMPLITUDE_NAME = "component_amplitude_mass_pie.png"
CHART_ENERGY_RATIO_NAME = "component_energy_ratio_pie.png"
CHART_ENERGY_LEGACY_NAME = "component_energy_pie.png"

# --- Sheet merge priority (first wins for non-null when coalescing) ---
MERGE_SHEETS: Tuple[str, ...] = (
    "Density_Metrics",
    "Canonical_Metrics",
    "Diagnostic_Metrics",
    "Legacy_Compatibility",
    "Validation_Metrics",
    "Debug_Counts",
    "Per_Note_Processing_Metadata",
)

# Canonical output column -> ordered list of source aliases (first match wins per sheet).
COLUMN_ALIASES: Dict[str, Tuple[str, ...]] = {
    "harmonic_energy_ratio": ("harmonic_energy_ratio", "component_harmonic_energy_ratio"),
    "inharmonic_energy_ratio": ("inharmonic_energy_ratio", "component_inharmonic_energy_ratio"),
    "subbass_energy_ratio": ("subbass_energy_ratio", "component_subbass_energy_ratio"),
    "harmonic_energy_sum": ("harmonic_energy_sum", "component_harmonic_energy_sum"),
    "inharmonic_energy_sum": ("inharmonic_energy_sum", "component_inharmonic_energy_sum"),
    "subbass_energy_sum": ("subbass_energy_sum", "component_subbass_energy_sum"),
    "harmonic_density_sum": (
        "harmonic_density_sum",
        "Harmonic Partials sum",
        "linear_sum_amplitude_harmonic",
    ),
    "inharmonic_density_sum": (
        "inharmonic_density_sum",
        "Inharmonic Partials sum",
        "linear_sum_amplitude_inharmonic_partial",
    ),
    "subbass_density_sum": (
        "subbass_density_sum",
        "Sub-bass sum",
        "linear_sum_amplitude_subbass_band",
    ),
    "Source_File": (
        "Source_File",
        "source_file_name",
        "filename",
        "file_name",
        "source_file",
        "input_file",
    ),
    "Source_Workbook": ("Source_Workbook", "source_workbook", "workbook_path", "compiled_from"),
    "f0_final_hz": ("f0_final_hz", "f0_estimated", "f0_final"),
    "harmonic_region_occupancy_count": (
        "harmonic_region_occupancy_count",
        "harmonic_occupancy_detected_order_count",
    ),
    "Instrument": ("Instrument", "instrument", "instrument_detected"),
    "Dynamic": ("Dynamic", "dynamic", "dynamic_detected"),
}

# Do not add dangerous reverse aliases (user forbade mapping Total sum -> density_metric_raw, etc.)

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FILL = PatternFill("solid", fgColor="1F4E79")
TITLE_FONT = Font(bold=True, color="FFFFFF", size=16)
SUBHEADER_FONT = Font(bold=True, size=12)
KPI_LABEL_FONT = Font(bold=True, size=10)
THIN = Side(style="thin", color="B4B4B4")
KPI_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
NEUTRAL_KPI_FILL = PatternFill("solid", fgColor="E7E6E6")
MISSING_WARN_FILL = PatternFill("solid", fgColor="F8CBAD")
F0_FALSE_FILL = PatternFill("solid", fgColor="FFEB9C")

# Research ``Spectral_Density_Metrics`` column highlights (soft fills; header + data).
RESEARCH_FILL_DENSITY_WEIGHTED_SUM = PatternFill("solid", fgColor="D6E4F0")
RESEARCH_FILL_COMBINED_DENSITY_METRIC = PatternFill("solid", fgColor="FFF2CC")
RESEARCH_FILL_DWS_CDM_MEAN = PatternFill("solid", fgColor="E8D5F2")
# Light blue highlight for the principled per-note scalar note_density_final.
RESEARCH_FILL_NOTE_DENSITY_FINAL = PatternFill("solid", fgColor="ADD8E6")
RESEARCH_HIGHLIGHT_HEADER_FONT = Font(bold=True, color="1F4E79", size=11)


PITCH_CLASS_MAP: Dict[str, int] = {
    "C": 0,
    "C#": 1,
    "DB": 1,
    "D": 2,
    "D#": 3,
    "EB": 3,
    "E": 4,
    "F": 5,
    "F#": 6,
    "GB": 6,
    "G": 7,
    "G#": 8,
    "AB": 8,
    "A": 9,
    "A#": 10,
    "BB": 10,
    "B": 11,
}


def parse_note(note_string: str | None) -> Tuple[str | None, int | None, int | None, str | None]:
    """
    Parse scientific pitch notation into (token, pitch_class, octave, letter).

    Returns (None, None, None, None) if parsing fails.
    """
    if note_string is None or (isinstance(note_string, float) and np.isnan(note_string)):
        return None, None, None, None
    s = str(note_string).strip()
    if not s:
        return None, None, None, None
    m = re.match(r"^([A-Ga-g])([#bB]?)(\d+)$", s)
    if not m:
        return None, None, None, None
    letter = m.group(1).upper()
    acc = m.group(2)
    if acc == "b" or acc == "B":
        acc_token = "b"
    elif acc == "#":
        acc_token = "#"
    else:
        acc_token = ""
    token_key = letter + acc_token
    if acc_token == "b":
        token_key = letter + "b"
    elif acc_token == "#":
        token_key = letter + "#"
    else:
        token_key = letter
    pc_key = token_key.replace("b", "B") if "b" in token_key else token_key
    pc_key = pc_key.upper()
    if pc_key not in PITCH_CLASS_MAP:
        return None, None, None, None
    pc = PITCH_CLASS_MAP[pc_key]
    octave = int(m.group(3))
    display = letter + (acc if acc in ("#", "b", "B") else "") + str(octave)
    if acc == "B":
        display = letter + "b" + str(octave)
    return display, pc, octave, letter


def note_to_midi(note_string: str | None) -> float | None:
    _tok, pc, octave, _ = parse_note(note_string)
    if pc is None or octave is None:
        return None
    return float(12 * (octave + 1) + pc)


def register_from_midi(midi: float | None) -> str | None:
    if midi is None or (isinstance(midi, float) and np.isnan(midi)):
        return None
    m = int(round(midi))
    if m < 36:
        return "Very low"
    if m < 48:
        return "Low"
    if m < 60:
        return "Middle"
    if m < 72:
        return "High"
    return "Very high"


def pitch_class_name(note_string: str | None) -> str | None:
    _t, pc, octave, letter = parse_note(note_string)
    if pc is None:
        return None
    m = re.match(r"^([A-Ga-g])([#b]?)", str(note_string).strip())
    if not m:
        return None
    base = m.group(1).upper() + (m.group(2) if m.group(2) in ("#", "b") else "")
    return base


def find_note_column(df: pd.DataFrame) -> str | None:
    for c in df.columns:
        if str(c).strip().lower() == "note":
            return str(c)
    return None


def _first_matching_column(df: pd.DataFrame, names: Sequence[str]) -> str | None:
    colset = {str(c): str(c) for c in df.columns}
    lower_map = {str(c).lower(): str(c) for c in df.columns}
    for name in names:
        if name in colset:
            return name
        low = name.lower()
        if low in lower_map:
            return lower_map[low]
    return None


def _rename_frame_to_canonical(df: pd.DataFrame) -> pd.DataFrame:
    """Rename known aliases to canonical names when the canonical column is absent."""
    out = df.copy()
    renames: Dict[str, str] = {}
    for col in list(out.columns):
        sc = str(col)
        if sc in renames:
            continue
        for canon, aliases in COLUMN_ALIASES.items():
            if sc == canon:
                break
            if sc in aliases or sc.lower() in {a.lower() for a in aliases}:
                if canon not in out.columns:
                    renames[col] = canon
                break
    return out.rename(columns=renames) if renames else out


def _coalesce_series(a: pd.Series, b: pd.Series) -> pd.Series:
    return a.where(a.notna() & (a.astype(str) != ""), b)


def merge_workbook_frames(path: Path, warnings: List[str]) -> pd.DataFrame:
    try:
        xl = pd.ExcelFile(path, engine="openpyxl")
    except Exception as e:  # noqa: BLE001
        raise ValueError(f"Cannot read workbook: {path}: {e}") from e
    names = set(xl.sheet_names)
    if "Density_Metrics" not in names:
        raise ValueError("Required sheet 'Density_Metrics' is missing from the compiled workbook.")

    merged: Optional[pd.DataFrame] = None
    note_key = "Note"
    density_raw = pd.read_excel(path, sheet_name="Density_Metrics", engine="openpyxl")
    if density_raw.empty:
        raise ValueError("Required sheet 'Density_Metrics' is empty.")
    density_nc = find_note_column(density_raw)
    if density_nc is None:
        raise ValueError("Required sheet 'Density_Metrics' has no 'Note' column.")
    merged = _rename_frame_to_canonical(density_raw.rename(columns={density_nc: note_key}))
    merged = merged.reset_index(drop=True).copy()
    merged["__source_density_row_id"] = np.arange(len(merged), dtype=int)

    for sheet in MERGE_SHEETS:
        if sheet == "Density_Metrics":
            continue
        if sheet not in names:
            warnings.append(f"Optional sheet '{sheet}' not found; skipped.")
            continue
        raw = pd.read_excel(path, sheet_name=sheet, engine="openpyxl")
        if raw.empty:
            warnings.append(f"Sheet '{sheet}' is empty; skipped.")
            continue
        nc = find_note_column(raw)
        if nc is None:
            warnings.append(f"Sheet '{sheet}' has no 'Note' column; skipped.")
            continue
        frame = raw.rename(columns={nc: note_key})
        frame = _rename_frame_to_canonical(frame)
        if frame.duplicated(subset=[note_key]).any():
            # Prevent cartesian expansion in Spectral_Density_Metrics.
            frame = frame.groupby(note_key, as_index=False, sort=False).last()
        # Left merge anchored to Density_Metrics row cardinality.
        merged = merged.merge(frame, on=note_key, how="left", suffixes=("", "_y"))
        drop_y: List[str] = []
        for col in list(merged.columns):
            if not col.endswith("_y"):
                continue
            base = col[:-2]
            if base == note_key:
                drop_y.append(col)
                continue
            if base in merged.columns:
                merged[base] = _coalesce_series(merged[base], merged[col])
            else:
                merged[base] = merged[col]
            drop_y.append(col)
        merged = merged.drop(columns=drop_y, errors="ignore")
        # Drop duplicate Note columns if any
        dup_note = [c for c in merged.columns if str(c).lower() == "note" and c != note_key]
        merged = merged.drop(columns=dup_note, errors="ignore")

    if merged is None or merged.empty:
        raise ValueError("No per-note rows could be loaded from Density_Metrics.")

    return merged


def _augment_source_file_name_from_per_note_processing_metadata(
    merged: pd.DataFrame,
    compiled_workbook: Path,
) -> pd.DataFrame:
    """Populate source-file hints from Per_Note_Processing_Metadata when available.

    Some real runs carry reliable per-note filename provenance only in
    ``Per_Note_Processing_Metadata.source_file_name``. This helper backfills
    ``source_file_name`` in the merged frame by Note key without changing
    analysis semantics.
    """
    if merged is None or merged.empty or "Note" not in merged.columns:
        return merged
    out = merged.copy()
    if "source_file_name" not in out.columns:
        out["source_file_name"] = np.nan
    try:
        pnp = pd.read_excel(
            compiled_workbook,
            sheet_name="Per_Note_Processing_Metadata",
            engine="openpyxl",
        )
    except Exception:
        return out
    if pnp is None or pnp.empty:
        return out
    note_col = find_note_column(pnp)
    if note_col is None:
        return out
    src_col = _first_matching_column(
        pnp,
        ("source_file_name", "Source_File", "source_file", "filename", "file_name"),
    )
    if src_col is None:
        return out
    pnp_map = (
        pnp[[note_col, src_col]]
        .copy()
        .rename(columns={note_col: "Note", src_col: "source_file_name"})
    )
    pnp_map["Note"] = pnp_map["Note"].astype(str)
    pnp_map["source_file_name"] = pnp_map["source_file_name"].astype(str)
    pnp_map = pnp_map[
        pnp_map["Note"].str.strip().ne("")
        & pnp_map["source_file_name"].str.strip().ne("")
        & pnp_map["source_file_name"].str.lower().ne("nan")
    ]
    if pnp_map.empty:
        return out
    note_to_src = dict(zip(pnp_map["Note"], pnp_map["source_file_name"], strict=False))
    cur = out["source_file_name"]
    out["source_file_name"] = [
        cur.iloc[i]
        if (
            pd.notna(cur.iloc[i])
            and str(cur.iloc[i]).strip() != ""
            and str(cur.iloc[i]).strip().lower() != "nan"
        )
        else note_to_src.get(str(out.iloc[i]["Note"]), np.nan)
        for i in range(len(out))
    ]
    return out


def _series_or_nan(df: pd.DataFrame, name: str) -> pd.Series:
    if name in df.columns:
        return pd.to_numeric(df[name], errors="coerce")
    return pd.Series(np.nan, index=df.index)


def _series_str(df: pd.DataFrame, name: str) -> pd.Series:
    if name in df.columns:
        return df[name].astype(str).replace({"nan": np.nan, "None": np.nan})
    return pd.Series(np.nan, index=df.index)


def min_max_normalize(s: pd.Series) -> Tuple[pd.Series, bool]:
    """Returns (normalized, warn_constant)."""
    v = pd.to_numeric(s, errors="coerce")
    vmin = v.min(skipna=True)
    vmax = v.max(skipna=True)
    if pd.isna(vmin) or pd.isna(vmax) or vmax == vmin:
        return pd.Series(np.nan, index=s.index), True
    return (v - vmin) / (vmax - vmin), False


def _pick_series(df: pd.DataFrame, canonical: str) -> pd.Series:
    """First matching column among aliases, else canonical name, else NaN column."""
    if canonical in COLUMN_ALIASES:
        for a in COLUMN_ALIASES[canonical]:
            if a in df.columns:
                return df[a]
    if canonical in df.columns:
        return df[canonical]
    return pd.Series(np.nan, index=df.index)


def _note_to_source_file_lookup(merged: pd.DataFrame) -> Dict[str, Any]:
    """Last row per note for ``Source_File`` (or aliases); used for chart path tie-break without exporting the column."""
    if merged.empty or "Note" not in merged.columns:
        return {}
    src = _pick_series(merged, "Source_File")
    m = merged[["Note"]].copy()
    m["_src"] = src.reindex(m.index).to_numpy()
    last = m.groupby("Note", as_index=False, sort=False).last()
    return {str(r["Note"]).strip(): r["_src"] for _, r in last.iterrows()}


def _all_blank_or_nan(series: pd.Series) -> bool:
    if series.isna().all():
        return True
    t = series.dropna().astype(str).str.strip()
    return len(t) == 0 or t.eq("").all()


def _first_non_blank(series: pd.Series) -> Any:
    if series is None:
        return np.nan
    s = series.dropna()
    if s.empty:
        return np.nan
    if s.dtype == object:
        t = s.astype(str).str.strip()
        t = t[t.ne("")]
        if t.empty:
            return np.nan
        return t.iloc[0]
    return s.iloc[0]


def _as_optional_float(v: Any) -> Optional[float]:
    try:
        if pd.isna(v):
            return None
    except Exception:
        pass
    try:
        return float(v)
    except Exception:
        return None


def _derive_zero_padding_from_fft(merged: pd.DataFrame) -> Any:
    if "zero_padding" in merged.columns:
        v = _first_non_blank(pd.to_numeric(merged["zero_padding"], errors="coerce"))
        if pd.notna(v):
            return int(v)
    if "n_fft_effective" in merged.columns and "n_fft" in merged.columns:
        n_eff = _first_non_blank(pd.to_numeric(merged["n_fft_effective"], errors="coerce"))
        n = _first_non_blank(pd.to_numeric(merged["n_fft"], errors="coerce"))
        if pd.notna(n_eff) and pd.notna(n) and float(n) > 0:
            ratio = float(n_eff) / float(n)
            if ratio >= 1.0:
                return int(round(ratio))
    return UNKNOWN_NOT_PARSEABLE


def _numeric_series_for_note(
    out_note: pd.Series,
    lookup: pd.DataFrame,
    candidates: Sequence[str],
) -> pd.Series:
    for c in candidates:
        if c in lookup.columns:
            mapped = out_note.map(lookup[c].to_dict())
            num = pd.to_numeric(mapped, errors="coerce")
            if num.notna().any():
                return num
    return pd.Series(np.nan, index=out_note.index)


def _meta_numeric(meta: Mapping[str, Any], *keys: str) -> Optional[float]:
    lmeta = {str(k).strip().lower(): v for k, v in meta.items()}
    for k in keys:
        if not k:
            continue
        raw = lmeta.get(str(k).strip().lower(), np.nan)
        f = _as_optional_float(raw)
        if f is not None:
            return f
    return None


def _resolve_freq_mag_field(
    out_note: pd.Series,
    lookup: pd.DataFrame,
    meta: Mapping[str, Any],
    *,
    lookup_candidates: Sequence[str],
    meta_candidates: Sequence[str],
) -> Tuple[pd.Series, Any]:
    per_note = _numeric_series_for_note(out_note, lookup, lookup_candidates)
    if per_note.notna().any():
        uniq = sorted(float(v) for v in pd.unique(per_note.dropna()))
        if len(uniq) == 1:
            global_val: Any = uniq[0]
        else:
            global_val = TIER_DEPENDENT_LABEL
        return per_note, global_val
    mv = _meta_numeric(meta, *meta_candidates)
    if mv is not None:
        return pd.Series(float(mv), index=out_note.index), float(mv)
    return pd.Series(UNKNOWN_NOT_PARSEABLE, index=out_note.index), UNKNOWN_NOT_PARSEABLE


def _detect_tier_strategy_used(meta: Mapping[str, Any], merged: pd.DataFrame) -> bool:
    tier_keys = (
        "tier_strategy",
        "n_fft_strategy_or_tier_strategy",
        "hop_length_strategy_or_tier_strategy",
    )
    for k in tier_keys:
        if k in meta:
            v = str(meta.get(k, "")).strip().lower()
            if "tier" in v and v not in {"", "nan"}:
                return True
    if "tier" in merged.columns:
        tier_vals = (
            merged["tier"]
            .dropna()
            .astype(str)
            .str.strip()
        )
        tier_vals = tier_vals[tier_vals.ne("")]
        if not tier_vals.empty:
            return True
    for c in ("n_fft", "hop_length"):
        if c in merged.columns:
            u = pd.to_numeric(merged[c], errors="coerce").dropna().unique()
            if len(u) > 1:
                return True
    return False


def _derive_source_corpus_path(path: Path, meta: Mapping[str, Any]) -> str:
    v = str(meta.get("source_corpus_path", "") or "").strip()
    if v and "spectral_analysis.xlsx" not in v.lower():
        return v
    compiled_from = str(meta.get("compiled_from", "") or "").strip()
    if compiled_from and "spectral_analysis.xlsx" not in compiled_from.lower():
        return compiled_from
    # compiled workbook usually lives under <corpus>/analysis_results
    if path.parent.name.lower() in {"analysis_results", "analysis_results_final_density_acceptance"}:
        return str(path.parent.parent)
    return str(path.parent)


def build_analysis_settings_by_note(
    merged: pd.DataFrame,
    sd: pd.DataFrame,
    meta: Mapping[str, Any],
) -> pd.DataFrame:
    by_note = merged.groupby("Note", as_index=False, sort=False).last() if "Note" in merged.columns else pd.DataFrame()
    out = sd[["Note"]].copy() if "Note" in sd.columns else pd.DataFrame({"Note": by_note.get("Note", pd.Series(dtype=object))})
    if "MIDI" in sd.columns:
        out["MIDI"] = pd.to_numeric(sd["MIDI"], errors="coerce")
    else:
        out["MIDI"] = pd.to_numeric(out["Note"].map(note_to_midi), errors="coerce")

    for c in ("f0_used_for_density_hz", "f0_used_for_density_source", "acoustic_f0_status"):
        out[c] = sd[c] if c in sd.columns else "unavailable_not_recorded"

    if not by_note.empty:
        lookup = by_note.set_index("Note")
    else:
        lookup = pd.DataFrame(index=out["Note"].astype(str))

    tier_mode = _detect_tier_strategy_used(meta, merged)
    tier_series = lookup["tier"] if "tier" in lookup.columns else pd.Series(np.nan, index=lookup.index)
    out["tier_name"] = out["Note"].map(tier_series.to_dict()) if not tier_series.empty else np.nan
    if tier_mode:
        out["tier_name"] = out["tier_name"].fillna(TIER_STRATEGY_LABEL)
    else:
        out["tier_name"] = out["tier_name"].fillna(FIXED_FFT_MODE_LABEL)

    def _per_note_value(col: str, fallback: Any) -> pd.Series:
        if col in lookup.columns:
            return out["Note"].map(lookup[col].to_dict())
        return pd.Series(fallback, index=out.index)

    out["n_fft"] = _per_note_value("n_fft", TIER_DEPENDENT_LABEL if tier_mode else UNKNOWN_NOT_PARSEABLE)
    out["hop_length"] = _per_note_value("hop_length", TIER_DEPENDENT_LABEL if tier_mode else UNKNOWN_NOT_PARSEABLE)
    out["zero_padding"] = _per_note_value("zero_padding", TIER_DEPENDENT_LABEL if tier_mode else _derive_zero_padding_from_fft(merged))
    out["window_type"] = _per_note_value("window_type", _first_non_blank(lookup["window"]) if "window" in lookup.columns else UNKNOWN_NOT_PARSEABLE)
    out["harmonic_tolerance_hz"] = _per_note_value(
        "harmonic_tolerance",
        TIER_DEPENDENT_LABEL if tier_mode else UNKNOWN_NOT_PARSEABLE,
    )

    freq_min_series, _ = _resolve_freq_mag_field(
        out["Note"],
        lookup,
        meta,
        lookup_candidates=("frequency_min_hz", "freq_min"),
        meta_candidates=("frequency_min_hz", "freq_min"),
    )
    freq_max_series, _ = _resolve_freq_mag_field(
        out["Note"],
        lookup,
        meta,
        lookup_candidates=("frequency_max_hz", "freq_max"),
        meta_candidates=("frequency_max_hz", "freq_max"),
    )
    mag_min_series, _ = _resolve_freq_mag_field(
        out["Note"],
        lookup,
        meta,
        lookup_candidates=("magnitude_min_db", "db_min"),
        meta_candidates=("magnitude_min_db", "db_min"),
    )
    mag_max_series, _ = _resolve_freq_mag_field(
        out["Note"],
        lookup,
        meta,
        lookup_candidates=("magnitude_max_db", "db_max"),
        meta_candidates=("magnitude_max_db", "db_max"),
    )
    out["frequency_min_hz"] = freq_min_series
    out["frequency_max_hz"] = freq_max_series
    out["magnitude_min_db"] = mag_min_series
    out["magnitude_max_db"] = mag_max_series

    for c in (
        "density_summation_mode",
        "harmonic_density_weight",
        "inharmonic_density_weight",
        "subbass_density_weight",
        "density_salience_threshold_db",
        "density_frequency_ceiling_hz",
    ):
        if c in sd.columns:
            out[c] = sd[c]
        elif c in lookup.columns:
            out[c] = out["Note"].map(lookup[c].to_dict())
        else:
            out[c] = meta.get(c, UNKNOWN_NOT_PARSEABLE)

    cols = [
        "Note",
        "MIDI",
        "f0_used_for_density_hz",
        "f0_used_for_density_source",
        "density_component_body_weighted_sum_body_ceiling",
        "harmonic_component_energy_sum_body_ceiling",
        "inharmonic_component_energy_sum_body_ceiling",
        "density_component_body_weighted_sum_body_ceiling",
        "harmonic_component_energy_sum_body_ceiling",
        "inharmonic_component_energy_sum_body_ceiling",
        "subbass_component_energy_sum",
        "spectral_slope_db_per_harmonic",
        "density_body_weighted_sum_body_ceiling",
        "harmonic_body_energy_sum_body_ceiling",
        "inharmonic_body_energy_sum_body_ceiling",
        "subbass_rumble_energy_sum",
        "acoustic_f0_status",
        "tier_name",
        "n_fft",
        "hop_length",
        "zero_padding",
        "window_type",
        "harmonic_tolerance_hz",
        "frequency_min_hz",
        "frequency_max_hz",
        "magnitude_min_db",
        "magnitude_max_db",
        "density_summation_mode",
        "harmonic_density_weight",
        "inharmonic_density_weight",
        "subbass_density_weight",
        "density_salience_threshold_db",
        "density_frequency_ceiling_hz",
        "body_density_frequency_ceiling_hz",
        "full_spectrum_frequency_ceiling_hz",
        "density_full_spectrum_weighted_sum_20khz",
        "harmonic_full_spectrum_energy_sum_20khz",
        "inharmonic_full_spectrum_energy_sum_20khz",
        "high_frequency_spectral_activity_sum",
        "spectral_extension_index_20khz",
        "brightness_or_upper_spectral_activity_index_20khz",
        "full_spectrum_harmonic_candidate_count_20khz",
        "probable_harmonic_component_count_body_ceiling",
        "probable_harmonic_component_energy_sum_body_ceiling",
    ]
    out = out[[c for c in cols if c in out.columns]]
    return out.sort_values("MIDI", na_position="last", kind="mergesort")


# Longest dynamic first (prefer pp over p)
_DYNAMIC_ORDER: Tuple[str, ...] = ("fff", "ppp", "ff", "pp", "mp", "mf", "p", "f")
_DYNAMIC_TOKEN_OK = frozenset(_DYNAMIC_ORDER)


def _ascii_fold(text: str) -> str:
    s = unicodedata.normalize("NFKD", text)
    return "".join(c for c in s if not unicodedata.combining(c))


def _tokenize_metadata_text(text: str | None) -> List[str]:
    if text is None or (isinstance(text, float) and np.isnan(text)):
        return []
    s = _ascii_fold(str(text).strip()).lower()
    if not s:
        return []
    s = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", s)
    parts = re.split(r"[_\s\.\/\\-]+", s)
    return [p for p in parts if len(p) >= 2 or p in _DYNAMIC_TOKEN_OK]


# (canonical name, substring phrases (longest match wins within rule), whole-token abbrev set)
_INSTRUMENT_RULES: Tuple[Tuple[str, Tuple[str, ...], Tuple[str, ...]], ...] = (
    ("English Horn", ("englishhorn", "english horn", "cor anglais", "coranglais", "cornoingles", "corno ingles"), ("eh",)),
    ("Clarinet", ("clarinet", "clarinete", "clarinetto"), ("clar", "cl")),
    ("Bassoon", ("bassoon", "basson", "fagote", "fagott", "fagotto", "fgt"), ("bn",)),
    ("Flute", ("flute", "flauta", "flûte"), ("fl",)),
    ("Oboe", ("oboe", "oboé"), ("ob",)),
    ("Violin", ("violin", "violino"), ("vln",)),
    ("Viola", ("viola",), ("vla",)),
    ("Cello", ("violoncello", "violoncelo", "cello"), ("vc",)),
    ("Contrabass", ("contrabass", "contrabaixo", "double bass", "doublebass"), ("cb",)),
    ("Trumpet", ("trumpet", "trompete"), ("tpt",)),
    ("Horn", ("frenchhorn", "french horn", "trompa"), ("hn",)),
    ("Trombone", ("trombone",), ("tbn",)),
    ("Tuba", ("tuba",), ("tba",)),
)


def infer_instrument_conservative(text: str | None) -> Optional[str]:
    """Infer instrument from path/filename text using token and phrase rules."""
    if not text:
        return None
    folded = _ascii_fold(str(text)).lower()
    compact = re.sub(r"[_\s\.\/\\-]+", "", folded)
    tokens = set(_tokenize_metadata_text(text))
    for canon, phrases, abbrevs in _INSTRUMENT_RULES:
        for ph in sorted(phrases, key=len, reverse=True):
            ph_c = re.sub(r"\s+", "", ph)
            if ph in folded or ph_c in compact:
                return canon
        for ab in abbrevs:
            if ab in tokens:
                return canon
    return None


def infer_dynamic_conservative(text: str | None) -> Optional[str]:
    if not text:
        return None
    folded = _ascii_fold(str(text)).lower()
    tokens = _tokenize_metadata_text(text)
    hits: List[str] = []
    for tok in tokens:
        if tok in _DYNAMIC_ORDER:
            hits.append(tok)
    for dyn in _DYNAMIC_ORDER:
        for sep in ("-", "_", " "):
            if f"{sep}{dyn}{sep}" in f"{sep}{folded}{sep}":
                hits.append(dyn)
                break
    if not hits:
        return None
    hits.sort(key=lambda d: (-len(d), _DYNAMIC_ORDER.index(d) if d in _DYNAMIC_ORDER else 99))
    return hits[0]


def infer_instrument_from_text(text: str | None) -> Optional[str]:
    return infer_instrument_conservative(text)


def infer_dynamic_from_text(text: str | None) -> Optional[str]:
    return infer_dynamic_conservative(text)


_TECHNIQUE_TOKEN_MAP: Dict[str, str] = {
    "ord": "ord",
    "ordinario": "ord",
    "stacc": "stacc",
    "staccato": "stacc",
    "leg": "legato",
    "legato": "legato",
    "pizz": "pizz",
    "pizzicato": "pizz",
    "trem": "tremolo",
    "tremolo": "tremolo",
    "vib": "vibrato",
    "vibrato": "vibrato",
}


def infer_technique_conservative(text: str | None) -> Optional[str]:
    if not text:
        return None
    tokens = _tokenize_metadata_text(text)
    for tok in tokens:
        if tok in _TECHNIQUE_TOKEN_MAP:
            return _TECHNIQUE_TOKEN_MAP[tok]
    return None


def _cell_str(row: pd.Series, col: str) -> Optional[str]:
    if col not in row.index:
        return None
    v = row.get(col)
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return None
    s = str(v).strip()
    return s or None


def _dedupe_preserve(seq: Sequence[str]) -> List[str]:
    seen: set[str] = set()
    out: List[str] = []
    for x in seq:
        if x in seen:
            continue
        seen.add(x)
        out.append(x)
    return out


def _instrument_inference_sources(merged: pd.DataFrame, i: int, compiled_workbook: Path) -> List[str]:
    """
    Priority-ordered strings for conservative instrument inference (first hit wins).

    Skips canonical ``Instrument`` / ``Dynamic`` here: those are handled as explicit workbook
    columns in ``_build_instrument_dynamic_series`` before inference runs.
    """
    row = merged.iloc[i]
    texts: List[str] = []
    hint_cols = (
        "instrument_detected",
        "instrument_family",
        "instrument_guess",
        "instrument_name",
        "instrument_label",
        "selected_instrument",
        "inferred_instrument",
        "per_note_instrument",
    )
    for col in hint_cols:
        s = _cell_str(row, col)
        if s:
            texts.append(s)
    for col in merged.columns:
        lc = str(col).lower()
        if lc in {"instrument", "note"} or col in hint_cols:
            continue
        if "instrument" in lc and "instrumentation" not in lc:
            s = _cell_str(row, str(col))
            if s and s.lower() not in {"unknown", "nan", "none"}:
                texts.append(s)
    for _src_col in (
        "Source_File",
        "source_file_name",
        "source_file",
        "filename",
        "file_name",
        "input_file",
    ):
        _src_series = _pick_series(merged, _src_col)
        if len(_src_series) > i:
            v = _src_series.iloc[i]
            if pd.notna(v) and str(v).strip():
                texts.append(str(v))
    sw = _pick_series(merged, "Source_Workbook")
    if len(sw) > i:
        v = sw.iloc[i]
        if pd.notna(v) and str(v).strip():
            texts.append(str(v))
    if publication_clean_export_enabled():
        texts.append(compiled_workbook.name)
    else:
        texts.append(str(compiled_workbook.resolve()))
        p = compiled_workbook.resolve()
        cur: Optional[Path] = p.parent
        for _ in range(6):
            if cur is None:
                break
            texts.append(str(cur))
            if cur.name:
                texts.append(cur.name)
            parent = cur.parent
            if parent == cur:
                break
            cur = parent
    return _dedupe_preserve(texts)


def _dynamic_inference_sources(merged: pd.DataFrame, i: int, compiled_workbook: Path) -> List[str]:
    row = merged.iloc[i]
    texts: List[str] = []
    hint_cols = (
        "dynamic_detected",
        "dynamics",
        "written_dynamic",
        "dynamic_marking",
        "notated_dynamic",
        "per_note_dynamic",
    )
    for col in hint_cols:
        s = _cell_str(row, col)
        if s:
            texts.append(s)
    for col in merged.columns:
        lc = str(col).lower()
        if lc == "dynamic" or col in hint_cols:
            continue
        if "dynamic" in lc:
            s = _cell_str(row, str(col))
            if s and s.lower() not in {"unknown", "nan", "none"}:
                texts.append(s)
    for _src_col in (
        "Source_File",
        "source_file_name",
        "source_file",
        "filename",
        "file_name",
        "input_file",
    ):
        _src_series = _pick_series(merged, _src_col)
        if len(_src_series) > i:
            v = _src_series.iloc[i]
            if pd.notna(v) and str(v).strip():
                texts.append(str(v))
    sw = _pick_series(merged, "Source_Workbook")
    if len(sw) > i:
        v = sw.iloc[i]
        if pd.notna(v) and str(v).strip():
            texts.append(str(v))
    if publication_clean_export_enabled():
        texts.append(compiled_workbook.name)
    else:
        texts.append(str(compiled_workbook.resolve()))
        p = compiled_workbook.resolve()
        cur: Optional[Path] = p.parent
        for _ in range(6):
            if cur is None:
                break
            texts.append(str(cur))
            if cur.name:
                texts.append(cur.name)
            parent = cur.parent
            if parent == cur:
                break
            cur = parent
    return _dedupe_preserve(texts)


def _infer_first_from_sources(sources: Sequence[str], infer_fn) -> Optional[str]:
    for raw in sources:
        g = infer_fn(raw)
        if g:
            return g
    return None


def _build_instrument_dynamic_series(
    merged: pd.DataFrame,
    compiled_workbook: Path,
    warnings: List[str],
    meta: ResearchExportMetadata,
) -> Tuple[pd.Series, pd.Series, pd.Series, pd.Series, pd.Series]:
    """Resolve Instrument, Dynamic, Technique and inference-status columns."""
    n = len(merged)
    orig_inst = _pick_series(merged, "Instrument").copy()
    orig_dyn = _pick_series(merged, "Dynamic").copy()

    inst_out = orig_inst.astype(object)
    dyn_out = orig_dyn.astype(object)
    tech_out = _pick_series(merged, "Technique").astype(object)
    infer_status = pd.Series("provided_or_inferred", index=merged.index, dtype=object)
    infer_reason = pd.Series("", index=merged.index, dtype=object)

    for i in range(n):
        if not (pd.isna(inst_out.iloc[i]) or str(inst_out.iloc[i]).strip() == ""):
            continue
        guess = _infer_first_from_sources(
            _instrument_inference_sources(merged, i, compiled_workbook),
            infer_instrument_conservative,
        )
        if guess:
            inst_out.iloc[i] = guess

    for i in range(n):
        if not (pd.isna(dyn_out.iloc[i]) or str(dyn_out.iloc[i]).strip() == ""):
            continue
        guess = _infer_first_from_sources(
            _dynamic_inference_sources(merged, i, compiled_workbook),
            infer_dynamic_conservative,
        )
        if guess:
            dyn_out.iloc[i] = guess

    for i in range(n):
        if not (pd.isna(tech_out.iloc[i]) or str(tech_out.iloc[i]).strip() == ""):
            continue
        sources = _dedupe_preserve(
            _instrument_inference_sources(merged, i, compiled_workbook)
            + _dynamic_inference_sources(merged, i, compiled_workbook)
        )
        guess = _infer_first_from_sources(sources, infer_technique_conservative)
        if guess:
            tech_out.iloc[i] = guess

    if meta.instrument:
        for i in range(n):
            row_has = pd.notna(orig_inst.iloc[i]) and str(orig_inst.iloc[i]).strip() != ""
            if meta.force_metadata or not row_has:
                inst_out.iloc[i] = meta.instrument
    if meta.dynamic:
        for i in range(n):
            row_has = pd.notna(orig_dyn.iloc[i]) and str(orig_dyn.iloc[i]).strip() != ""
            if meta.force_metadata or not row_has:
                dyn_out.iloc[i] = meta.dynamic

    if _all_blank_or_nan(inst_out):
        warnings.append("Instrument column missing and could not be inferred; left blank.")
    if _all_blank_or_nan(dyn_out):
        warnings.append("Dynamic column missing or ambiguous; could not be inferred confidently; left blank.")

    for i in range(n):
        miss = []
        if pd.isna(inst_out.iloc[i]) or str(inst_out.iloc[i]).strip() == "":
            miss.append("instrument")
        if pd.isna(dyn_out.iloc[i]) or str(dyn_out.iloc[i]).strip() == "":
            miss.append("dynamic")
        if pd.isna(tech_out.iloc[i]) or str(tech_out.iloc[i]).strip() == "":
            miss.append("technique")
        if miss:
            infer_status.iloc[i] = "incomplete_inference"
            infer_reason.iloc[i] = "missing_" + "_".join(miss)
        else:
            infer_status.iloc[i] = "ok"
            infer_reason.iloc[i] = ""

    return inst_out, dyn_out, tech_out, infer_status, infer_reason


def _collect_note_named_directories(analysis_root: Path) -> Dict[str, List[Path]]:
    """Map note folder name -> directories under ``analysis_root`` whose final component parses as a note."""
    out: Dict[str, List[Path]] = defaultdict(list)
    if not analysis_root.is_dir():
        return dict(out)
    try:
        for dirpath, dirnames, _filenames in os.walk(analysis_root):
            for d in dirnames:
                if parse_note(d)[0]:
                    out[d].append(Path(dirpath) / d)
    except OSError:
        pass
    return dict(out)


def _pick_note_folder_for_row(
    note: str,
    candidates: List[Path],
    source_file_val: Any,
    warnings: List[str],
    warn_key: str,
) -> Optional[Path]:
    if not candidates:
        return None
    if len(candidates) == 1:
        return candidates[0]
    stem = None
    if source_file_val is not None and pd.notna(source_file_val):
        stem = Path(str(source_file_val)).stem.lower()
    scored: List[Tuple[int, str, Path]] = []
    for c in candidates:
        pnorm = str(c).lower().replace("\\", "/")
        score = 1 if stem and stem in pnorm else 0
        scored.append((score, pnorm, c))
    scored.sort(key=lambda t: (-t[0], t[1]))
    if len(scored) > 1 and scored[0][0] == scored[1][0]:
        warnings.append(
            f"{warn_key}: multiple folders for note {note!r}; using {scored[0][2]} (tie-break: sorted path)."
        )
    return scored[0][2]


def _rel_under_root(path: Path, root: Path) -> str:
    try:
        return path.resolve().relative_to(root.resolve()).as_posix()
    except ValueError:
        return path.as_posix()


def apply_per_note_chart_paths(
    sd: pd.DataFrame,
    compiled_workbook: Path,
    merged: pd.DataFrame,
    warnings: List[str],
) -> None:
    """Fill ``amplitude_mass_chart_file`` / ``energy_ratio_chart_file`` using filesystem search."""
    analysis_root = compiled_workbook.parent.resolve()
    note_index = _collect_note_named_directories(analysis_root)
    note_src = _note_to_source_file_lookup(merged)
    any_amp_still_missing = False
    any_ratio_still_missing = False

    for pos in range(len(sd)):
        idx = sd.index[pos]
        note = sd.iloc[pos].get("Note")
        if note is None or (isinstance(note, float) and np.isnan(note)):
            continue
        note_s = str(note).strip()
        src = note_src.get(note_s)
        amp_col = "amplitude_mass_chart_file"
        erg_col = "energy_ratio_chart_file"
        if amp_col not in sd.columns:
            sd[amp_col] = np.nan
        if erg_col not in sd.columns:
            sd[erg_col] = np.nan
        cur_amp = sd.at[idx, amp_col]
        cur_erg = sd.at[idx, erg_col]
        if pd.isna(cur_amp) or str(cur_amp).strip() == "":
            cands = note_index.get(note_s, [])
            folder = _pick_note_folder_for_row(
                note_s,
                cands,
                src,
                warnings,
                "Chart path",
            )
            if folder and (folder / CHART_AMPLITUDE_NAME).is_file():
                sd.at[idx, amp_col] = _rel_under_root(folder / CHART_AMPLITUDE_NAME, analysis_root)
            else:
                any_amp_still_missing = True
        if pd.isna(cur_erg) or str(cur_erg).strip() == "":
            cands = note_index.get(note_s, [])
            folder = _pick_note_folder_for_row(
                note_s,
                cands,
                src,
                warnings,
                "Chart path",
            )
            rel = None
            if folder:
                if (folder / CHART_ENERGY_RATIO_NAME).is_file():
                    rel = _rel_under_root(folder / CHART_ENERGY_RATIO_NAME, analysis_root)
                elif (folder / CHART_ENERGY_LEGACY_NAME).is_file():
                    rel = _rel_under_root(folder / CHART_ENERGY_LEGACY_NAME, analysis_root)
            if rel:
                sd.at[idx, erg_col] = rel
            else:
                any_ratio_still_missing = True

    if any_amp_still_missing:
        warnings.append(
            "Per-note component_amplitude_mass_pie.png not found under analysis folder for at least one note "
            "(see amplitude_mass_chart_file on Spectral_Density_Metrics)."
        )
    if any_ratio_still_missing:
        warnings.append(
            "Per-note component_energy_ratio_pie.png / component_energy_pie.png not found under analysis folder "
            "for at least one note (see energy_ratio_chart_file on Spectral_Density_Metrics)."
        )
    for _path_col in ("amplitude_mass_chart_file", "energy_ratio_chart_file"):
        if _path_col in sd.columns:
            _s = sd[_path_col].astype(object)
            _missing = _s.isna() | _s.astype(str).str.strip().eq("")
            sd[_path_col] = _s.where(~_missing, "not_found_under_analysis_folder")


def build_spectral_density_metrics(
    merged: pd.DataFrame,
    warnings: List[str],
    compiled_workbook: Path,
    meta: Optional[ResearchExportMetadata] = None,
    *,
    include_legacy_cdm_mean: bool = False,
) -> pd.DataFrame:
    meta = meta or ResearchExportMetadata()
    merged = _augment_source_file_name_from_per_note_processing_metadata(
        merged, compiled_workbook
    )
    note_col = "Note"
    notes = merged[note_col] if note_col in merged.columns else pd.Series(np.nan, index=merged.index)

    midi_list = [note_to_midi(x) for x in notes]
    midi = pd.Series(midi_list, index=merged.index, dtype=float)

    norm_warns: Dict[str, bool] = {}

    (
        instrument,
        dynamic,
        technique,
        metadata_inference_status,
        metadata_missing_reason,
    ) = _build_instrument_dynamic_series(merged, compiled_workbook, warnings, meta)

    f0_source_series = _series_str(merged, "f0_source")
    f0_final_source_series = _series_str(merged, "f0_final_source")
    f0_final_source_series = f0_final_source_series.where(
        f0_final_source_series.notna() & f0_final_source_series.astype(str).str.strip().ne(""),
        "unknown",
    )
    f0_fit_accepted_series = (
        merged["f0_fit_accepted"] if "f0_fit_accepted" in merged.columns else pd.Series(np.nan, index=merged.index)
    )
    f0_fit_rejection_reason_series = _series_str(merged, "f0_fit_rejection_reason")
    f0_used_for_density_source_series = _series_str(merged, "f0_used_for_density_source")
    f0_used_for_density_source_series = f0_used_for_density_source_series.where(
        f0_used_for_density_source_series.notna() & f0_used_for_density_source_series.astype(str).str.strip().ne(""),
        f0_source_series,
    )

    acoustic_f0_status_series = _series_str(merged, "acoustic_f0_status")
    if acoustic_f0_status_series.isna().all():
        derived: List[str] = []
        for i in range(len(merged)):
            acc = f0_fit_accepted_series.iloc[i]
            src = str(f0_source_series.iloc[i] if i < len(f0_source_series) else "").strip().lower()
            rej = str(
                f0_fit_rejection_reason_series.iloc[i] if i < len(f0_fit_rejection_reason_series) else ""
            ).strip().lower()
            acc_true = bool(acc is True or str(acc).strip().lower() in ("true", "1"))
            if acc_true:
                derived.append("fit_accepted_acoustically_verified")
            elif "fallback" in src or "fit_rejected" in src or rej:
                derived.append("nominal_fallback_used_not_acoustically_verified")
            else:
                derived.append("not_acoustically_verified")
        acoustic_f0_status_series = pd.Series(derived, index=merged.index)

    expected_harmonic_slot_count = _series_or_nan(merged, "expected_harmonic_slot_count")
    harmonic_slot_expected_count = _series_or_nan(merged, "harmonic_slot_expected_count")
    if expected_harmonic_slot_count.isna().all():
        expected_harmonic_slot_count = harmonic_slot_expected_count
    if harmonic_slot_expected_count.isna().all():
        harmonic_slot_expected_count = expected_harmonic_slot_count

    harmonic_region_occupancy_count = _series_or_nan(merged, "harmonic_region_occupancy_count")
    if harmonic_region_occupancy_count.isna().all():
        harmonic_region_occupancy_count = _series_or_nan(merged, "harmonic_occupancy_detected_order_count")
    if harmonic_region_occupancy_count.isna().all():
        harmonic_region_occupancy_count = _series_or_nan(merged, "detected_harmonic_slot_count")
    if harmonic_region_occupancy_count.isna().all():
        harmonic_region_occupancy_count = _series_or_nan(merged, "harmonic_order_count")
    harmonic_occupancy_detected_order_count = harmonic_region_occupancy_count
    detected_harmonic_slot_count = harmonic_region_occupancy_count

    harmonic_slot_matched_count = _series_or_nan(merged, "harmonic_slot_matched_count")
    if harmonic_slot_matched_count.isna().all():
        harmonic_slot_matched_count = detected_harmonic_slot_count

    with np.errstate(divide="ignore", invalid="ignore"):
        harmonic_slot_coverage_ratio = pd.to_numeric(harmonic_slot_matched_count, errors="coerce") / pd.to_numeric(
            harmonic_slot_expected_count, errors="coerce"
        ).replace(0, np.nan)

    component_harmonic_energy_ratio = _series_or_nan(merged, "component_harmonic_energy_ratio")
    component_inharmonic_energy_ratio = _series_or_nan(merged, "component_inharmonic_energy_ratio")
    component_subbass_energy_ratio = _series_or_nan(merged, "component_subbass_energy_ratio")
    if component_harmonic_energy_ratio.isna().all():
        component_harmonic_energy_ratio = _series_or_nan(merged, "harmonic_energy_ratio")
    if component_inharmonic_energy_ratio.isna().all():
        component_inharmonic_energy_ratio = _series_or_nan(merged, "inharmonic_energy_ratio")
    if component_subbass_energy_ratio.isna().all():
        component_subbass_energy_ratio = _series_or_nan(merged, "subbass_energy_ratio")

    _comp_h = pd.to_numeric(component_harmonic_energy_ratio, errors="coerce")
    _comp_i = pd.to_numeric(component_inharmonic_energy_ratio, errors="coerce")
    _comp_s = pd.to_numeric(component_subbass_energy_ratio, errors="coerce")
    _comp_sum = _comp_h + _comp_i + _comp_s
    with np.errstate(divide="ignore", invalid="ignore"):
        component_harmonic_energy_ratio = _comp_h / _comp_sum.replace(0.0, np.nan)
        component_inharmonic_energy_ratio = _comp_i / _comp_sum.replace(0.0, np.nan)
        component_subbass_energy_ratio = _comp_s / _comp_sum.replace(0.0, np.nan)

    core_harmonic_energy_ratio = _series_or_nan(merged, "harmonic_energy_ratio")
    core_residual_energy_ratio = _series_or_nan(merged, "residual_energy_ratio")
    if core_residual_energy_ratio.isna().all():
        core_residual_energy_ratio = _series_or_nan(merged, "component_residual_noise_energy_ratio")
    core_subbass_energy_ratio = _series_or_nan(merged, "subbass_energy_ratio")
    if core_residual_energy_ratio.isna().all():
        h_core = pd.to_numeric(core_harmonic_energy_ratio, errors="coerce")
        s_core = pd.to_numeric(core_subbass_energy_ratio, errors="coerce")
        core_residual_energy_ratio = 1.0 - h_core - s_core
    _core_h = pd.to_numeric(core_harmonic_energy_ratio, errors="coerce")
    _core_r = pd.to_numeric(core_residual_energy_ratio, errors="coerce")
    _core_s = pd.to_numeric(core_subbass_energy_ratio, errors="coerce")
    _core_sum = _core_h + _core_r + _core_s
    with np.errstate(divide="ignore", invalid="ignore"):
        core_harmonic_energy_ratio = _core_h / _core_sum.replace(0.0, np.nan)
        core_residual_energy_ratio = _core_r / _core_sum.replace(0.0, np.nan)
        core_subbass_energy_ratio = _core_s / _core_sum.replace(0.0, np.nan)

    body_weighted_effective_density = _series_or_nan(merged, "body_weighted_effective_density")
    low_mid_energy_ratio = _series_or_nan(merged, "low_mid_energy_ratio")
    harmonic_body_density = _series_or_nan(merged, "harmonic_body_density")
    harmonic_body_density_normalized = _series_or_nan(merged, "harmonic_body_density_normalized")
    expected_harmonic_slots_up_to_body_ceiling = _series_or_nan(merged, "expected_harmonic_slots_up_to_body_ceiling")
    residual_body_contribution = _series_or_nan(merged, "residual_body_contribution")
    if residual_body_contribution.isna().all():
        residual_body_contribution = (
            pd.to_numeric(core_residual_energy_ratio, errors="coerce")
            * pd.to_numeric(_series_or_nan(merged, "residual_log_frequency_occupancy"), errors="coerce")
        )
    residual_body_contribution_capped = _series_or_nan(merged, "residual_body_contribution_capped")
    if residual_body_contribution_capped.isna().all():
        residual_body_contribution_capped = pd.to_numeric(residual_body_contribution, errors="coerce").clip(upper=0.25)
    salient_harmonic_order_count_up_to_body_ceiling = _series_or_nan(
        merged, "salient_harmonic_order_count_up_to_body_ceiling"
    )
    expected_harmonic_order_count_up_to_body_ceiling = _series_or_nan(
        merged, "expected_harmonic_order_count_up_to_body_ceiling"
    )
    salient_harmonic_coverage_up_to_body_ceiling = _series_or_nan(
        merged, "salient_harmonic_coverage_up_to_body_ceiling"
    )
    if salient_harmonic_coverage_up_to_body_ceiling.isna().all():
        with np.errstate(divide="ignore", invalid="ignore"):
            salient_harmonic_coverage_up_to_body_ceiling = pd.to_numeric(
                salient_harmonic_order_count_up_to_body_ceiling, errors="coerce"
            ) / pd.to_numeric(expected_harmonic_order_count_up_to_body_ceiling, errors="coerce").replace(0, np.nan)
    theoretical_harmonic_order_count_up_to_body_ceiling = pd.to_numeric(
        expected_harmonic_order_count_up_to_body_ceiling, errors="coerce"
    )
    detected_salient_harmonic_order_count_up_to_body_ceiling = pd.to_numeric(
        salient_harmonic_order_count_up_to_body_ceiling, errors="coerce"
    )
    salient_harmonic_coverage_ratio_up_to_body_ceiling = pd.to_numeric(
        salient_harmonic_coverage_up_to_body_ceiling, errors="coerce"
    )
    salient_harmonic_mass_up_to_body_ceiling = _series_or_nan(
        merged, "salient_harmonic_mass_up_to_body_ceiling"
    )
    salient_harmonic_order_count_up_to_density_ceiling_hz = _series_or_nan(
        merged, "salient_harmonic_order_count_up_to_density_ceiling_hz"
    )
    if salient_harmonic_order_count_up_to_density_ceiling_hz.isna().all():
        salient_harmonic_order_count_up_to_density_ceiling_hz = pd.to_numeric(
            salient_harmonic_order_count_up_to_body_ceiling, errors="coerce"
        )
    expected_harmonic_order_count_up_to_density_ceiling_hz = _series_or_nan(
        merged, "expected_harmonic_order_count_up_to_density_ceiling_hz"
    )
    if expected_harmonic_order_count_up_to_density_ceiling_hz.isna().all():
        expected_harmonic_order_count_up_to_density_ceiling_hz = pd.to_numeric(
            expected_harmonic_order_count_up_to_body_ceiling, errors="coerce"
        )
    salient_harmonic_coverage_up_to_density_ceiling_hz = _series_or_nan(
        merged, "salient_harmonic_coverage_up_to_density_ceiling_hz"
    )
    if salient_harmonic_coverage_up_to_density_ceiling_hz.isna().all():
        with np.errstate(divide="ignore", invalid="ignore"):
            salient_harmonic_coverage_up_to_density_ceiling_hz = pd.to_numeric(
                salient_harmonic_order_count_up_to_density_ceiling_hz, errors="coerce"
            ) / pd.to_numeric(
                expected_harmonic_order_count_up_to_density_ceiling_hz, errors="coerce"
            ).replace(0, np.nan)
    salient_harmonic_mass_up_to_density_ceiling_hz = _series_or_nan(
        merged, "salient_harmonic_mass_up_to_density_ceiling_hz"
    )
    if salient_harmonic_mass_up_to_density_ceiling_hz.isna().all():
        salient_harmonic_mass_up_to_density_ceiling_hz = pd.to_numeric(
            salient_harmonic_mass_up_to_body_ceiling, errors="coerce"
        )
    salient_odd_harmonic_count_up_to_body_ceiling = _series_or_nan(
        merged, "salient_odd_harmonic_count_up_to_body_ceiling"
    )
    salient_even_harmonic_count_up_to_body_ceiling = _series_or_nan(
        merged, "salient_even_harmonic_count_up_to_body_ceiling"
    )
    odd_even_harmonic_energy_ratio = _series_or_nan(merged, "odd_even_harmonic_energy_ratio")
    salient_inharmonic_log_bin_count_up_to_body_ceiling = _series_or_nan(
        merged, "salient_inharmonic_log_bin_count_up_to_body_ceiling"
    )
    salient_subbass_particle_count = _series_or_nan(merged, "salient_subbass_particle_count")
    salient_inharmonic_log_bin_count_up_to_density_ceiling_hz = _series_or_nan(
        merged, "salient_inharmonic_log_bin_count_up_to_density_ceiling_hz"
    )
    if salient_inharmonic_log_bin_count_up_to_density_ceiling_hz.isna().all():
        salient_inharmonic_log_bin_count_up_to_density_ceiling_hz = pd.to_numeric(
            salient_inharmonic_log_bin_count_up_to_body_ceiling, errors="coerce"
        )
    salient_subbass_particle_count_up_to_density_ceiling_hz = _series_or_nan(
        merged, "salient_subbass_particle_count_up_to_density_ceiling_hz"
    )
    if salient_subbass_particle_count_up_to_density_ceiling_hz.isna().all():
        salient_subbass_particle_count_up_to_density_ceiling_hz = pd.to_numeric(
            salient_subbass_particle_count, errors="coerce"
        )
    harmonic_density_component = _series_or_nan(merged, "harmonic_density_component")
    inharmonic_density_component = _series_or_nan(merged, "inharmonic_density_component")
    subbass_density_component = _series_or_nan(merged, "subbass_density_component")
    harmonic_density_weight = _series_or_nan(merged, "harmonic_density_weight")
    inharmonic_density_weight = _series_or_nan(merged, "inharmonic_density_weight")
    subbass_density_weight = _series_or_nan(merged, "subbass_density_weight")
    density_summation_mode = _series_str(merged, "density_summation_mode")
    density_salience_threshold_db = _series_or_nan(merged, "density_salience_threshold_db")
    density_frequency_ceiling_hz = _series_or_nan(merged, "density_frequency_ceiling_hz")
    final_note_density_count_based = _series_or_nan(merged, "final_note_density_count_based")
    final_note_density_salience_weighted = _series_or_nan(merged, "final_note_density_salience_weighted")
    density_body_weighted_sum_body_ceiling = _series_or_nan(merged, "density_body_weighted_sum_body_ceiling")
    density_component_body_weighted_sum_body_ceiling = _series_or_nan(
        merged, "density_component_body_weighted_sum_body_ceiling"
    )
    if density_component_body_weighted_sum_body_ceiling.isna().all():
        density_component_body_weighted_sum_body_ceiling = _series_or_nan(
            merged, "density_component_body_weighted_sum_body_ceiling"
        )
    density_component_body_weighted_sum_body_ceiling = _series_or_nan(
        merged, "density_component_body_weighted_sum_body_ceiling"
    )
    if density_component_body_weighted_sum_body_ceiling.isna().all():
        density_component_body_weighted_sum_body_ceiling = pd.to_numeric(
            density_component_body_weighted_sum_body_ceiling, errors="coerce"
        )
    if density_body_weighted_sum_body_ceiling.isna().all():
        density_body_weighted_sum_body_ceiling = pd.to_numeric(final_note_density_salience_weighted, errors="coerce")
    if density_body_weighted_sum_body_ceiling.isna().all():
        density_body_weighted_sum_body_ceiling = _series_or_nan(merged, "density_metric_raw")
    harmonic_effective_component_count_body_ceiling = _series_or_nan(
        merged, "harmonic_effective_component_count_body_ceiling"
    )
    normalized_harmonic_richness_body_ceiling = _series_or_nan(
        merged, "normalized_harmonic_richness_body_ceiling"
    )
    body_density_per_expected_harmonic_slot_body_ceiling = _series_or_nan(
        merged, "body_density_per_expected_harmonic_slot_body_ceiling"
    )
    pitch_normalized_harmonic_component_energy_body_ceiling = _series_or_nan(
        merged, "pitch_normalized_harmonic_component_energy_body_ceiling"
    )
    richness_weighted_body_density_body_ceiling = _series_or_nan(
        merged, "richness_weighted_body_density_body_ceiling"
    )
    harmonic_component_energy_sum_body_ceiling = _series_or_nan(
        merged, "harmonic_component_energy_sum_body_ceiling"
    )
    if harmonic_component_energy_sum_body_ceiling.isna().all():
        harmonic_component_energy_sum_body_ceiling = _series_or_nan(merged, "harmonic_component_energy_sum_body_ceiling")
    inharmonic_component_energy_sum_body_ceiling = _series_or_nan(
        merged, "inharmonic_component_energy_sum_body_ceiling"
    )
    if inharmonic_component_energy_sum_body_ceiling.isna().all():
        inharmonic_component_energy_sum_body_ceiling = _series_or_nan(merged, "inharmonic_component_energy_sum_body_ceiling")
    subbass_component_energy_sum_body_ceiling = _series_or_nan(
        merged, "subbass_component_energy_sum_body_ceiling"
    )
    if subbass_component_energy_sum_body_ceiling.isna().all():
        subbass_component_energy_sum_body_ceiling = _series_or_nan(merged, "subbass_component_energy_sum")
    harmonic_component_energy_sum_body_ceiling = _series_or_nan(merged, "harmonic_component_energy_sum_body_ceiling")
    inharmonic_component_energy_sum_body_ceiling = _series_or_nan(merged, "inharmonic_component_energy_sum_body_ceiling")
    subbass_component_energy_sum = _series_or_nan(merged, "subbass_component_energy_sum")
    harmonic_body_energy_sum_body_ceiling = _series_or_nan(merged, "harmonic_body_energy_sum_body_ceiling")
    inharmonic_body_energy_sum_body_ceiling = _series_or_nan(merged, "inharmonic_body_energy_sum_body_ceiling")
    subbass_rumble_energy_sum = _series_or_nan(merged, "subbass_rumble_energy_sum")
    harmonic_full_spectrum_energy_sum_20khz = _series_or_nan(merged, "harmonic_full_spectrum_energy_sum_20khz")
    inharmonic_full_spectrum_energy_sum_20khz = _series_or_nan(merged, "inharmonic_full_spectrum_energy_sum_20khz")
    density_full_spectrum_weighted_sum_20khz = _series_or_nan(merged, "density_full_spectrum_weighted_sum_20khz")
    high_frequency_spectral_activity_sum = _series_or_nan(merged, "high_frequency_spectral_activity_sum")
    spectral_extension_index_20khz = _series_or_nan(merged, "spectral_extension_index_20khz")
    brightness_or_upper_spectral_activity_index_20khz = _series_or_nan(
        merged, "brightness_or_upper_spectral_activity_index_20khz"
    )
    full_spectrum_harmonic_candidate_count_20khz = _series_or_nan(
        merged, "full_spectrum_harmonic_candidate_count_20khz"
    )
    harmonic_candidate_count_20khz = _series_or_nan(merged, "harmonic_candidate_count_20khz")
    validated_harmonic_component_count_body_ceiling = _series_or_nan(
        merged, "validated_harmonic_component_count_body_ceiling"
    )
    probable_harmonic_component_count_body_ceiling = _series_or_nan(
        merged, "probable_harmonic_component_count_body_ceiling"
    )
    probable_harmonic_component_energy_sum_body_ceiling = _series_or_nan(
        merged, "probable_harmonic_component_energy_sum_body_ceiling"
    )
    if validated_harmonic_component_count_body_ceiling.isna().all():
        validated_harmonic_component_count_body_ceiling = _series_or_nan(
            merged, "validated_harmonic_component_count_body_ceiling"
        )
    validated_harmonic_component_count_body_ceiling = _series_or_nan(
        merged, "validated_harmonic_component_count_body_ceiling"
    )
    spectral_slope_db_per_harmonic = _series_or_nan(merged, "spectral_slope_db_per_harmonic")
    body_density_frequency_ceiling_hz = _series_or_nan(merged, "body_density_frequency_ceiling_hz")
    if body_density_frequency_ceiling_hz.isna().all():
        body_density_frequency_ceiling_hz = pd.to_numeric(density_frequency_ceiling_hz, errors="coerce")
    if body_density_frequency_ceiling_hz.isna().all():
        body_density_frequency_ceiling_hz = pd.Series(float(BODY_DENSITY_MAX_HZ), index=merged.index)
    full_spectrum_frequency_ceiling_hz = _series_or_nan(merged, "full_spectrum_frequency_ceiling_hz")
    if full_spectrum_frequency_ceiling_hz.isna().all():
        full_spectrum_frequency_ceiling_hz = pd.Series(float(FULL_SPECTRUM_MAX_HZ), index=merged.index)

    if harmonic_density_weight.isna().all():
        harmonic_density_weight = pd.Series(1.0, index=merged.index)
    if inharmonic_density_weight.isna().all():
        inharmonic_density_weight = pd.Series(0.5, index=merged.index)
    if subbass_density_weight.isna().all():
        subbass_density_weight = pd.Series(0.25, index=merged.index)
    if density_summation_mode.isna().all() or density_summation_mode.astype(str).str.strip().eq("").all():
        density_summation_mode = pd.Series("his_note_adaptive", index=merged.index)
    if density_salience_threshold_db.isna().all():
        density_salience_threshold_db = _series_or_nan(merged, "db_min")
    if density_salience_threshold_db.isna().all():
        density_salience_threshold_db = pd.Series(-80.0, index=merged.index)
    if density_frequency_ceiling_hz.isna().all():
        density_frequency_ceiling_hz = pd.to_numeric(full_spectrum_frequency_ceiling_hz, errors="coerce")
    if density_frequency_ceiling_hz.isna().all():
        density_frequency_ceiling_hz = pd.Series(float(FULL_SPECTRUM_MAX_HZ), index=merged.index)
    _valid_primary_str = _series_str(merged, "valid_for_primary_statistics").fillna("")
    valid_for_primary_statistics = _valid_primary_str.astype(str).str.strip().str.lower().isin(
        {"true", "1", "yes", "y"}
    )

    _mode_norm = density_summation_mode.astype(str).str.strip().str.lower()
    _harm_only = _mode_norm.isin(["harmonic_only", "harmonic-only", "h_only"])
    _w_h = pd.to_numeric(harmonic_density_weight, errors="coerce")
    _w_i = pd.to_numeric(inharmonic_density_weight, errors="coerce")
    _w_s = pd.to_numeric(subbass_density_weight, errors="coerce")
    _w_h_eff = _w_h.where(~_harm_only, 1.0)
    _w_i_eff = _w_i.where(~_harm_only, 0.0)
    _w_s_eff = _w_s.where(~_harm_only, 0.0)
    harmonic_density_weight = _w_h_eff
    inharmonic_density_weight = _w_i_eff
    subbass_density_weight = _w_s_eff

    if final_note_density_count_based.isna().all():
        final_note_density_count_based = (
            _w_h_eff * pd.to_numeric(salient_harmonic_order_count_up_to_body_ceiling, errors="coerce").fillna(0.0)
            + _w_i_eff * pd.to_numeric(salient_inharmonic_log_bin_count_up_to_body_ceiling, errors="coerce").fillna(0.0)
            + _w_s_eff * pd.to_numeric(salient_subbass_particle_count, errors="coerce").fillna(0.0)
        )
    if final_note_density_salience_weighted.isna().all():
        final_note_density_salience_weighted = (
            _w_h_eff * pd.to_numeric(harmonic_density_component, errors="coerce").fillna(0.0)
            + _w_i_eff * pd.to_numeric(inharmonic_density_component, errors="coerce").fillna(0.0)
            + _w_s_eff * pd.to_numeric(subbass_density_component, errors="coerce").fillna(0.0)
        )

    harmonic_effective_power_density_normalized = _series_or_nan(merged, "harmonic_effective_power_density_normalized")
    if harmonic_effective_power_density_normalized.isna().all():
        harmonic_effective_power_density_normalized = _series_or_nan(
            merged, "harmonic_effective_power_density_normalized_by_expected_slots"
        )
    if harmonic_effective_power_density_normalized.isna().all():
        harmonic_effective_power_density_normalized = (
            pd.to_numeric(_pick_series(merged, "harmonic_density_sum"), errors="coerce")
            / pd.to_numeric(expected_harmonic_slot_count, errors="coerce")
        )

    cdm_series = _pick_series(merged, "Combined Density Metric")

    # note_density_final — principled scalar density for each note:
    #   ER_H * harmonic_density_sum + ER_I * inharmonic_density_sum
    #   + ER_S * subbass_density_sum
    # Recomputed here from this sheet's own (normalized) measured energy
    # ratios and per-band density sums so the displayed columns stay
    # internally consistent. NaN in any of the six inputs propagates to NaN.
    _ndf_h = pd.to_numeric(component_harmonic_energy_ratio, errors="coerce") * pd.to_numeric(
        _pick_series(merged, "harmonic_density_sum"), errors="coerce"
    )
    _ndf_i = pd.to_numeric(component_inharmonic_energy_ratio, errors="coerce") * pd.to_numeric(
        _pick_series(merged, "inharmonic_density_sum"), errors="coerce"
    )
    _ndf_s = pd.to_numeric(component_subbass_energy_ratio, errors="coerce") * pd.to_numeric(
        _pick_series(merged, "subbass_density_sum"), errors="coerce"
    )
    note_density_final = _ndf_h + _ndf_i + _ndf_s

    out = pd.DataFrame(
        {
            "Instrument": instrument,
            "Note": notes,
            "MIDI": midi,
            "Pitch_Class": [pitch_class_name(x) for x in notes],
            "Octave": [parse_note(x)[2] for x in notes],
            "Register": [register_from_midi(m) for m in midi_list],
            "Dynamic": dynamic,
            "Technique": technique,
            "metadata_inference_status": metadata_inference_status,
            "metadata_missing_reason": metadata_missing_reason,
            "f0_nominal_hz": _series_or_nan(merged, "f0_nominal_hz"),
            "f0_final_hz": _pick_series(merged, "f0_final_hz"),
            "f0_source": f0_source_series,
            "f0_final_source": f0_final_source_series,
            "acoustic_f0_status": acoustic_f0_status_series,
            "f0_used_for_density_hz": _series_or_nan(merged, "f0_used_for_density_hz"),
            "f0_used_for_density_source": f0_used_for_density_source_series,
            "f0_used_for_harmonic_validation_hz": _series_or_nan(
                merged, "f0_used_for_harmonic_validation_hz"
            ),
            "f0_fit_accepted": f0_fit_accepted_series,
            "f0_fit_rejection_reason": f0_fit_rejection_reason_series,
            "f0_epistemic_status": _series_str(merged, "f0_epistemic_status"),
            "f0_validation_mode": _series_str(merged, "f0_validation_mode"),
            "nominal_prior_hz": _series_or_nan(merged, "nominal_prior_hz"),
            "f0_candidate_hz": _series_or_nan(merged, "f0_candidate_hz"),
            "f0_deviation_cents": _series_or_nan(merged, "f0_deviation_cents"),
            "low_order_match_count": _series_or_nan(merged, "low_order_match_count"),
            "odd_harmonic_match_count": _series_or_nan(merged, "odd_harmonic_match_count"),
            "even_harmonic_match_count": _series_or_nan(merged, "even_harmonic_match_count"),
            "median_abs_error_cents": _series_or_nan(merged, "median_abs_error_cents"),
            "p90_abs_error_cents": _series_or_nan(merged, "p90_abs_error_cents"),
            "harmonic_comb_score": _series_or_nan(merged, "harmonic_comb_score"),
            "f0_validation_max_hz": _series_or_nan(merged, "f0_validation_max_hz"),
            "arithmetic_validation_status": _series_str(merged, "arithmetic_validation_status"),
            "acoustic_validation_status": _series_str(merged, "acoustic_validation_status"),
            "f0_detuning_cents_from_nominal": _series_or_nan(merged, "f0_detuning_cents_from_nominal"),
            "density_metric_raw": _series_or_nan(merged, "density_metric_raw"),
            "density_component_body_weighted_sum_body_ceiling": density_component_body_weighted_sum_body_ceiling,
            "density_component_body_weighted_sum_body_ceiling": density_component_body_weighted_sum_body_ceiling,
            "density_body_weighted_sum_body_ceiling": density_body_weighted_sum_body_ceiling,
            "harmonic_effective_component_count_body_ceiling": harmonic_effective_component_count_body_ceiling,
            "normalized_harmonic_richness_body_ceiling": normalized_harmonic_richness_body_ceiling,
            "body_density_per_expected_harmonic_slot_body_ceiling": body_density_per_expected_harmonic_slot_body_ceiling,
            "pitch_normalized_harmonic_component_energy_body_ceiling": pitch_normalized_harmonic_component_energy_body_ceiling,
            "richness_weighted_body_density_body_ceiling": richness_weighted_body_density_body_ceiling,
            "harmonic_component_energy_sum_body_ceiling": harmonic_component_energy_sum_body_ceiling,
            "inharmonic_component_energy_sum_body_ceiling": inharmonic_component_energy_sum_body_ceiling,
            "subbass_component_energy_sum_body_ceiling": subbass_component_energy_sum_body_ceiling,
            "harmonic_component_energy_sum_body_ceiling": harmonic_component_energy_sum_body_ceiling,
            "inharmonic_component_energy_sum_body_ceiling": inharmonic_component_energy_sum_body_ceiling,
            "subbass_component_energy_sum": subbass_component_energy_sum,
            "harmonic_body_energy_sum_body_ceiling": harmonic_body_energy_sum_body_ceiling,
            "inharmonic_body_energy_sum_body_ceiling": inharmonic_body_energy_sum_body_ceiling,
            "subbass_rumble_energy_sum": subbass_rumble_energy_sum,
            "spectral_slope_db_per_harmonic": spectral_slope_db_per_harmonic,
            "density_full_spectrum_weighted_sum_20khz": density_full_spectrum_weighted_sum_20khz,
            "harmonic_full_spectrum_energy_sum_20khz": harmonic_full_spectrum_energy_sum_20khz,
            "inharmonic_full_spectrum_energy_sum_20khz": inharmonic_full_spectrum_energy_sum_20khz,
            "high_frequency_spectral_activity_sum": high_frequency_spectral_activity_sum,
            "spectral_extension_index_20khz": spectral_extension_index_20khz,
            "brightness_or_upper_spectral_activity_index_20khz": brightness_or_upper_spectral_activity_index_20khz,
            "full_spectrum_harmonic_candidate_count_20khz": full_spectrum_harmonic_candidate_count_20khz,
            "harmonic_candidate_count_20khz": harmonic_candidate_count_20khz,
            "validated_harmonic_component_count_body_ceiling": validated_harmonic_component_count_body_ceiling,
            "probable_harmonic_component_count_body_ceiling": probable_harmonic_component_count_body_ceiling,
            "probable_harmonic_component_energy_sum_body_ceiling": probable_harmonic_component_energy_sum_body_ceiling,
            "validated_harmonic_component_count_body_ceiling": validated_harmonic_component_count_body_ceiling,
            "density_metric_raw_source_sheet": "Density_Metrics",
            "energy_weighted_component_density_diagnostic": _series_or_nan(
                merged, "density_metric_raw"
            ),
            "density_metric_normalized": _series_or_nan(merged, "density_metric_normalized"),
            "density_weighted_sum": (
                richness_weighted_body_density_body_ceiling
                if not richness_weighted_body_density_body_ceiling.isna().all()
                else _series_or_nan(merged, "density_weighted_sum")
            ),
            "density_log_weighted": _series_or_nan(merged, "density_log_weighted"),
            "Total sum": _series_or_nan(merged, "Total sum"),
            "effective_partial_density": _series_or_nan(merged, "effective_partial_density"),
            "body_weighted_effective_density": body_weighted_effective_density,
            "low_mid_energy_ratio": low_mid_energy_ratio,
            "harmonic_body_density": harmonic_body_density,
            "expected_harmonic_slots_up_to_body_ceiling": expected_harmonic_slots_up_to_body_ceiling,
            "harmonic_body_density_normalized": harmonic_body_density_normalized,
            "residual_body_contribution": residual_body_contribution,
            "residual_body_contribution_capped": residual_body_contribution_capped,
            "salient_harmonic_order_count_up_to_body_ceiling": salient_harmonic_order_count_up_to_body_ceiling,
            "expected_harmonic_order_count_up_to_body_ceiling": expected_harmonic_order_count_up_to_body_ceiling,
            "salient_harmonic_coverage_up_to_body_ceiling": salient_harmonic_coverage_up_to_body_ceiling,
            "theoretical_harmonic_order_count_up_to_body_ceiling": theoretical_harmonic_order_count_up_to_body_ceiling,
            "detected_salient_harmonic_order_count_up_to_body_ceiling": detected_salient_harmonic_order_count_up_to_body_ceiling,
            "salient_harmonic_coverage_ratio_up_to_body_ceiling": salient_harmonic_coverage_ratio_up_to_body_ceiling,
            "salient_harmonic_mass_up_to_body_ceiling": salient_harmonic_mass_up_to_body_ceiling,
            "salient_harmonic_order_count_up_to_density_ceiling_hz": salient_harmonic_order_count_up_to_density_ceiling_hz,
            "expected_harmonic_order_count_up_to_density_ceiling_hz": expected_harmonic_order_count_up_to_density_ceiling_hz,
            "salient_harmonic_coverage_up_to_density_ceiling_hz": salient_harmonic_coverage_up_to_density_ceiling_hz,
            "salient_harmonic_mass_up_to_density_ceiling_hz": salient_harmonic_mass_up_to_density_ceiling_hz,
            "salient_odd_harmonic_count_up_to_body_ceiling": salient_odd_harmonic_count_up_to_body_ceiling,
            "salient_even_harmonic_count_up_to_body_ceiling": salient_even_harmonic_count_up_to_body_ceiling,
            "odd_even_harmonic_energy_ratio": odd_even_harmonic_energy_ratio,
            "salient_inharmonic_log_bin_count_up_to_body_ceiling": salient_inharmonic_log_bin_count_up_to_body_ceiling,
            "salient_subbass_particle_count": salient_subbass_particle_count,
            "salient_inharmonic_log_bin_count_up_to_density_ceiling_hz": salient_inharmonic_log_bin_count_up_to_density_ceiling_hz,
            "salient_subbass_particle_count_up_to_density_ceiling_hz": salient_subbass_particle_count_up_to_density_ceiling_hz,
            "final_note_density_count_based": final_note_density_count_based,
            "final_note_density_salience_weighted": final_note_density_salience_weighted,
            "note_density_final": note_density_final,
            "note_density_final_ci_low": _series_or_nan(merged, "note_density_final_ci_low"),
            "note_density_final_ci_high": _series_or_nan(merged, "note_density_final_ci_high"),
            "note_density_final_rel_uncertainty": _series_or_nan(
                merged, "note_density_final_rel_uncertainty"
            ),
            "note_density_final_uncertainty_sources": _series_or_nan(
                merged, "note_density_final_uncertainty_sources"
            ),
            "harmonic_density_component": harmonic_density_component,
            "inharmonic_density_component": inharmonic_density_component,
            "subbass_density_component": subbass_density_component,
            "harmonic_density_weight": harmonic_density_weight,
            "inharmonic_density_weight": inharmonic_density_weight,
            "subbass_density_weight": subbass_density_weight,
            "density_summation_mode": density_summation_mode,
            "valid_for_primary_statistics": valid_for_primary_statistics,
            "is_primary_comparable_profile": _series_str(
                merged, "is_primary_comparable_profile"
            ).fillna("").astype(str).str.strip().str.lower().isin({"true", "1", "yes", "y"}),
            "analysis_parameter_profile_id": _series_str(merged, "analysis_parameter_profile_id"),
            "primary_comparable_profile_definition": _series_str(
                merged, "primary_comparable_profile_definition"
            ),
            "density_confidence": _series_or_nan(merged, "density_confidence"),
            "f0_confidence": _series_or_nan(merged, "f0_confidence"),
            "harmonic_assignment_confidence": _series_or_nan(merged, "harmonic_assignment_confidence"),
            "spectral_stability_confidence": _series_or_nan(merged, "spectral_stability_confidence"),
            "qc_status": _series_str(merged, "qc_status"),
            "outlier_ratio_max_to_mean": _series_or_nan(merged, "outlier_ratio_max_to_mean"),
            "outlier_policy_applied": _series_str(merged, "outlier_policy_applied"),
            "density_winsorized": _series_or_nan(merged, "density_winsorized"),
            "density_median_based": _series_or_nan(merged, "density_median_based"),
            "density_trimmed_mean": _series_or_nan(merged, "density_trimmed_mean"),
            "sethares_status": _series_str(merged, "sethares_status"),
            "sethares_value_status": _series_str(merged, "sethares_value_status"),
            "sethares_curve_status": _series_str(merged, "sethares_curve_status"),
            "sethares_plot_status": _series_str(merged, "sethares_plot_status"),
            "density_weighted_sum_alias_of": _series_str(merged, "density_weighted_sum_alias_of"),
            "density_weighted_sum_semantic_status": _series_str(merged, "density_weighted_sum_semantic_status"),
            "density_salience_threshold_db": density_salience_threshold_db,
            "density_frequency_ceiling_hz": density_frequency_ceiling_hz,
            "body_density_frequency_ceiling_hz": body_density_frequency_ceiling_hz,
            "full_spectrum_frequency_ceiling_hz": full_spectrum_frequency_ceiling_hz,
            "harmonic_region_occupancy_count": harmonic_region_occupancy_count,
            "harmonic_occupancy_detected_order_count": harmonic_occupancy_detected_order_count,
            "harmonic_occupancy_ratio": _series_or_nan(merged, "harmonic_occupancy_ratio"),
            "expected_harmonic_slot_count": expected_harmonic_slot_count,
            "detected_harmonic_slot_count": detected_harmonic_slot_count,
            "harmonic_slot_expected_count": harmonic_slot_expected_count,
            "harmonic_slot_matched_count": harmonic_slot_matched_count,
            "harmonic_slot_coverage_ratio": harmonic_slot_coverage_ratio,
            "harmonic_effective_power_density_normalized": harmonic_effective_power_density_normalized,
            "residual_log_frequency_occupancy": _series_or_nan(
                merged, "residual_log_frequency_occupancy"
            ),
            "core_harmonic_energy_ratio": core_harmonic_energy_ratio,
            "core_residual_energy_ratio": core_residual_energy_ratio,
            "core_subbass_energy_ratio": core_subbass_energy_ratio,
            "residual_energy_ratio": core_residual_energy_ratio,
            "spectral_entropy": _series_or_nan(merged, "spectral_entropy"),
            "harmonic_density_sum": _pick_series(merged, "harmonic_density_sum"),
            "inharmonic_density_sum": _pick_series(merged, "inharmonic_density_sum"),
            "subbass_density_sum": _pick_series(merged, "subbass_density_sum"),
            "weighted_harmonic_density_contribution": _series_or_nan(merged, "weighted_harmonic_density_contribution"),
            "weighted_inharmonic_density_contribution": _series_or_nan(merged, "weighted_inharmonic_density_contribution"),
            "weighted_subbass_density_contribution": _series_or_nan(merged, "weighted_subbass_density_contribution"),
            "harmonic_energy_sum": _pick_series(merged, "harmonic_energy_sum"),
            "inharmonic_energy_sum": _pick_series(merged, "inharmonic_energy_sum"),
            "subbass_energy_sum": _pick_series(merged, "subbass_energy_sum"),
            "total_component_energy": _series_or_nan(merged, "total_component_energy"),
            "harmonic_energy_ratio": component_harmonic_energy_ratio,
            "inharmonic_energy_ratio": component_inharmonic_energy_ratio,
            "subbass_energy_ratio": component_subbass_energy_ratio,
            "component_harmonic_energy_ratio": component_harmonic_energy_ratio,
            "component_inharmonic_energy_ratio": component_inharmonic_energy_ratio,
            "component_subbass_energy_ratio": component_subbass_energy_ratio,
            "harmonic_alignment_status": _series_str(merged, "harmonic_alignment_status"),
            "harmonic_alignment_coverage_ratio": _series_or_nan(merged, "harmonic_alignment_coverage_ratio"),
            "mean_abs_harmonic_deviation_cents": _series_or_nan(merged, "mean_abs_harmonic_deviation_cents"),
            "max_abs_harmonic_deviation_cents": _series_or_nan(merged, "max_abs_harmonic_deviation_cents"),
            "debug_counts_invariant_status": _series_str(merged, "debug_counts_invariant_status"),
            "publication_output_allowed": merged["publication_output_allowed"]
            if "publication_output_allowed" in merged.columns
            else pd.Series(np.nan, index=merged.index),
        }
    )

    def _zscore(s: pd.Series) -> pd.Series:
        x = pd.to_numeric(s, errors="coerce")
        mu = float(x.mean())
        sigma = float(x.std(ddof=0))
        if not np.isfinite(sigma) or sigma <= 0.0:
            return pd.Series(np.nan, index=x.index)
        return (x - mu) / sigma

    out["spectral_body_thickness_index"] = (
        0.45 * _zscore(out["body_weighted_effective_density"])
        + 0.25 * _zscore(out["low_mid_energy_ratio"])
        + 0.20 * _zscore(out["harmonic_body_density_normalized"])
        + 0.10 * _zscore(out["residual_body_contribution_capped"])
    )

    for extra in (
        "harmonic_amplitude_sum",
        "inharmonic_amplitude_sum",
        "subbass_amplitude_sum",
        "amplitude_mass_chart_file",
        "energy_ratio_chart_file",
    ):
        out[extra] = merged[extra] if extra in merged.columns else np.nan
    for _path_col in ("amplitude_mass_chart_file", "energy_ratio_chart_file"):
        out[_path_col] = out[_path_col].astype(object)

    for col in (
        "density_metric_raw",
        "density_weighted_sum",
        "Total sum",
        "effective_partial_density",
        "body_weighted_effective_density",
        "low_mid_energy_ratio",
        "harmonic_body_density_normalized",
        "residual_body_contribution_capped",
        "spectral_body_thickness_index",
        "harmonic_occupancy_ratio",
        "harmonic_slot_coverage_ratio",
        "residual_log_frequency_occupancy",
        "core_residual_energy_ratio",
        "residual_energy_ratio",
        "spectral_entropy",
        "final_note_density_count_based",
        "final_note_density_salience_weighted",
    ):
        s = pd.to_numeric(out[col], errors="coerce")
        n, w = min_max_normalize(s)
        out[f"{col}_norm_for_chart"] = n
        if w:
            norm_warns[col] = True

    if norm_warns:
        warnings.append(
            "Min-max chart normalization undefined for: "
            + ", ".join(sorted(norm_warns.keys()))
            + " (constant or all-missing); chart columns set to NaN."
        )

    if include_legacy_cdm_mean:
        dws = pd.to_numeric(out["density_weighted_sum"], errors="coerce")
        cdm_map = pd.DataFrame({"Note": notes, "_cdm": pd.to_numeric(cdm_series, errors="coerce")})
        cdm_by_note = cdm_map.groupby("Note", as_index=True)["_cdm"].last()
        cdm = pd.to_numeric(out["Note"].map(cdm_by_note), errors="coerce")
        out["density_weighted_sum_cdm_mean"] = (dws + cdm) / 2.0
    else:
        warnings.append(
            "Legacy editorial mean density_weighted_sum_cdm_mean omitted by default "
            "(use --include-legacy-cdm-mean to export it)."
        )

    if "is_primary_comparable_profile" in out.columns:
        _pp = pd.to_numeric(out["is_primary_comparable_profile"], errors="coerce")
        _pp_true = int((_pp == 1).sum())
        _pp_total = int(len(out))
        if _pp_true < _pp_total:
            warnings.append(
                "Run-parameter comparability warning: "
                f"{_pp_total - _pp_true}/{_pp_total} note rows are not in the primary "
                "comparable profile (wf=log, threshold/ceiling runtime-configured). "
                "Use Primary_Statistics_Filtered for thesis-grade primary statistics."
            )

    out = out.sort_values("MIDI", na_position="last", kind="mergesort")
    return out


def build_component_balance(sd: pd.DataFrame, warnings: List[str]) -> pd.DataFrame:
    cols = [
        "Instrument",
        "Note",
        "MIDI",
        "Register",
        "Dynamic",
        "harmonic_density_sum",
        "inharmonic_density_sum",
        "subbass_density_sum",
        "Total sum",
        "component_harmonic_energy_ratio",
        "component_inharmonic_energy_ratio",
        "component_subbass_energy_ratio",
        "core_harmonic_energy_ratio",
        "core_residual_energy_ratio",
        "core_subbass_energy_ratio",
        "weighted_harmonic_density_contribution",
        "weighted_inharmonic_density_contribution",
        "weighted_subbass_density_contribution",
        "density_metric_raw",
        "harmonic_amplitude_sum",
        "inharmonic_amplitude_sum",
        "subbass_amplitude_sum",
        "density_weighted_sum",
        "density_log_weighted",
        "amplitude_mass_chart_file",
        "energy_ratio_chart_file",
    ]
    cb = pd.DataFrame({c: sd[c] if c in sd.columns else np.nan for c in cols})

    w_h = pd.to_numeric(cb["weighted_harmonic_density_contribution"], errors="coerce")
    w_i = pd.to_numeric(cb["weighted_inharmonic_density_contribution"], errors="coerce")
    w_s = pd.to_numeric(cb["weighted_subbass_density_contribution"], errors="coerce")
    if w_h.isna().all() and w_i.isna().all() and w_s.isna().all():
        warnings.append("Component_Balance: weighted density contributions missing; recomputed checks set to NaN.")

    cb["component_energy_ratio_sum"] = (
        pd.to_numeric(cb["component_harmonic_energy_ratio"], errors="coerce")
        + pd.to_numeric(cb["component_inharmonic_energy_ratio"], errors="coerce")
        + pd.to_numeric(cb["component_subbass_energy_ratio"], errors="coerce")
    )
    cb["core_energy_ratio_sum"] = (
        pd.to_numeric(cb["core_harmonic_energy_ratio"], errors="coerce")
        + pd.to_numeric(cb["core_residual_energy_ratio"], errors="coerce")
        + pd.to_numeric(cb["core_subbass_energy_ratio"], errors="coerce")
    )
    cb["density_metric_raw_recomputed"] = w_h + w_i + w_s
    raw = pd.to_numeric(cb["density_metric_raw"], errors="coerce")
    rec = pd.to_numeric(cb["density_metric_raw_recomputed"], errors="coerce")
    cb["density_metric_raw_difference"] = raw - rec

    h_d = pd.to_numeric(cb["harmonic_density_sum"], errors="coerce")
    i_d = pd.to_numeric(cb["inharmonic_density_sum"], errors="coerce")
    s_d = pd.to_numeric(cb["subbass_density_sum"], errors="coerce")
    if h_d.isna().all() and i_d.isna().all() and s_d.isna().all():
        warnings.append("Component_Balance: component density sums missing; total_sum_recomputed set to NaN.")

    cb["total_sum_recomputed"] = h_d + i_d + s_d
    tot = pd.to_numeric(cb["Total sum"], errors="coerce")
    cb["total_sum_difference"] = tot - pd.to_numeric(cb["total_sum_recomputed"], errors="coerce")

    def row_status(row: pd.Series) -> str:
        ers_comp = row["component_energy_ratio_sum"]
        ers_core = row["core_energy_ratio_sum"]
        dmd = row["density_metric_raw_difference"]
        tsd = row["total_sum_difference"]
        try:
            er_comp_ok = bool(pd.isna(ers_comp)) or abs(float(ers_comp) - 1.0) <= 0.01
        except (TypeError, ValueError):
            er_comp_ok = bool(pd.isna(ers_comp))
        try:
            er_core_ok = bool(pd.isna(ers_core)) or abs(float(ers_core) - 1.0) <= 0.01
        except (TypeError, ValueError):
            er_core_ok = bool(pd.isna(ers_core))

        dm_raw = row.get("density_metric_raw")
        dm_rec = row.get("density_metric_raw_recomputed")
        try:
            if pd.isna(dmd):
                dm_ok = bool(pd.isna(dm_raw) and pd.isna(dm_rec))
            else:
                dm_ok = abs(float(dmd)) <= 1e-6
        except (TypeError, ValueError):
            dm_ok = False

        tot = row.get("Total sum")
        tre = row.get("total_sum_recomputed")
        try:
            if pd.isna(tsd):
                ts_ok = pd.isna(tot) and pd.isna(tre)
            else:
                ts_ok = abs(float(tsd)) <= 1e-6
        except (TypeError, ValueError):
            ts_ok = False

        if er_comp_ok and er_core_ok and dm_ok and ts_ok:
            return "passed"
        return "warning"

    cb["component_balance_status"] = cb.apply(row_status, axis=1)
    return cb.sort_values("MIDI", na_position="last", kind="mergesort")


def build_validation_summary(merged: pd.DataFrame, sd: pd.DataFrame, warnings: List[str]) -> pd.DataFrame:
    cols = [
        "Instrument",
        "Note",
        "MIDI",
        "Register",
        "f0_nominal_hz",
        "f0_final_hz",
        "f0_source",
        "f0_final_source",
        "f0_fit_accepted",
        "acoustic_f0_status",
        "f0_fit_quality",
        "f0_fit_residual_std_hz",
        "f0_fit_rejection_reason",
        "f0_detuning_cents_from_nominal",
        "harmonic_alignment_status",
        "harmonic_alignment_coverage_ratio",
        "harmonic_alignment_energy_coverage_ratio",
        "mean_abs_harmonic_deviation_cents",
        "max_abs_harmonic_deviation_cents",
        "rms_harmonic_deviation_cents",
        "debug_counts_invariant_status",
        "debug_counts_invariant_failures",
        "input_schema_validation_status",
        "publication_output_allowed",
    ]
    base_prefix = ("Instrument", "Note", "MIDI", "Register")
    base_cols = [c for c in base_prefix if c in sd.columns]
    vs_base = sd[base_cols].copy()
    m = merged.copy()
    if "Note" not in m.columns:
        nc = find_note_column(m)
        if nc:
            m = m.rename(columns={nc: "Note"})
    m = m.groupby("Note", as_index=False).last()
    rest = [c for c in cols if c not in base_cols and c in m.columns]
    vs = vs_base.merge(m[["Note"] + rest], on="Note", how="left")
    for c in cols:
        if c not in vs.columns:
            vs[c] = np.nan
    vs = vs[cols]

    def f0_contradiction(row: pd.Series) -> bool:
        src = str(row.get("f0_source", "") or "")
        fsrc = str(row.get("f0_final_source", "") or "")
        if "prior_constrained_harmonic_fit" in (src, fsrc):
            acc = row.get("f0_fit_accepted")
            if acc is False or str(acc).lower() == "false":
                return True
        return False

    ok_align = {"ok", "excellent", "good", "passed"}

    def arithmetic_status(row: pd.Series) -> str:
        dci = str(row.get("debug_counts_invariant_status", "") or "").lower().strip()
        if dci in ("failed", "fail", "warning"):
            return "warning"
        return "passed"

    def acoustic_status(row: pd.Series) -> str:
        if f0_contradiction(row):
            return "failed_f0_provenance_contradiction"
        acc = row.get("f0_fit_accepted")
        acc_false = bool(acc is False or str(acc).strip().lower() in ("false", "0", "0.0"))
        af0 = str(row.get("acoustic_f0_status", "") or "").strip()
        if af0.lower() in ("nan", "none"):
            af0 = ""
        src = str(row.get("f0_source", "") or "").strip().lower()
        rej = str(row.get("f0_fit_rejection_reason", "") or "").strip().lower()
        if not af0 and acc_false and ("fallback" in src or "fit_rejected" in src or rej):
            af0 = "nominal_fallback_used_not_acoustically_verified"
        if acc_false and af0 != "nominal_fallback_used_not_acoustically_verified":
            return "failed_rejected_fit_missing_explicit_nominal_fallback_status"
        if af0:
            if af0 == "nominal_fallback_used_not_acoustically_verified":
                return af0
            if "not_acoustically_verified" in af0:
                return af0
        ha = str(row.get("harmonic_alignment_status", "") or "").lower().strip()
        if ha and ha not in ok_align:
            return "warning_harmonic_alignment"
        return "passed"

    vs["arithmetic_validation_status"] = vs.apply(arithmetic_status, axis=1)
    vs["acoustic_validation_status"] = vs.apply(acoustic_status, axis=1)
    vs["validation_summary_status"] = np.where(
        (vs["arithmetic_validation_status"] == "passed")
        & (vs["acoustic_validation_status"].isin(["passed"])),
        "passed",
        "warning",
    )
    if vs["f0_final_source"].isna().all():
        warnings.append("Validation_Summary: f0_final_source column missing from source workbook.")
    return vs.sort_values("MIDI", na_position="last", kind="mergesort")


def build_charts_data(sd: pd.DataFrame) -> pd.DataFrame:
    cols = [
        "Note",
        "MIDI",
        "spectral_body_thickness_index",
        "body_weighted_effective_density",
        "low_mid_energy_ratio",
        "harmonic_body_density_normalized",
        "core_residual_energy_ratio",
        "spectral_entropy",
        "salient_harmonic_order_count_up_to_body_ceiling",
        "expected_harmonic_order_count_up_to_body_ceiling",
        "salient_harmonic_coverage_up_to_body_ceiling",
        "salient_harmonic_order_count_up_to_density_ceiling_hz",
        "expected_harmonic_order_count_up_to_density_ceiling_hz",
        "salient_harmonic_coverage_up_to_density_ceiling_hz",
        "salient_inharmonic_log_bin_count_up_to_body_ceiling",
        "salient_subbass_particle_count",
        "final_note_density_count_based",
        "final_note_density_salience_weighted",
        "final_note_density_salience_weighted_norm_for_chart",
        "harmonic_density_component",
        "inharmonic_density_component",
        "subbass_density_component",
        "harmonic_density_weight",
        "inharmonic_density_weight",
        "subbass_density_weight",
        "density_summation_mode",
        "density_salience_threshold_db",
        "density_frequency_ceiling_hz",
        "harmonic_occupancy_ratio",
        "residual_log_frequency_occupancy",
        "effective_partial_density",
        "spectral_body_thickness_index_norm_for_chart",
        "body_weighted_effective_density_norm_for_chart",
        "low_mid_energy_ratio_norm_for_chart",
        "harmonic_body_density_normalized_norm_for_chart",
        "core_residual_energy_ratio_norm_for_chart",
        "spectral_entropy_norm_for_chart",
        "harmonic_occupancy_ratio_norm_for_chart",
        "residual_log_frequency_occupancy_norm_for_chart",
        "effective_partial_density_norm_for_chart",
        "density_metric_raw",
        "density_metric_raw_norm_for_chart",
        "density_weighted_sum",
        "density_weighted_sum_norm_for_chart",
        "weighted_harmonic_density_contribution",
        "weighted_inharmonic_density_contribution",
        "weighted_subbass_density_contribution",
        "core_harmonic_energy_ratio",
        "core_residual_energy_ratio",
        "core_subbass_energy_ratio",
        "component_harmonic_energy_ratio",
        "component_inharmonic_energy_ratio",
        "component_subbass_energy_ratio",
    ]
    cols = list(dict.fromkeys(cols))
    cd = pd.DataFrame({c: sd[c] for c in cols if c in sd.columns})
    for c in cols:
        if c not in cd.columns:
            cd[c] = np.nan
    return cd[cols].sort_values("MIDI", na_position="last", kind="mergesort")


def load_analysis_metadata(path: Path, warnings: List[str]) -> Dict[str, Any]:
    try:
        xl = pd.ExcelFile(path, engine="openpyxl")
        if "Analysis_Metadata" not in xl.sheet_names:
            warnings.append("Analysis_Metadata sheet missing; Metadata sheet will have blanks.")
            return {}
        am = pd.read_excel(path, sheet_name="Analysis_Metadata", engine="openpyxl", header=0)
        if am is None or am.empty:
            return {}
        cols_lower = {str(c).strip().lower() for c in am.columns}
        if "parameter" in cols_lower and "value" in cols_lower:
            pcol = next(c for c in am.columns if str(c).strip().lower() == "parameter")
            vcol = next(c for c in am.columns if str(c).strip().lower() == "value")
            out: Dict[str, Any] = {}
            for _, r in am.iterrows():
                k = str(r.get(pcol, "")).strip()
                if k:
                    out[k] = r.get(vcol)
            return out
        if len(am) == 1:
            row = am.iloc[0]
            return {str(c).strip(): row.iloc[i] for i, c in enumerate(am.columns)}
        am2 = pd.read_excel(path, sheet_name="Analysis_Metadata", engine="openpyxl", header=None)
        if am2.shape[1] < 2:
            return {}
        keys = am2.iloc[:, 0].astype(str)
        vals = am2.iloc[:, 1]
        return {str(k).strip(): v for k, v in zip(keys, vals, strict=False)}
    except Exception:
        warnings.append("Could not parse Analysis_Metadata.")
        return {}


def build_metadata_rows(
    path: Path,
    meta: Mapping[str, Any],
    sd: pd.DataFrame,
    merged: pd.DataFrame,
    warnings: List[str],
) -> pd.DataFrame:
    now = format_utc_publication_timestamp()
    source_workbook_sha256 = hashlib.sha256(path.read_bytes()).hexdigest()
    git_status_reason = "ok"
    git_commit = "unavailable_not_recorded"
    git_branch = "unavailable_not_recorded"
    try:
        _git_commit = subprocess.run(
            ["git", "rev-parse", "HEAD"],
            capture_output=True,
            text=True,
            cwd=str(path.parent),
            check=False,
        )
        if _git_commit.returncode == 0 and str(_git_commit.stdout).strip():
            git_commit = str(_git_commit.stdout).strip()
        else:
            _msg = f"{_git_commit.stderr} {_git_commit.stdout}".lower()
            if "not a git repository" in _msg:
                git_commit = "unavailable_not_a_git_repository"
                git_status_reason = "not_a_git_repository"
            else:
                git_status_reason = "git_rev_parse_failed"
    except Exception:
        git_status_reason = "git_rev_parse_exception"
    try:
        _git_branch = subprocess.run(
            ["git", "branch", "--show-current"],
            capture_output=True,
            text=True,
            cwd=str(path.parent),
            check=False,
        )
        if _git_branch.returncode == 0 and str(_git_branch.stdout).strip():
            git_branch = str(_git_branch.stdout).strip()
        elif git_status_reason == "ok":
            git_status_reason = "git_branch_lookup_failed"
    except Exception:
        if git_status_reason == "ok":
            git_status_reason = "git_branch_lookup_exception"

    meta_missing: set[str] = set()

    def mget(key: str) -> Any:
        if key in meta:
            return meta[key]
        lk = key.lower()
        for k, v in meta.items():
            if str(k).strip().lower() == lk:
                return v
        meta_missing.add(key)
        return np.nan

    def mget_required(*keys: str) -> Any:
        for k in keys:
            v = mget(k)
            if isinstance(v, str):
                vv = v.strip()
                if vv.lower() in {"not_available_at_compile_stage", "not available at compile stage"}:
                    continue
                if vv:
                    return v
            elif pd.notna(v):
                return v
            if k in sd.columns:
                s = sd[k]
                s = s[s.notna()] if isinstance(s, pd.Series) else s
                if isinstance(s, pd.Series) and not s.empty:
                    return s.iloc[0]
        return "unavailable_not_recorded"

    pitch_range = np.nan
    if sd["MIDI"].notna().any():
        pitch_range = f"{int(sd['MIDI'].min())}-{int(sd['MIDI'].max())}"

    tier_mode = _detect_tier_strategy_used(meta, merged)

    if tier_mode:
        n_fft_strategy = TIER_STRATEGY_LABEL
        n_fft_value: Any = TIER_DEPENDENT_LABEL
        hop_strategy = TIER_STRATEGY_LABEL
        hop_value: Any = TIER_DEPENDENT_LABEL
        zero_padding_value: Any = TIER_DEPENDENT_LABEL
        harmonic_tolerance_strategy = TIER_STRATEGY_LABEL
        harmonic_tolerance_value: Any = TIER_DEPENDENT_LABEL
    else:
        n_fft_strategy = mget_required("n_fft_strategy_or_tier_strategy", "tier_strategy", "tier")
        n_fft_value = mget_required("n_fft")
        if n_fft_value == "unavailable_not_recorded" and "n_fft" in merged.columns:
            v = _first_non_blank(pd.to_numeric(merged["n_fft"], errors="coerce"))
            if pd.notna(v):
                n_fft_value = v
        hop_strategy = mget_required("hop_length_strategy_or_tier_strategy", "tier_strategy", "tier")
        hop_value = mget_required("hop_length")
        if hop_value == "unavailable_not_recorded" and "hop_length" in merged.columns:
            v = _first_non_blank(pd.to_numeric(merged["hop_length"], errors="coerce"))
            if pd.notna(v):
                hop_value = v
        zero_padding_value = mget_required("zero_padding")
        if zero_padding_value == "unavailable_not_recorded":
            zero_padding_value = _derive_zero_padding_from_fft(merged)
        harmonic_tolerance_strategy = mget_required(
            "harmonic_tolerance_strategy", "use_adaptive_tolerance", "adaptive_tolerance"
        )
        harmonic_tolerance_value = mget_required("harmonic_tolerance", "tolerance")
        if harmonic_tolerance_value == "unavailable_not_recorded" and "harmonic_tolerance" in merged.columns:
            v = _first_non_blank(pd.to_numeric(merged["harmonic_tolerance"], errors="coerce"))
            if pd.notna(v):
                harmonic_tolerance_value = v
        if str(n_fft_strategy).strip() == "unavailable_not_recorded":
            n_fft_strategy = FIXED_FFT_MODE_LABEL
        if str(hop_strategy).strip() == "unavailable_not_recorded":
            hop_strategy = FIXED_FFT_MODE_LABEL
        if str(harmonic_tolerance_strategy).strip() == "unavailable_not_recorded":
            harmonic_tolerance_strategy = FIXED_FFT_MODE_LABEL

    by_note = merged.groupby("Note", as_index=False, sort=False).last() if "Note" in merged.columns else pd.DataFrame()
    lookup = by_note.set_index("Note") if (not by_note.empty and "Note" in by_note.columns) else pd.DataFrame()
    note_series = sd["Note"] if "Note" in sd.columns else pd.Series(dtype=object)
    _, frequency_min_meta = _resolve_freq_mag_field(
        note_series,
        lookup,
        meta,
        lookup_candidates=("frequency_min_hz", "freq_min"),
        meta_candidates=("frequency_min_hz", "freq_min"),
    )
    _, frequency_max_meta = _resolve_freq_mag_field(
        note_series,
        lookup,
        meta,
        lookup_candidates=("frequency_max_hz", "freq_max"),
        meta_candidates=("frequency_max_hz", "freq_max"),
    )
    _, magnitude_min_meta = _resolve_freq_mag_field(
        note_series,
        lookup,
        meta,
        lookup_candidates=("magnitude_min_db", "db_min"),
        meta_candidates=("magnitude_min_db", "db_min"),
    )
    _, magnitude_max_meta = _resolve_freq_mag_field(
        note_series,
        lookup,
        meta,
        lookup_candidates=("magnitude_max_db", "db_max"),
        meta_candidates=("magnitude_max_db", "db_max"),
    )
    freq_mag_unknown_remaining = any(
        str(v).strip() == UNKNOWN_NOT_PARSEABLE
        for v in (frequency_min_meta, frequency_max_meta, magnitude_min_meta, magnitude_max_meta)
    )

    instrument_detected = mget_required("instrument_detected", "Instrument")
    if instrument_detected == "unavailable_not_recorded":
        instrument_detected = _first_non_blank(sd["Instrument"]) if "Instrument" in sd.columns else np.nan
        if pd.isna(instrument_detected):
            instrument_detected = infer_instrument_conservative(str(path.parent.parent)) or UNKNOWN_NOT_PARSEABLE
    dynamic_detected = mget_required("dynamic_detected", "Dynamic")
    if dynamic_detected == "unavailable_not_recorded":
        dynamic_detected = _first_non_blank(sd["Dynamic"]) if "Dynamic" in sd.columns else np.nan
        if pd.isna(dynamic_detected):
            dynamic_detected = infer_dynamic_conservative(str(path.parent.parent)) or UNKNOWN_NOT_PARSEABLE

    source_corpus_path = _derive_source_corpus_path(path, meta)
    output_path_val = mget_required("output_path")
    if output_path_val == "unavailable_not_recorded":
        output_path_val = str(path.parent)

    rows = {
        "source_compiled_workbook": str(path.resolve()),
        "source_corpus_path": source_corpus_path,
        "output_path": output_path_val,
        "source_workbook_sha256": source_workbook_sha256,
        "git_commit": git_commit,
        "git_status_reason": git_status_reason,
        "git_branch": git_branch,
        "research_export_created_at": now,
        "research_export_script": SCRIPT_NAME,
        "research_export_version": SCRIPT_VERSION,
        "pipeline_contract_version": mget("pipeline_contract_version"),
        "analysis_schema_version": mget_required("ANALYSIS_SCHEMA_VERSION", "analysis_schema_version"),
        "stage1_module": mget("stage1_module"),
        "stage1_class": mget("stage1_class"),
        "stage2_module": mget("stage2_module"),
        "stage2_function": mget("stage2_function"),
        "compiled_from": mget("compiled_from"),
        "accepted_input_engine": mget("accepted_input_engine"),
        "legacy_pipeline_used": mget("legacy_pipeline_used"),
        "legacy_super_json_allowed": mget("legacy_super_json_allowed"),
        "publication_output_allowed": mget("publication_output_allowed"),
        "input_schema_validation_status": mget("input_schema_validation_status"),
        "weight_function": mget("weight_function"),
        "window_type": mget_required("window_type", "window"),
        "n_fft_strategy_or_tier_strategy": n_fft_strategy,
        "n_fft": n_fft_value,
        "hop_length_strategy_or_tier_strategy": hop_strategy,
        "hop_length": hop_value,
        "zero_padding": zero_padding_value,
        "harmonic_tolerance_strategy": harmonic_tolerance_strategy,
        "harmonic_tolerance": harmonic_tolerance_value,
        "frequency_min_hz": frequency_min_meta,
        "frequency_max_hz": frequency_max_meta,
        "magnitude_min_db": magnitude_min_meta,
        "magnitude_max_db": magnitude_max_meta,
        "density_summation_mode": mget_required("density_summation_mode"),
        "harmonic_density_weight": mget_required("harmonic_density_weight"),
        "inharmonic_density_weight": mget_required("inharmonic_density_weight"),
        "subbass_density_weight": mget_required("subbass_density_weight"),
        "density_salience_threshold_db": mget_required("density_salience_threshold_db"),
        "density_frequency_ceiling_hz": mget_required("density_frequency_ceiling_hz"),
        "notes_count": len(sd),
        "pitch_range": pitch_range,
        "harmonic_slot_coverage_ratio_formula": "harmonic_slot_matched_count / harmonic_slot_expected_count",
        "harmonic_region_occupancy_count_definition": (
            "Occupancy/slot-derived descriptor; not a strict count of detected harmonic partial orders "
            "and not bounded by floor(ceiling/f0)."
        ),
        "harmonic_occupancy_detected_order_count_definition": (
            "DEPRECATED alias of harmonic_region_occupancy_count. "
            "Occupancy/slot-derived descriptor; not a strict count of detected harmonic partial orders "
            "and not bounded by floor(ceiling/f0)."
        ),
        "theoretical_harmonic_order_count_up_to_body_ceiling_definition": (
            "alias of expected_harmonic_order_count_up_to_body_ceiling (= floor(body_density_frequency_ceiling_hz/f0))"
        ),
        "detected_salient_harmonic_order_count_up_to_body_ceiling_definition": (
            "alias of salient_harmonic_order_count_up_to_body_ceiling"
        ),
        "salient_harmonic_coverage_ratio_up_to_body_ceiling_definition": (
            "alias of salient_harmonic_coverage_up_to_body_ceiling"
        ),
        "legacy_high_ceiling_harmonic_slot_index_count_definition": (
            "legacy high-ceiling harmonic-slot index count (historical Harmonic Count/Harmonic Count (N)); "
            "not a fixed-ceiling physical harmonic-order count"
        ),
        "density_body_weighted_sum_body_ceiling_definition": (
            "alias of density_component_body_weighted_sum_body_ceiling (component-based body metric)"
        ),
        "density_component_body_weighted_sum_body_ceiling_definition": (
            "canonical component-based body metric at configured body ceiling "
            "(see body_density_frequency_ceiling_hz); legacy *_body_ceiling field is an alias"
        ),
        "density_component_body_weighted_sum_body_ceiling_definition": (
            "legacy alias of density_component_body_weighted_sum_body_ceiling"
        ),
        "harmonic_component_energy_sum_body_ceiling_definition": (
            "sum Power_raw of Harmonic Spectrum rows where include_for_density=True and "
            "Frequency<=body_density_frequency_ceiling_hz"
        ),
        "harmonic_component_energy_sum_body_ceiling_definition": (
            "legacy alias of harmonic_component_energy_sum_body_ceiling"
        ),
        "inharmonic_component_energy_sum_body_ceiling_definition": (
            "sum Power_raw of inharmonic/nonharmonic peak-candidate rows with "
            "Frequency<=body_density_frequency_ceiling_hz"
        ),
        "inharmonic_component_energy_sum_body_ceiling_definition": (
            "legacy alias of inharmonic_component_energy_sum_body_ceiling"
        ),
        "subbass_component_energy_sum_body_ceiling_definition": (
            "sum Power_raw of accepted subbass residual candidates under subbass policy "
            "with current body-ceiling configuration"
        ),
        "subbass_component_energy_sum_definition": (
            "legacy alias of subbass_component_energy_sum_body_ceiling"
        ),
        "harmonic_body_energy_sum_body_ceiling_definition": "harmonic component energy sum with runtime body ceiling",
        "inharmonic_body_energy_sum_body_ceiling_definition": "inharmonic component energy sum with runtime body ceiling",
        "subbass_rumble_energy_sum_definition": "subbass/rumble component energy sum used in body-density weighting",
        "body_band_harmonic_bin_energy_sum_body_ceiling_definition": (
            "diagnostic-only bin-integrated harmonic-band energy within runtime body ceiling"
        ),
        "body_band_residual_bin_energy_sum_body_ceiling_definition": (
            "diagnostic-only bin-integrated residual-band energy within runtime body ceiling"
        ),
        "body_band_total_bin_energy_sum_body_ceiling_definition": (
            "diagnostic-only total bin-integrated body-band energy within runtime body ceiling"
        ),
        "density_body_band_bin_integrated_index_body_ceiling_definition": (
            "diagnostic-only weighted index from bin-integrated body-band energies; not the primary fatness metric"
        ),
        "density_full_spectrum_weighted_sum_20khz_definition": (
            "full-spectrum weighted diagnostic sum (<=20000 Hz); not the primary body/fatness metric"
        ),
        "harmonic_full_spectrum_energy_sum_20khz_definition": (
            "harmonic full-spectrum energy sum (<=20000 Hz diagnostic/extension family)"
        ),
        "inharmonic_full_spectrum_energy_sum_20khz_definition": (
            "inharmonic full-spectrum energy sum (<=20000 Hz diagnostic/extension family)"
        ),
        "high_frequency_spectral_activity_sum_definition": (
            "upper-band spectral activity above body ceiling; diagnostic for brightness/extension"
        ),
        "spectral_extension_index_20khz_definition": (
            "ratio-like extension indicator comparing full-spectrum to body-limited weighted energy"
        ),
        "brightness_or_upper_spectral_activity_index_20khz_definition": (
            "upper-band activity index in the full-spectrum diagnostic family (20 kHz ceiling)"
        ),
        "full_spectrum_harmonic_candidate_count_20khz_definition": (
            "harmonic candidate count under full-spectrum ceiling (20 kHz); not a body-density harmonic-order metric"
        ),
        "harmonic_occupancy_ratio_formula": (
            "unique harmonic-order bins (nearest n*f0 within harmonic_tolerance_cents, excluding subbass) "
            "/ expected_harmonic_slot_count"
        ),
        "harmonic_occupancy_ratio_definition": (
            "acoustic_density_core occupancy metric based on accepted harmonic-order bins; "
            "not identical to validation-slot coverage"
        ),
        "body_weighted_effective_density_formula": (
            "(sum(w_body_i*sqrt(P_i))^2)/sum((w_body_i*sqrt(P_i))^2), "
            "with w_body(f)=1/(1+(f/1800)^2) on salient peaks up to runtime body ceiling"
        ),
        "spectral_body_thickness_index_formula": (
            "0.45*z(body_weighted_effective_density)+0.25*z(low_mid_energy_ratio)"
            "+0.20*z(harmonic_body_density_normalized)+0.10*z(residual_body_contribution_capped)"
        ),
        "instrument_detected": instrument_detected,
        "dynamic_detected": dynamic_detected,
    }
    if freq_mag_unknown_remaining:
        rows["frequency_magnitude_fields_recovery_status"] = FREQ_MAG_RECOVERY_PARTIAL
    rows = apply_publication_clean_research_metadata_fields(rows, workbook_basename=path.name)
    if meta_missing:
        warnings.append(
            "Metadata sheet missing keys (left blank in Metadata sheet): " + ", ".join(sorted(meta_missing))
        )
    return pd.DataFrame([{"Field": k, "Value": rows[k]} for k in rows])


def readme_lines(
    source: Path,
    warnings: List[str],
    n_notes: int,
    pitch_range: str,
    instrument: str,
    dynamic: str,
    generated: str,
    include_legacy_cdm_mean: bool = False,
) -> List[str]:
    if publication_clean_export_enabled():
        lines = [
            "Spectral density — research export workbook",
            "",
            "Source compiled workbook (filename only):",
            source.name,
            "",
            f"Generated: {generated}",
            f"Script: {SCRIPT_NAME}  version {SCRIPT_VERSION}",
            f"Number of notes in export: {n_notes}",
            "",
            "Purpose",
            "-------",
            "This workbook is a reduced, professionally formatted export intended for secondary research,",
            "plotting, thesis writing, and visual inspection.",
            "",
            "The file compiled_density_metrics.xlsx remains the complete technical and audit export.",
            "This workbook (compiled_density_metrics_research.xlsx) is the recommended research and reporting export.",
            "",
            "Metric hierarchy (read carefully)",
            "-----------------------------------",
            "density_metric_raw:",
            "    D_H·w_H + D_I·w_I + D_S·w_S where each D_* uses the compile weight_function (linear, log, sqrt,",
            "    d3, …) and w_* are measured component_harmonic / inharmonic / subbass energy ratios.",
            "",
            "density_weighted_sum:",
            "    Numerically equal to density_metric_raw. Changes when you change the compile weight_function.",
            "    harmonic_amplitude_sum (if present) is a separate linear diagnostic and does not follow that key.",
            "    Highlighted (soft blue) on Spectral_Density_Metrics.",
            "",
            "harmonic_slot_coverage_ratio:",
            "    Defined as harmonic_slot_matched_count / harmonic_slot_expected_count.",
            "",
            "harmonic_occupancy_ratio:",
            "    Acoustic core occupancy: unique harmonic-order bins passing n·f0 proximity and subbass exclusion,",
            "    divided by expected_harmonic_slot_count. This is intentionally separate from slot-coverage ratio.",
            "",
            "Combined Density Metric:",
            "    Legacy Stage-1 combined harmonic/inharmonic scalar (log/expm1 path in proc_audio).",
            "    Legacy-only (see Legacy_Compatibility sheet), not a primary Spectral_Density_Metrics field.",
            "",
            "density_weighted_sum_cdm_mean (legacy editorial blend):",
            "    Deprecated and not acoustically/dimensionally valid as a final scalar.",
            "    Exported only with --include-legacy-cdm-mean.",
            "",
            "Total sum:",
            "    Unweighted sum of per-band D values (D_H + D_I + D_S); diagnostic, not energy-ratio-weighted.",
            "",
            "effective_partial_density:",
            "    Effective spectral component participation descriptor (not the primary perceived thickness answer).",
            "",
            "spectral_body_thickness_index:",
            "    Recommended note-thickness index combining body-weighted effective density, low-mid ratio,",
            "    harmonic body density, and capped residual body contribution (corpus-relative z-score blend).",
            "",
            "spectral_entropy:",
            "    Distributional spread of spectral power.",
            "",
            "core_harmonic_energy_ratio / core_residual_energy_ratio / core_subbass_energy_ratio:",
            "    Acoustic-core peak-classification energy family (sums to ~1).",
            "",
            "component_harmonic_energy_ratio / component_inharmonic_energy_ratio / component_subbass_energy_ratio:",
            "    Component-balance energy family (sums to ~1).",
            "",
            "nonharmonic / inharmonic fields:",
            "    Interpret as nonharmonic candidate material unless stricter validation is explicitly present.",
            "",
            "subbass / low-frequency fields:",
            "    Interpret as low-frequency residual material, not automatically 'ground noise'.",
            "",
            "Instrument and dynamic metadata",
            "-------------------------------",
            "Instrument and dynamic may be read from source metadata, inferred conservatively from filenames,",
            "folder paths, and the compiled workbook location, or supplied manually with --instrument and --dynamic.",
            "Use --force-metadata to override non-empty workbook values with the CLI arguments.",
            "",
            "Sheets",
            "------",
            "README — This sheet.",
            "Dashboard — Overview KPIs and charts (charts omitted if export was run with --no-charts).",
            "Spectral_Density_Metrics — Main clean per-note table for analysis.",
            "Component_Balance — How density_metric_raw and totals relate to components and energy ratios.",
            "Validation_Summary — Compact QC fields including f0 and alignment.",
            "Charts_Data — Chart source data (sorted by MIDI); Dashboard charts reference this sheet.",
            "Metadata — Pipeline and export metadata extracted from Analysis_Metadata plus export audit fields.",
            "",
            "Warnings / Missing Fields",
            "-------------------------",
        ]
    else:
        lines = [
            "Spectral density — research export workbook",
            "",
            "Source compiled workbook:",
            str(source.resolve()),
            "",
            f"Generated: {generated}",
            f"Script: {SCRIPT_NAME}  version {SCRIPT_VERSION}",
            f"Number of notes in export: {n_notes}",
            "",
            "Purpose",
            "-------",
            "This workbook is a reduced, professionally formatted export intended for secondary research,",
            "plotting, thesis writing, and visual inspection.",
            "",
            "The file compiled_density_metrics.xlsx remains the complete technical and audit export.",
            "This workbook (compiled_density_metrics_research.xlsx) is the recommended research and reporting export.",
            "",
            "Metric hierarchy (read carefully)",
            "-----------------------------------",
            "density_metric_raw:",
            "    D_H·w_H + D_I·w_I + D_S·w_S where each D_* uses the compile weight_function (linear, log, sqrt,",
            "    d3, …) and w_* are measured component_harmonic / inharmonic / subbass energy ratios.",
            "",
            "density_weighted_sum:",
            "    Numerically equal to density_metric_raw. Changes when you change the compile weight_function.",
            "    harmonic_amplitude_sum (if present) is a separate linear diagnostic and does not follow that key.",
            "    Highlighted (soft blue) on Spectral_Density_Metrics.",
            "",
            "harmonic_slot_coverage_ratio:",
            "    Defined as harmonic_slot_matched_count / harmonic_slot_expected_count.",
            "",
            "harmonic_occupancy_ratio:",
            "    Acoustic core occupancy: unique harmonic-order bins passing n·f0 proximity and subbass exclusion,",
            "    divided by expected_harmonic_slot_count. This is intentionally separate from slot-coverage ratio.",
            "",
            "Combined Density Metric:",
            "    Legacy Stage-1 combined harmonic/inharmonic scalar (log/expm1 path in proc_audio).",
            "    Legacy-only (see Legacy_Compatibility sheet), not a primary Spectral_Density_Metrics field.",
            "",
            "density_weighted_sum_cdm_mean (legacy editorial blend):",
            "    Deprecated and not acoustically/dimensionally valid as a final scalar.",
            "    Exported only with --include-legacy-cdm-mean.",
            "",
            "Total sum:",
            "    Unweighted sum of per-band D values (D_H + D_I + D_S); diagnostic, not energy-ratio-weighted.",
            "",
            "effective_partial_density:",
            "    Effective spectral component participation descriptor (not the primary perceived thickness answer).",
            "",
            "spectral_body_thickness_index:",
            "    Recommended note-thickness index combining body-weighted effective density, low-mid ratio,",
            "    harmonic body density, and capped residual body contribution (corpus-relative z-score blend).",
            "",
            "spectral_entropy:",
            "    Distributional spread of spectral power.",
            "",
            "core_harmonic_energy_ratio / core_residual_energy_ratio / core_subbass_energy_ratio:",
            "    Acoustic-core peak-classification energy family (sums to ~1).",
            "",
            "component_harmonic_energy_ratio / component_inharmonic_energy_ratio / component_subbass_energy_ratio:",
            "    Component-balance energy family (sums to ~1).",
            "",
            "nonharmonic / inharmonic fields:",
            "    Interpret as nonharmonic candidate material unless stricter validation is explicitly present.",
            "",
            "subbass / low-frequency fields:",
            "    Interpret as low-frequency residual material, not automatically 'ground noise'.",
            "",
            "Instrument and dynamic metadata",
            "-------------------------------",
            "Instrument and dynamic may be read from source metadata, inferred conservatively from filenames,",
            "folder paths, and the compiled workbook location, or supplied manually with --instrument and --dynamic.",
            "Use --force-metadata to override non-empty workbook values with the CLI arguments.",
            "",
            "Sheets",
            "------",
            "README — This sheet.",
            "Dashboard — Overview KPIs and charts (charts omitted if export was run with --no-charts).",
            "Spectral_Density_Metrics — Main clean per-note table for analysis.",
            "Component_Balance — How density_metric_raw and totals relate to components and energy ratios.",
            "Validation_Summary — Compact QC fields including f0 and alignment.",
            "Charts_Data — Chart source data (sorted by MIDI); Dashboard charts reference this sheet.",
            "Metadata — Pipeline and export metadata extracted from Analysis_Metadata plus export audit fields.",
            "",
            "Warnings / Missing Fields",
            "-------------------------",
        ]
    if warnings:
        for w in warnings:
            lines.append(f"- {w}")
    else:
        lines.append("(none)")
    return lines


def _autosize_columns(ws, max_width: float = 48.0) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        maxlen = 0
        for cell in col:
            try:
                v = str(cell.value) if cell.value is not None else ""
            except Exception:
                v = ""
            maxlen = max(maxlen, min(len(v), 80))
        ws.column_dimensions[letter].width = min(max(10, maxlen + 2), max_width)


def _style_header_row(ws, row: int = 1, ncol: int | None = None) -> None:
    ncol = ncol or ws.max_column
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def _sanitize_dataframe_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Ensure non-blank, unique column names for Excel (duplicates get ``_2``, ``_3``, …).
    """
    out = df.copy()
    new_names: List[str] = []
    seen_count: Dict[str, int] = {}
    for i, col in enumerate(out.columns):
        if col is None or (isinstance(col, float) and pd.isna(col)):
            base = f"column_{i + 1}"
        else:
            base = str(col).strip()
        if not base:
            base = f"column_{i + 1}"
        n = seen_count.get(base, 0)
        seen_count[base] = n + 1
        if n == 0:
            new_names.append(base)
        else:
            new_names.append(f"{base}_{n + 1}")
    out.columns = new_names
    return out


def apply_simple_autofilter(ws) -> None:
    """Worksheet AutoFilter only (does not create ``xl/tables/table*.xml`` Table parts)."""
    if ws.max_row >= 2 and ws.max_column >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"


def _apply_number_formats(ws, headers: Sequence[str], formats: Mapping[str, str]) -> None:
    if ws.max_row < 2:
        return
    hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_index = {str(h): i + 1 for i, h in enumerate(hdr) if h is not None}
    for name, fmt in formats.items():
        if name not in col_index:
            continue
        ci = col_index[name]
        for r in range(2, ws.max_row + 1):
            ws.cell(r, ci).number_format = fmt


def _write_data_sheet(
    wb: Workbook,
    title: str,
    df: pd.DataFrame,
    ratio_cols: Tuple[str, ...],
    metric_cols: Tuple[str, ...],
) -> None:
    df = _sanitize_dataframe_columns(df)
    ws = wb.create_sheet(title)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    ncol = max(1, df.shape[1])
    _style_header_row(ws, 1, ncol)
    ws.freeze_panes = "A2"
    apply_simple_autofilter(ws)
    fmt_map: Dict[str, str] = {}
    for c in metric_cols:
        fmt_map[c] = "0.000000"
    for c in ratio_cols:
        fmt_map[c] = "0.00%"
    _apply_number_formats(ws, list(df.columns), fmt_map)
    _autosize_columns(ws)


def _apply_research_column_highlights(
    ws,
    column_fills: Sequence[Tuple[str, PatternFill]],
) -> None:
    """Soft column fills on ``Spectral_Density_Metrics`` (research workbook only)."""
    if ws.max_row < 1 or ws.max_column < 1:
        return
    hdr = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    for col_name, fill in column_fills:
        ci = hdr.get(col_name)
        if ci is None:
            continue
        for r in range(1, ws.max_row + 1):
            cell = ws.cell(r, ci)
            cell.fill = fill
            if r == 1:
                cell.font = RESEARCH_HIGHLIGHT_HEADER_FONT


def _apply_sdm_conditional(ws, headers: List[str | None]) -> None:
    def col_idx(name: str) -> int | None:
        for i, h in enumerate(headers, start=1):
            if h == name:
                return i
        return None

    last = ws.max_row
    if last < 2:
        return

    def add_rule(col_name: str, rule: Any) -> None:
        ci = col_idx(col_name)
        if not ci:
            return
        letter = get_column_letter(ci)
        ws.conditional_formatting.add(f"{letter}2:{letter}{last}", rule)

    green = PatternFill("solid", fgColor="C6EFCE")

    add_rule(
        "harmonic_energy_ratio",
        ColorScaleRule(
            start_type="num",
            start_value=0,
            start_color="F8696B",
            mid_type="num",
            mid_value=0.85,
            mid_color="FFEB84",
            end_type="num",
            end_value=1,
            end_color="63BE7B",
        ),
    )
    ci_d = col_idx("debug_counts_invariant_status")
    if ci_d:
        letter = get_column_letter(ci_d)
        ws.conditional_formatting.add(
            f"{letter}2:{letter}{last}",
            CellIsRule(operator="equal", formula=["passed"], fill=green, stopIfTrue=False),
        )
    ci_f = col_idx("f0_fit_accepted")
    if ci_f:
        letter = get_column_letter(ci_f)
        ws.conditional_formatting.add(f"{letter}2:{letter}{last}", CellIsRule(operator="equal", formula=["TRUE"], fill=green))
        ws.conditional_formatting.add(
            f"{letter}2:{letter}{last}", CellIsRule(operator="equal", formula=["FALSE"], fill=F0_FALSE_FILL)
        )


def _write_dashboard_layout(
    dash,
    source: Path,
    sd: pd.DataFrame,
    vs: pd.DataFrame,
    *,
    generated: str,
    pr: str,
    ins: str,
    dyn: str,
) -> int:
    """Two-column dashboard header + KPIs; returns suggested first row for charts."""
    widths = {1: 22.0, 2: 54.0, 3: 3.0, 4: 26.0, 5: 16.0, 6: 3.0, 7: 26.0, 8: 16.0}
    for ci, wi in widths.items():
        dash.column_dimensions[get_column_letter(ci)].width = wi

    dash.merge_cells("A1:B2")
    c0 = dash["A1"]
    c0.value = "Spectral Density Research Dashboard"
    c0.fill = TITLE_FILL
    c0.font = TITLE_FONT
    c0.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    dash.merge_cells("D1:H2")
    c1 = dash["D1"]
    c1.value = "Summary statistics"
    c1.font = Font(bold=True, size=14, color="FFFFFF")
    c1.fill = PatternFill("solid", fgColor="44546A")
    c1.alignment = Alignment(horizontal="center", vertical="center")

    lbl_fill = PatternFill("solid", fgColor="E7E6E6")
    r = 4
    dash.cell(r, 1, "Run context").font = SUBHEADER_FONT
    dash.merge_cells(f"D{r}:H{r}")
    dash.cell(r, 4, "Key performance indicators").font = SUBHEADER_FONT
    r += 1
    fields = [
        ("Source workbook", str(source.resolve())),
        ("Number of notes", len(sd)),
        ("Pitch range (MIDI)", pr or "n/a"),
        ("Instrument", ins or "n/a"),
        ("Dynamic", dyn or "n/a"),
        ("Generated", generated),
    ]
    if publication_clean_export_enabled():
        fields = [
            ("Compiled workbook", source.name),
            ("Number of notes", len(sd)),
            ("Pitch range (MIDI)", pr or "n/a"),
        ]
        if ins:
            fields.append(("Instrument", ins))
        if dyn:
            fields.append(("Dynamic", dyn))
        fields.append(("Generated", generated))
    for lab, val in fields:
        dash.cell(r, 1, lab).font = KPI_LABEL_FONT
        dash.cell(r, 1).fill = lbl_fill
        dash.cell(r, 1).border = KPI_BORDER
        vv = dash.cell(r, 2, val)
        vv.border = KPI_BORDER
        vv.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    def mean_col(name: str) -> float:
        if name not in sd.columns:
            return float("nan")
        return float(pd.to_numeric(sd[name], errors="coerce").mean())

    def corr_with_midi(name: str) -> float:
        if name not in sd.columns or "MIDI" not in sd.columns:
            return float("nan")
        return float(
            pd.to_numeric(sd[name], errors="coerce").corr(pd.to_numeric(sd["MIDI"], errors="coerce"))
        )

    def corr_between(a: str, b: str) -> float:
        if a not in sd.columns or b not in sd.columns:
            return float("nan")
        return float(pd.to_numeric(sd[a], errors="coerce").corr(pd.to_numeric(sd[b], errors="coerce")))

    kpis: List[Tuple[Any, Any]] = [
        ("Mean spectral_body_thickness_index", mean_col("spectral_body_thickness_index")),
        ("Mean body_weighted_effective_density", mean_col("body_weighted_effective_density")),
        ("Mean low_mid_energy_ratio", mean_col("low_mid_energy_ratio")),
        ("Mean harmonic_body_density_normalized", mean_col("harmonic_body_density_normalized")),
        (
            "Mean salient_harmonic_order_count_up_to_body_ceiling",
            mean_col("salient_harmonic_order_count_up_to_body_ceiling"),
        ),
        (
            "Corr(MIDI, salient_harmonic_order_count_up_to_body_ceiling)",
            corr_with_midi("salient_harmonic_order_count_up_to_body_ceiling"),
        ),
        (
            "Mean final_note_density_salience_weighted",
            mean_col("final_note_density_salience_weighted"),
        ),
        (
            "Corr(MIDI, final_note_density_salience_weighted)",
            corr_with_midi("final_note_density_salience_weighted"),
        ),
        (
            "Corr(final_note_density_salience_weighted, salient_harmonic_order_count_up_to_body_ceiling)",
            corr_between(
                "final_note_density_salience_weighted",
                "salient_harmonic_order_count_up_to_body_ceiling",
            ),
        ),
        ("Mean harmonic_occupancy_ratio", mean_col("harmonic_occupancy_ratio")),
        ("Mean residual_log_frequency_occupancy", mean_col("residual_log_frequency_occupancy")),
        ("Mean effective_partial_density", mean_col("effective_partial_density")),
        ("Mean spectral_entropy", mean_col("spectral_entropy")),
        ("Mean core_harmonic_energy_ratio", mean_col("core_harmonic_energy_ratio")),
        ("Arithmetic validation passed count", int((vs["arithmetic_validation_status"] == "passed").sum())),
        ("Acoustic f0 verified count", int((vs["acoustic_validation_status"] == "passed").sum())),
        (
            "f0 fallback / acoustically unverified count",
            int(
                (
                    sd.get("acoustic_f0_status", pd.Series(dtype=object))
                    .astype(str)
                    .str.contains("nominal_fallback_used_not_acoustically_verified", case=False, na=False)
                ).sum()
            ),
        ),
        (
            "f0 accepted count",
            int(sd["f0_fit_accepted"].apply(lambda x: str(x).lower() in ("true", "1")).sum())
            if "f0_fit_accepted" in sd.columns
            else 0,
        ),
    ]
    r0 = 5
    half = (len(kpis) + 1) // 2
    for i, (label, val) in enumerate(kpis):
        if i < half:
            rr = r0 + i
            lc = 4
        else:
            rr = r0 + (i - half)
            lc = 7
        dash.cell(rr, lc, label).font = KPI_LABEL_FONT
        dash.cell(rr, lc).fill = NEUTRAL_KPI_FILL
        dash.cell(rr, lc).border = KPI_BORDER
        cell_v = dash.cell(rr, lc + 1, val)
        cell_v.fill = NEUTRAL_KPI_FILL
        cell_v.border = KPI_BORDER
        if isinstance(val, float) and not np.isnan(val):
            if "ratio" in str(label).lower():
                cell_v.number_format = "0.00%"
            elif "Mean" in str(label):
                cell_v.number_format = "0.000000"

    if {"Note", "spectral_body_thickness_index"}.issubset(sd.columns):
        rank_df = sd[["Note", "spectral_body_thickness_index"]].copy()
        rank_df["spectral_body_thickness_index"] = pd.to_numeric(
            rank_df["spectral_body_thickness_index"], errors="coerce"
        )
        rank_df = rank_df.dropna(subset=["spectral_body_thickness_index"])
        if not rank_df.empty:
            top5 = rank_df.nlargest(5, "spectral_body_thickness_index")
            bot5 = rank_df.nsmallest(5, "spectral_body_thickness_index")
            start_row = r0 + max(half, len(kpis) - half) + 2
            dash.cell(start_row, 4, "Top 5 thickest notes by spectral_body_thickness_index").font = SUBHEADER_FONT
            rr = start_row + 1
            for _, row in top5.iterrows():
                dash.cell(rr, 4, str(row["Note"]))
                v = dash.cell(rr, 5, float(row["spectral_body_thickness_index"]))
                v.number_format = "0.000"
                rr += 1
            start_row_r = start_row
            dash.cell(start_row_r, 7, "Bottom 5 thinnest notes by spectral_body_thickness_index").font = SUBHEADER_FONT
            rr = start_row_r + 1
            for _, row in bot5.iterrows():
                dash.cell(rr, 7, str(row["Note"]))
                v = dash.cell(rr, 8, float(row["spectral_body_thickness_index"]))
                v.number_format = "0.000"
                rr += 1
    if {"Note", "salient_harmonic_order_count_up_to_body_ceiling"}.issubset(sd.columns):
        rank_df2 = sd[["Note", "salient_harmonic_order_count_up_to_body_ceiling"]].copy()
        rank_df2["salient_harmonic_order_count_up_to_body_ceiling"] = pd.to_numeric(
            rank_df2["salient_harmonic_order_count_up_to_body_ceiling"], errors="coerce"
        )
        rank_df2 = rank_df2.dropna(subset=["salient_harmonic_order_count_up_to_body_ceiling"])
        if not rank_df2.empty:
            top5c = rank_df2.nlargest(5, "salient_harmonic_order_count_up_to_body_ceiling")
            bot5c = rank_df2.nsmallest(5, "salient_harmonic_order_count_up_to_body_ceiling")
            start_row2 = r0 + max(half, len(kpis) - half) + 10
            dash.cell(start_row2, 4, "Top 5 by salient_harmonic_order_count_up_to_body_ceiling").font = SUBHEADER_FONT
            rr = start_row2 + 1
            for _, row in top5c.iterrows():
                dash.cell(rr, 4, str(row["Note"]))
                dash.cell(rr, 5, float(row["salient_harmonic_order_count_up_to_body_ceiling"]))
                rr += 1
            dash.cell(start_row2, 7, "Bottom 5 by salient_harmonic_order_count_up_to_body_ceiling").font = SUBHEADER_FONT
            rr = start_row2 + 1
            for _, row in bot5c.iterrows():
                dash.cell(rr, 7, str(row["Note"]))
                dash.cell(rr, 8, float(row["salient_harmonic_order_count_up_to_body_ceiling"]))
                rr += 1
    if {"Note", "final_note_density_salience_weighted"}.issubset(sd.columns):
        rank_df3 = sd[["Note", "final_note_density_salience_weighted"]].copy()
        rank_df3["final_note_density_salience_weighted"] = pd.to_numeric(
            rank_df3["final_note_density_salience_weighted"], errors="coerce"
        )
        rank_df3 = rank_df3.dropna(subset=["final_note_density_salience_weighted"])
        if not rank_df3.empty:
            top5f = rank_df3.nlargest(5, "final_note_density_salience_weighted")
            bot5f = rank_df3.nsmallest(5, "final_note_density_salience_weighted")
            start_row3 = r0 + max(half, len(kpis) - half) + 18
            dash.cell(start_row3, 4, "Top 5 densest notes by final_note_density_salience_weighted").font = SUBHEADER_FONT
            rr = start_row3 + 1
            for _, row in top5f.iterrows():
                dash.cell(rr, 4, str(row["Note"]))
                v = dash.cell(rr, 5, float(row["final_note_density_salience_weighted"]))
                v.number_format = "0.000"
                rr += 1
            dash.cell(start_row3, 7, "Bottom 5 least dense notes by final_note_density_salience_weighted").font = SUBHEADER_FONT
            rr = start_row3 + 1
            for _, row in bot5f.iterrows():
                dash.cell(rr, 7, str(row["Note"]))
                v = dash.cell(rr, 8, float(row["final_note_density_salience_weighted"]))
                v.number_format = "0.000"
                rr += 1
    bottom = r0 + max(half, len(kpis) - half)
    return bottom + 3


def _dashboard_charts(wb: Workbook, ws, charts_df: pd.DataFrame, data_start_row: int) -> None:
    """Append charts below KPI block. Charts reference Charts_Data sheet."""
    cd_sheet = wb["Charts_Data"]
    n = len(charts_df)
    if n == 0:
        return
    data_end = 1 + n
    anchor_row = data_start_row

    def ref_col(col_name: str) -> int:
        headers = list(charts_df.columns)
        return headers.index(col_name) + 1

    # Line chart 1: Note vs harmonic occupancy
    chart1 = LineChart()
    chart1.title = "Harmonic occupancy profile"
    chart1.y_axis.title = "harmonic_occupancy_ratio"
    chart1.x_axis.title = "Note"
    cats = Reference(cd_sheet, min_col=1, min_row=2, max_row=data_end)
    if "harmonic_occupancy_ratio" in charts_df.columns:
        v1 = Reference(cd_sheet, min_col=ref_col("harmonic_occupancy_ratio"), min_row=1, max_row=data_end)
        chart1.add_data(v1, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.height = 8
    chart1.width = 18
    ws.add_chart(chart1, f"A{anchor_row}")
    anchor_row += 20

    chart2 = LineChart()
    chart2.title = "Residual log-frequency occupancy"
    cats = Reference(cd_sheet, min_col=1, min_row=2, max_row=data_end)
    if "residual_log_frequency_occupancy" in charts_df.columns:
        v2 = Reference(cd_sheet, min_col=ref_col("residual_log_frequency_occupancy"), min_row=1, max_row=data_end)
        chart2.add_data(v2, titles_from_data=True)
    chart2.set_categories(cats)
    chart2.height = 8
    chart2.width = 18
    ws.add_chart(chart2, f"A{anchor_row}")
    anchor_row += 20

    chart3 = LineChart()
    chart3.title = "Normalized descriptor comparison"
    added_norm = False
    for col in (
        "spectral_body_thickness_index_norm_for_chart",
        "body_weighted_effective_density_norm_for_chart",
        "low_mid_energy_ratio_norm_for_chart",
        "harmonic_body_density_normalized_norm_for_chart",
        "harmonic_occupancy_ratio_norm_for_chart",
        "residual_log_frequency_occupancy_norm_for_chart",
        "core_residual_energy_ratio_norm_for_chart",
        "spectral_entropy_norm_for_chart",
        "effective_partial_density_norm_for_chart",
    ):
        if col not in charts_df.columns:
            continue
        v = Reference(cd_sheet, min_col=ref_col(col), min_row=1, max_row=data_end)
        chart3.add_data(v, titles_from_data=True)
        added_norm = True
    if added_norm:
        chart3.set_categories(Reference(cd_sheet, min_col=1, min_row=2, max_row=data_end))
        chart3.height = 9
        chart3.width = 20
        ws.add_chart(chart3, f"A{anchor_row}")
        anchor_row += 22

    chart4 = BarChart()
    chart4.type = "col"
    chart4.grouping = "stacked"
    chart4.title = "Weighted component contributions"
    chart4.overlap = 100
    for col in (
        "weighted_harmonic_density_contribution",
        "weighted_inharmonic_density_contribution",
        "weighted_subbass_density_contribution",
    ):
        v = Reference(cd_sheet, min_col=ref_col(col), min_row=1, max_row=data_end)
        chart4.add_data(v, titles_from_data=True)
    chart4.set_categories(Reference(cd_sheet, min_col=1, min_row=2, max_row=data_end))
    chart4.height = 10
    chart4.width = 18
    ws.add_chart(chart4, f"A{anchor_row}")
    anchor_row += 24

    chart5 = BarChart()
    chart5.type = "col"
    chart5.grouping = "percentStacked"
    chart5.title = "Component energy ratios"
    for col in (
        "component_harmonic_energy_ratio",
        "component_inharmonic_energy_ratio",
        "component_subbass_energy_ratio",
    ):
        if col not in charts_df.columns:
            continue
        v = Reference(cd_sheet, min_col=ref_col(col), min_row=1, max_row=data_end)
        chart5.add_data(v, titles_from_data=True)
    chart5.set_categories(Reference(cd_sheet, min_col=1, min_row=2, max_row=data_end))
    chart5.height = 10
    chart5.width = 18
    ws.add_chart(chart5, f"A{anchor_row}")
    anchor_row += 22

    if "salient_harmonic_order_count_up_to_body_ceiling" in charts_df.columns:
        chart6 = LineChart()
        chart6.title = "MIDI vs salient_harmonic_order_count_up_to_body_ceiling"
        chart6.y_axis.title = "salient_harmonic_order_count_up_to_body_ceiling"
        chart6.x_axis.title = "MIDI"
        v = Reference(
            cd_sheet,
            min_col=ref_col("salient_harmonic_order_count_up_to_body_ceiling"),
            min_row=1,
            max_row=data_end,
        )
        chart6.add_data(v, titles_from_data=True)
        chart6.set_categories(Reference(cd_sheet, min_col=ref_col("MIDI"), min_row=2, max_row=data_end))
        chart6.height = 8
        chart6.width = 18
        ws.add_chart(chart6, f"A{anchor_row}")


def build_workbook(
    source: Path,
    output: Path,
    *,
    no_charts: bool,
    overwrite: bool,
    research_metadata: Optional[ResearchExportMetadata] = None,
    include_legacy_cdm_mean: bool = False,
) -> List[str]:
    warnings: List[str] = []
    if not source.is_file():
        raise FileNotFoundError(f"Input file does not exist: {source}")
    if output.exists() and not overwrite:
        raise FileExistsError(
            f"Output file already exists: {output}\n"
            "Pass --overwrite to replace it, or choose a different --output path."
        )

    meta = research_metadata or ResearchExportMetadata()
    merged = merge_workbook_frames(source, warnings)
    merged = _rename_frame_to_canonical(merged)
    merged = publication_research_canonical_density_columns(merged)
    sd = build_spectral_density_metrics(
        merged,
        warnings,
        source,
        meta,
        include_legacy_cdm_mean=include_legacy_cdm_mean,
    )
    _dws_candidates = (
        "richness_weighted_body_density_body_ceiling",
        "body_density_per_expected_harmonic_slot_body_ceiling",
        "density_component_body_weighted_sum_body_ceiling",
        "density_component_body_weighted_sum_body_ceiling",
        "density_body_weighted_sum_body_ceiling",
    )
    _dws_selected = None
    for _c in _dws_candidates:
        if _c not in sd.columns:
            continue
        _s = pd.to_numeric(sd[_c], errors="coerce")
        if _s.notna().any():
            sd["density_weighted_sum"] = _s
            _dws_selected = _c
            break
    if _dws_selected is not None:
        sd["density_weighted_sum_alias_of"] = _dws_selected
        sd["density_weighted_sum_semantic_status"] = "body_limited_runtime_ceiling_primary_metric"
    # Keep debug/bin/candidate counts out of research-facing sheets.
    for _diag_col in ("harmonic_bin_count", "harmonic_peak_candidate_count"):
        if _diag_col in sd.columns:
            sd = sd.drop(columns=[_diag_col], errors="ignore")
    if "legacy_high_ceiling_harmonic_slot_index_count" not in sd.columns:
        if "Harmonic Count (N)" in sd.columns:
            sd["legacy_high_ceiling_harmonic_slot_index_count"] = pd.to_numeric(
                sd["Harmonic Count (N)"], errors="coerce"
            )
        elif "Harmonic Count" in sd.columns:
            sd["legacy_high_ceiling_harmonic_slot_index_count"] = pd.to_numeric(
                sd["Harmonic Count"], errors="coerce"
            )
    sd = sd.drop(
        columns=[
            "Harmonic Count",
            "Harmonic Count (N)",
            "Harmonic Count (relative)",
            "Harmonic Ceiling (relative)",
            "harmonic_order_count",
            "harmonic_occupancy_detected_order_count",
        ],
        errors="ignore",
    )
    apply_per_note_chart_paths(sd, source, merged, warnings)
    required_front_cols = [
        "Technique",
        "metadata_inference_status",
        "metadata_missing_reason",
        "f0_final_source",
        "amplitude_mass_chart_file",
        "energy_ratio_chart_file",
        "f0_used_for_density_hz",
        "f0_used_for_density_source",
        "acoustic_f0_status",
        "density_weighted_sum",
        "density_weighted_sum_alias_of",
        "density_weighted_sum_semantic_status",
        "density_component_body_weighted_sum_body_ceiling",
        "harmonic_component_energy_sum_body_ceiling",
        "inharmonic_component_energy_sum_body_ceiling",
        "density_component_body_weighted_sum_body_ceiling",
        "harmonic_component_energy_sum_body_ceiling",
        "inharmonic_component_energy_sum_body_ceiling",
        "subbass_component_energy_sum",
        "spectral_slope_db_per_harmonic",
        "validated_harmonic_component_count_body_ceiling",
        "probable_harmonic_component_count_body_ceiling",
        "probable_harmonic_component_energy_sum_body_ceiling",
        "validated_harmonic_component_count_body_ceiling",
        "harmonic_candidate_count_20khz",
        "spectral_body_thickness_index",
        "body_weighted_effective_density",
        "low_mid_energy_ratio",
        "harmonic_body_density_normalized",
        "salient_harmonic_order_count_up_to_body_ceiling",
        "expected_harmonic_order_count_up_to_body_ceiling",
        "salient_harmonic_coverage_up_to_body_ceiling",
        "theoretical_harmonic_order_count_up_to_body_ceiling",
        "detected_salient_harmonic_order_count_up_to_body_ceiling",
        "salient_harmonic_coverage_ratio_up_to_body_ceiling",
        "salient_harmonic_mass_up_to_body_ceiling",
        "salient_harmonic_order_count_up_to_density_ceiling_hz",
        "expected_harmonic_order_count_up_to_density_ceiling_hz",
        "salient_harmonic_coverage_up_to_density_ceiling_hz",
        "salient_harmonic_mass_up_to_density_ceiling_hz",
        "salient_odd_harmonic_count_up_to_body_ceiling",
        "salient_even_harmonic_count_up_to_body_ceiling",
        "odd_even_harmonic_energy_ratio",
        "salient_inharmonic_log_bin_count_up_to_body_ceiling",
        "salient_subbass_particle_count",
        "salient_inharmonic_log_bin_count_up_to_density_ceiling_hz",
        "salient_subbass_particle_count_up_to_density_ceiling_hz",
        "final_note_density_count_based",
        "final_note_density_salience_weighted",
        "final_note_density_salience_weighted_norm_for_chart",
        "harmonic_density_component",
        "inharmonic_density_component",
        "subbass_density_component",
        "harmonic_density_weight",
        "inharmonic_density_weight",
        "subbass_density_weight",
        "density_summation_mode",
        "density_salience_threshold_db",
        "density_frequency_ceiling_hz",
        "harmonic_region_occupancy_count",
        "legacy_high_ceiling_harmonic_slot_index_count",
        "harmonic_occupancy_ratio",
        "expected_harmonic_slot_count",
        "detected_harmonic_slot_count",
        "harmonic_slot_expected_count",
        "harmonic_slot_matched_count",
        "harmonic_slot_coverage_ratio",
        "harmonic_effective_power_density_normalized",
        "residual_log_frequency_occupancy",
        "core_harmonic_energy_ratio",
        "core_residual_energy_ratio",
        "core_subbass_energy_ratio",
        "residual_body_contribution_capped",
        "component_harmonic_energy_ratio",
        "component_inharmonic_energy_ratio",
        "component_subbass_energy_ratio",
    ]
    required_front_backup = {
        c: sd[c].copy()
        for c in required_front_cols
        if c in sd.columns
    }
    if publication_clean_export_enabled():
        sd = publication_clean_drop_known_sparse_columns(sd)
        sd = drop_publication_noise_columns_from_dataframe(sd)
        for c, s in required_front_backup.items():
            if c not in sd.columns:
                sd[c] = s
    _front = [c for c in required_front_cols if c in sd.columns]
    _rest = [c for c in sd.columns if c not in _front]
    sd = sd.loc[:, _front + _rest].copy()
    cb = build_component_balance(sd, warnings)
    vs = build_validation_summary(merged, sd, warnings)
    try:
        _vs_cols = ["Note", "arithmetic_validation_status", "acoustic_validation_status"]
        _vs_map = vs[_vs_cols].drop_duplicates(subset=["Note"])
        sd = sd.merge(_vs_map, on="Note", how="left", suffixes=("", "_vs"))
        for _c in ("arithmetic_validation_status", "acoustic_validation_status"):
            _alt = f"{_c}_vs"
            if _alt in sd.columns:
                if _c in sd.columns:
                    sd[_c] = sd[_c].where(sd[_c].astype(str).str.strip().ne(""), sd[_alt])
                else:
                    sd[_c] = sd[_alt]
                sd = sd.drop(columns=[_alt])
    except Exception as _e_vs_merge:
        warnings.append(f"Validation status merge into Spectral_Density_Metrics failed: {_e_vs_merge}")
    cd = build_charts_data(sd)
    legacy_notes = merged["Note"] if "Note" in merged.columns else pd.Series(np.nan, index=merged.index)
    legacy_df = pd.DataFrame(
        {
            "Note": legacy_notes,
            "MIDI": pd.to_numeric(legacy_notes.map(note_to_midi), errors="coerce"),
            "Combined Density Metric": _pick_series(merged, "Combined Density Metric"),
            "Weighted Combined Metric": _series_or_nan(merged, "Weighted Combined Metric"),
            "Total Metric": _series_or_nan(merged, "Total Metric"),
        }
    )
    if include_legacy_cdm_mean:
        legacy_df["density_weighted_sum_cdm_mean"] = _series_or_nan(sd, "density_weighted_sum_cdm_mean")
    legacy_df = legacy_df.sort_values("MIDI", na_position="last", kind="mergesort")
    if publication_clean_export_enabled():
        cb = drop_publication_noise_columns_from_dataframe(cb)
        vs = drop_publication_noise_columns_from_dataframe(vs)
        cd = drop_publication_noise_columns_from_dataframe(cd)
        # Keep key final-density plotting columns visible in Charts_Data even when
        # a specific source workbook leaves them all-missing.
        for _cc in (
            "salient_harmonic_order_count_up_to_body_ceiling",
            "salient_inharmonic_log_bin_count_up_to_body_ceiling",
            "salient_subbass_particle_count",
            "final_note_density_count_based",
            "final_note_density_salience_weighted",
            "final_note_density_salience_weighted_norm_for_chart",
            "harmonic_density_component",
            "inharmonic_density_component",
            "subbass_density_component",
        ):
            if _cc not in cd.columns:
                cd[_cc] = np.nan
    meta_map = load_analysis_metadata(source, warnings)
    meta_df = build_metadata_rows(source, meta_map, sd, merged, warnings)
    settings_by_note = build_analysis_settings_by_note(merged, sd, meta_map)
    generated = format_utc_publication_timestamp()
    pr = ""
    if sd["MIDI"].notna().any():
        pr = f"{int(sd['MIDI'].min())}–{int(sd['MIDI'].max())}"
    ins = (
        str(sd["Instrument"].dropna().iloc[0])
        if "Instrument" in sd.columns and sd["Instrument"].notna().any()
        else ""
    )
    dyn = (
        str(sd["Dynamic"].dropna().iloc[0])
        if "Dynamic" in sd.columns and sd["Dynamic"].notna().any()
        else ""
    )

    wb = Workbook()
    # README
    rm = wb.active
    rm.title = "README"
    for line in readme_lines(
        source,
        warnings,
        len(sd),
        pr,
        ins,
        dyn,
        generated,
        include_legacy_cdm_mean=include_legacy_cdm_mean,
    ):
        rm.append([line])
    for row in range(1, rm.max_row + 1):
        v = rm.cell(row, 1).value
        if isinstance(v, str) and v and v[0].isupper() and not v.startswith(" "):
            if v in (
                "Purpose",
                "Metric hierarchy (read carefully)",
                "Instrument and dynamic metadata",
                "Sheets",
                "Warnings / Missing Fields",
            ):
                rm.cell(row, 1).font = SUBHEADER_FONT
    rm.column_dimensions["A"].width = 108

    # Dashboard (charts added after Charts_Data exists)
    dash = wb.create_sheet("Dashboard")
    chart_anchor = _write_dashboard_layout(
        dash, source, sd, vs, generated=generated, pr=pr, ins=ins, dyn=dyn
    )

    # Data sheets order: create Charts_Data before Dashboard charts
    ratio_cols = (
        "core_harmonic_energy_ratio",
        "core_residual_energy_ratio",
        "core_subbass_energy_ratio",
        "component_harmonic_energy_ratio",
        "component_inharmonic_energy_ratio",
        "component_subbass_energy_ratio",
        "harmonic_alignment_coverage_ratio",
    )
    metric_cols = [
        "f0_used_for_density_hz",
        "f0_used_for_density_source",
        "f0_fit_accepted",
        "acoustic_f0_status",
        "f0_validation_mode",
        "nominal_prior_hz",
        "f0_candidate_hz",
        "f0_deviation_cents",
        "low_order_match_count",
        "odd_harmonic_match_count",
        "even_harmonic_match_count",
        "median_abs_error_cents",
        "p90_abs_error_cents",
        "harmonic_comb_score",
        "f0_validation_max_hz",
        "arithmetic_validation_status",
        "acoustic_validation_status",
        "density_metric_raw",
        "density_component_body_weighted_sum_body_ceiling",
        "harmonic_component_energy_sum_body_ceiling",
        "inharmonic_component_energy_sum_body_ceiling",
        "density_component_body_weighted_sum_body_ceiling",
        "harmonic_component_energy_sum_body_ceiling",
        "inharmonic_component_energy_sum_body_ceiling",
        "subbass_component_energy_sum",
        "spectral_slope_db_per_harmonic",
        "density_body_weighted_sum_body_ceiling",
        "harmonic_body_energy_sum_body_ceiling",
        "inharmonic_body_energy_sum_body_ceiling",
        "subbass_rumble_energy_sum",
        "body_band_harmonic_bin_energy_sum_body_ceiling",
        "body_band_residual_bin_energy_sum_body_ceiling",
        "body_band_total_bin_energy_sum_body_ceiling",
        "density_body_band_bin_integrated_index_body_ceiling",
        "density_metric_raw_source_sheet",
        "energy_weighted_component_density_diagnostic",
        "density_weighted_sum",
        "density_log_weighted",
        "density_winsorized",
        "density_median_based",
        "density_trimmed_mean",
        "Total sum",
        "spectral_body_thickness_index",
        "body_weighted_effective_density",
        "low_mid_energy_ratio",
        "harmonic_body_density",
        "harmonic_body_density_normalized",
        "salient_harmonic_order_count_up_to_body_ceiling",
        "expected_harmonic_order_count_up_to_body_ceiling",
        "salient_harmonic_coverage_up_to_body_ceiling",
        "theoretical_harmonic_order_count_up_to_body_ceiling",
        "detected_salient_harmonic_order_count_up_to_body_ceiling",
        "salient_harmonic_coverage_ratio_up_to_body_ceiling",
        "salient_harmonic_mass_up_to_body_ceiling",
        "salient_harmonic_order_count_up_to_density_ceiling_hz",
        "expected_harmonic_order_count_up_to_density_ceiling_hz",
        "salient_harmonic_coverage_up_to_density_ceiling_hz",
        "salient_harmonic_mass_up_to_density_ceiling_hz",
        "salient_odd_harmonic_count_up_to_body_ceiling",
        "salient_even_harmonic_count_up_to_body_ceiling",
        "odd_even_harmonic_energy_ratio",
        "salient_inharmonic_log_bin_count_up_to_body_ceiling",
        "salient_subbass_particle_count",
        "salient_inharmonic_log_bin_count_up_to_density_ceiling_hz",
        "salient_subbass_particle_count_up_to_density_ceiling_hz",
        "final_note_density_count_based",
        "final_note_density_salience_weighted",
        "note_density_final",
        "note_density_final_ci_low",
        "note_density_final_ci_high",
        "note_density_final_rel_uncertainty",
        "note_density_final_uncertainty_sources",
        "final_note_density_salience_weighted_norm_for_chart",
        "harmonic_density_component",
        "inharmonic_density_component",
        "subbass_density_component",
        "harmonic_density_weight",
        "inharmonic_density_weight",
        "subbass_density_weight",
        "density_summation_mode",
        "valid_for_primary_statistics",
        "is_primary_comparable_profile",
        "analysis_parameter_profile_id",
        "primary_comparable_profile_definition",
        "density_confidence",
        "f0_confidence",
        "harmonic_assignment_confidence",
        "spectral_stability_confidence",
        "qc_status",
        "outlier_ratio_max_to_mean",
        "outlier_policy_applied",
        "sethares_value_status",
        "sethares_curve_status",
        "sethares_plot_status",
        "density_salience_threshold_db",
        "density_frequency_ceiling_hz",
        "body_density_frequency_ceiling_hz",
        "full_spectrum_frequency_ceiling_hz",
        "density_full_spectrum_weighted_sum_20khz",
        "harmonic_full_spectrum_energy_sum_20khz",
        "inharmonic_full_spectrum_energy_sum_20khz",
        "high_frequency_spectral_activity_sum",
        "spectral_extension_index_20khz",
        "brightness_or_upper_spectral_activity_index_20khz",
        "full_spectrum_harmonic_candidate_count_20khz",
        "harmonic_candidate_count_20khz",
        "validated_harmonic_component_count_body_ceiling",
        "probable_harmonic_component_count_body_ceiling",
        "probable_harmonic_component_energy_sum_body_ceiling",
        "validated_harmonic_component_count_body_ceiling",
        "residual_body_contribution",
        "residual_body_contribution_capped",
        "effective_partial_density",
        "harmonic_region_occupancy_count",
        "legacy_high_ceiling_harmonic_slot_index_count",
        "harmonic_occupancy_ratio",
        "expected_harmonic_slot_count",
        "detected_harmonic_slot_count",
        "harmonic_slot_expected_count",
        "harmonic_slot_matched_count",
        "harmonic_slot_coverage_ratio",
        "harmonic_effective_power_density_normalized",
        "residual_log_frequency_occupancy",
        "core_harmonic_energy_ratio",
        "core_residual_energy_ratio",
        "core_subbass_energy_ratio",
        "residual_energy_ratio",
        "spectral_entropy",
        "harmonic_density_sum",
        "inharmonic_density_sum",
        "subbass_density_sum",
        "weighted_harmonic_density_contribution",
        "weighted_inharmonic_density_contribution",
        "weighted_subbass_density_contribution",
        "harmonic_energy_sum",
        "inharmonic_energy_sum",
        "subbass_energy_sum",
        "total_component_energy",
        "component_harmonic_energy_ratio",
        "component_inharmonic_energy_ratio",
        "component_subbass_energy_ratio",
        "f0_nominal_hz",
        "f0_final_hz",
        "f0_detuning_cents_from_nominal",
        "mean_abs_harmonic_deviation_cents",
        "max_abs_harmonic_deviation_cents",
    ]
    if include_legacy_cdm_mean:
        metric_cols.insert(4, "density_weighted_sum_cdm_mean")

    _write_data_sheet(wb, "Spectral_Density_Metrics", sd, ratio_cols, tuple(metric_cols))
    sdm_ws = wb["Spectral_Density_Metrics"]
    hdrs = [sdm_ws.cell(1, c).value for c in range(1, sdm_ws.max_column + 1)]
    _hl = [
        ("density_weighted_sum", RESEARCH_FILL_DENSITY_WEIGHTED_SUM),
        ("note_density_final", RESEARCH_FILL_NOTE_DENSITY_FINAL),
    ]
    if include_legacy_cdm_mean:
        _hl.append(("density_weighted_sum_cdm_mean", RESEARCH_FILL_DWS_CDM_MEAN))
    _apply_research_column_highlights(sdm_ws, tuple(_hl))
    _apply_sdm_conditional(sdm_ws, hdrs)

    # QC-governed primary statistical table (defaults to acoustically validated rows).
    primary_sd = sd.copy()
    if "valid_for_primary_statistics" in primary_sd.columns:
        primary_sd = primary_sd[primary_sd["valid_for_primary_statistics"] == True].copy()  # noqa: E712
    if "is_primary_comparable_profile" in primary_sd.columns:
        primary_sd = primary_sd[primary_sd["is_primary_comparable_profile"] == True].copy()  # noqa: E712
    if primary_sd.empty:
        primary_sd = pd.DataFrame(
            [
                {
                    "primary_statistics_status": (
                        "no rows passed valid_for_primary_statistics + "
                        "is_primary_comparable_profile filters"
                    )
                }
            ]
        )
        _write_data_sheet(wb, "Primary_Statistics_Filtered", primary_sd, tuple(), ("primary_statistics_status",))
    else:
        _write_data_sheet(
            wb,
            "Primary_Statistics_Filtered",
            primary_sd,
            ratio_cols,
            tuple(metric_cols),
        )

    cb_ratios = (
        "component_harmonic_energy_ratio",
        "component_inharmonic_energy_ratio",
        "component_subbass_energy_ratio",
        "component_energy_ratio_sum",
        "core_harmonic_energy_ratio",
        "core_residual_energy_ratio",
        "core_subbass_energy_ratio",
        "core_energy_ratio_sum",
    )
    cb_metrics = (
        "harmonic_density_sum",
        "inharmonic_density_sum",
        "subbass_density_sum",
        "Total sum",
        "weighted_harmonic_density_contribution",
        "weighted_inharmonic_density_contribution",
        "weighted_subbass_density_contribution",
        "density_metric_raw",
        "density_metric_raw_recomputed",
        "density_metric_raw_difference",
        "total_sum_recomputed",
        "total_sum_difference",
        "harmonic_amplitude_sum",
        "inharmonic_amplitude_sum",
        "subbass_amplitude_sum",
        "density_weighted_sum",
        "density_log_weighted",
    )
    _write_data_sheet(wb, "Component_Balance", cb, cb_ratios, cb_metrics)

    vs_ratios = ("harmonic_alignment_coverage_ratio", "harmonic_alignment_energy_coverage_ratio")
    vs_metrics = (
        "f0_nominal_hz",
        "f0_final_hz",
        "f0_fit_residual_std_hz",
        "f0_detuning_cents_from_nominal",
        "mean_abs_harmonic_deviation_cents",
        "max_abs_harmonic_deviation_cents",
        "rms_harmonic_deviation_cents",
    )
    _write_data_sheet(wb, "Validation_Summary", vs, vs_ratios, vs_metrics)

    _write_data_sheet(
        wb,
        "Charts_Data",
        cd,
        (
            "core_harmonic_energy_ratio",
            "core_residual_energy_ratio",
            "core_subbass_energy_ratio",
            "component_harmonic_energy_ratio",
            "component_inharmonic_energy_ratio",
            "component_subbass_energy_ratio",
        ),
        (
            "spectral_body_thickness_index",
            "body_weighted_effective_density",
            "low_mid_energy_ratio",
            "harmonic_body_density_normalized",
            "core_residual_energy_ratio",
            "spectral_entropy",
            "salient_harmonic_order_count_up_to_body_ceiling",
            "expected_harmonic_order_count_up_to_body_ceiling",
            "salient_harmonic_coverage_up_to_body_ceiling",
            "final_note_density_salience_weighted",
            "final_note_density_count_based",
            "final_note_density_salience_weighted_norm_for_chart",
            "harmonic_density_component",
            "inharmonic_density_component",
            "subbass_density_component",
            "harmonic_occupancy_ratio",
            "residual_log_frequency_occupancy",
            "effective_partial_density",
            "spectral_body_thickness_index_norm_for_chart",
            "body_weighted_effective_density_norm_for_chart",
            "low_mid_energy_ratio_norm_for_chart",
            "harmonic_body_density_normalized_norm_for_chart",
            "harmonic_occupancy_ratio_norm_for_chart",
            "residual_log_frequency_occupancy_norm_for_chart",
            "core_residual_energy_ratio_norm_for_chart",
            "spectral_entropy_norm_for_chart",
            "effective_partial_density_norm_for_chart",
            "density_weighted_sum",
            "density_metric_raw",
            "density_weighted_sum_norm_for_chart",
            "density_metric_raw_norm_for_chart",
            "weighted_harmonic_density_contribution",
            "weighted_inharmonic_density_contribution",
            "weighted_subbass_density_contribution",
        ),
    )

    _write_data_sheet(
        wb,
        "Legacy_Compatibility",
        legacy_df,
        tuple(),
        (
            "Combined Density Metric",
            "Weighted Combined Metric",
            "Total Metric",
            "density_weighted_sum_cdm_mean",
        ),
    )

    _write_data_sheet(
        wb,
        "Analysis_Settings_By_Note",
        settings_by_note,
        tuple(),
        (
            "f0_used_for_density_hz",
            "f0_used_for_density_source",
            "acoustic_f0_status",
            "f0_epistemic_status",
            "f0_validation_mode",
            "nominal_prior_hz",
            "f0_candidate_hz",
            "f0_deviation_cents",
            "valid_for_primary_statistics",
            "is_primary_comparable_profile",
            "analysis_parameter_profile_id",
            "qc_status",
            "sethares_value_status",
            "sethares_curve_status",
            "harmonic_tolerance_hz",
            "density_frequency_ceiling_hz",
        ),
    )

    # Metadata sheet (worksheet AutoFilter only; no formal Table)
    meta_df_out = _sanitize_dataframe_columns(meta_df)
    mws = wb.create_sheet("Metadata")
    for row in dataframe_to_rows(meta_df_out, index=False, header=True):
        mws.append(row)
    _style_header_row(mws, 1, max(1, meta_df_out.shape[1]))
    mws.freeze_panes = "A2"
    apply_simple_autofilter(mws)
    _autosize_columns(mws)

    if not no_charts:
        try:
            _dashboard_charts(wb, dash, cd, chart_anchor)
        except Exception as e:  # noqa: BLE001
            warnings.append(f"Dashboard charts could not be fully created ({e}); data sheets are still valid.")
            dash.cell(chart_anchor, 1, f"Chart generation note: {e}")

    # Validation chart/table on dashboard
    last_row = dash.max_row + 3
    dash.cell(last_row, 1, "Validation snapshot (from Spectral_Density_Metrics)").font = SUBHEADER_FONT
    tbl_r = last_row + 1
    headers = (
        "Note",
        "f0_fit_accepted",
        "acoustic_validation_status",
        "debug_counts_invariant_status",
        "harmonic_alignment_status",
    )
    for i, h in enumerate(headers, start=1):
        dash.cell(tbl_r, i, h)
        dash.cell(tbl_r, i).fill = HEADER_FILL
        dash.cell(tbl_r, i).font = HEADER_FONT
    sdm = wb["Spectral_Density_Metrics"]
    col_map = {}
    for c in range(1, sdm.max_column + 1):
        val = sdm.cell(1, c).value
        if val in headers:
            col_map[str(val)] = c
    for ridx, (_, srow) in enumerate(sd.iterrows(), start=1):
        for j, h in enumerate(headers, start=1):
            ci = col_map.get(h)
            if ci:
                dash.cell(tbl_r + ridx, j, sdm.cell(ridx + 1, ci).value)

    wb.save(output)
    return warnings


def export_research_workbook(
    input_path: str | Path,
    output_path: str | Path | None = None,
    *,
    overwrite: bool = False,
    no_charts: bool = False,
    instrument: Optional[str] = None,
    dynamic: Optional[str] = None,
    force_metadata: bool = False,
    research_metadata: Optional[ResearchExportMetadata] = None,
    include_legacy_cdm_mean: bool = False,
) -> Path:
    """
    Build ``compiled_density_metrics_research.xlsx`` from a compiled workbook.

    Parameters
    ----------
    input_path:
        Path to ``compiled_density_metrics.xlsx``.
    output_path:
        If ``None``, writes ``compiled_density_metrics_research.xlsx`` next to the input.
    overwrite:
        If ``True``, replace an existing output file.
    no_charts:
        If ``True``, omit chart objects from the Dashboard sheet.
    instrument, dynamic, force_metadata:
        Optional CLI-style metadata (ignored if ``research_metadata`` is passed explicitly).
    research_metadata:
        If set, overrides the individual metadata keyword arguments.

    Returns
    -------
    Path
        Resolved path to the written research workbook.

    Raises
    ------
    FileNotFoundError
        If ``input_path`` is not an existing file.
    FileExistsError
        If the output exists and ``overwrite`` is ``False``.
    ValueError
        If the workbook cannot be read or merged (e.g. missing ``Density_Metrics``).
    """
    src = Path(input_path).expanduser().resolve()
    if output_path is None:
        out = src.parent / "compiled_density_metrics_research.xlsx"
    else:
        out = Path(output_path).expanduser().resolve()
    if research_metadata is not None:
        meta = research_metadata
    else:
        meta = ResearchExportMetadata(
            instrument=instrument.strip() if instrument else None,
            dynamic=dynamic.strip() if dynamic else None,
            force_metadata=force_metadata,
        )
    warns = build_workbook(
        src,
        out,
        no_charts=no_charts,
        overwrite=overwrite,
        research_metadata=meta,
        include_legacy_cdm_mean=include_legacy_cdm_mean,
    )
    for w in warns:
        print(f"WARNING: {w}", file=sys.stderr)
    return out


def main(argv: Optional[Sequence[str]] = None) -> int:
    p = argparse.ArgumentParser(description="Export reduced research workbook from compiled_density_metrics.xlsx")
    p.add_argument("input", type=Path, help="Path to compiled_density_metrics.xlsx")
    p.add_argument("--output", type=Path, default=None, help="Output path (default: sibling compiled_density_metrics_research.xlsx)")
    p.add_argument("--no-charts", action="store_true", help="Skip chart objects on Dashboard")
    p.add_argument("--overwrite", action="store_true", help="Overwrite existing output file")
    p.add_argument("--instrument", type=str, default=None, help="Override Instrument for all rows (see --force-metadata)")
    p.add_argument("--dynamic", type=str, default=None, help="Override Dynamic for all rows (see --force-metadata)")
    p.add_argument(
        "--include-legacy-cdm-mean",
        action="store_true",
        help="Include deprecated density_weighted_sum_cdm_mean editorial blend column",
    )
    p.add_argument(
        "--force-metadata",
        action="store_true",
        help="When set with --instrument/--dynamic, replace non-empty workbook values too",
    )
    args = p.parse_args(list(argv) if argv is not None else None)

    src = args.input.expanduser().resolve()
    out_arg = args.output
    if out_arg is None:
        out_resolved: Path | None = None
    else:
        out_resolved = out_arg.expanduser().resolve()

    try:
        out = export_research_workbook(
            src,
            out_resolved,
            overwrite=args.overwrite,
            no_charts=args.no_charts,
            instrument=args.instrument,
            dynamic=args.dynamic,
            force_metadata=args.force_metadata,
            include_legacy_cdm_mean=args.include_legacy_cdm_mean,
        )
    except FileNotFoundError as e:
        print(str(e), file=sys.stderr)
        return 1
    except FileExistsError as e:
        print(str(e), file=sys.stderr)
        return 1
    except ValueError as e:
        print(str(e), file=sys.stderr)
        return 1

    print(f"Wrote: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
