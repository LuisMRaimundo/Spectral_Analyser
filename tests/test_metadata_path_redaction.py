# -*- coding: utf-8 -*-
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from compile_metrics import validate_compiled_density_workbook
from metadata_sanitizer import (
    REDACT_TOKEN,
    enrich_and_redact_batch_audio_result,
    sanitize_dataframe_for_publication,
    sanitize_metadata_dict,
    sanitize_excel_workbook_for_publication,
    string_fails_publication_scan,
)


def test_a_windows_path_detected() -> None:
    s = r"C:\Users\lmr20\Desktop\teste 2\analysis_results"
    assert string_fails_publication_scan(s)


def test_b_posix_path_detected() -> None:
    assert string_fails_publication_scan("/home/user/project/analysis_results")
    assert string_fails_publication_scan("/Users/jane/Desktop/x")
    assert string_fails_publication_scan("/mnt/data/analysis_results")


def test_c_basename_safe() -> None:
    assert not string_fails_publication_scan("IOWA_Flute.mf_A4_Sustains.aif")


def test_d_dict_redacted() -> None:
    d = {"folder_path": r"C:\Users\lmr20\Desktop\teste 2"}
    out = sanitize_metadata_dict(d)
    assert out["folder_path"] == REDACT_TOKEN


def test_e_dataframe_column_redacted() -> None:
    df = pd.DataFrame({"folder_path": [r"C:\Users\x\y"], "Note": ["A4"]})
    out = sanitize_dataframe_for_publication(df)
    assert str(out["folder_path"].iloc[0]) == REDACT_TOKEN
    assert out["Note"].iloc[0] == "A4"


def test_f_workbook_scan_and_sanitize(tmp_path: Path) -> None:
    from metadata_sanitizer import list_publication_path_violations_in_excel

    wb = Workbook()
    ws = wb.active
    ws.title = "Analysis_Metadata"
    ws["A1"] = "Parameter"
    ws["B1"] = "Value"
    ws["A2"] = "x"
    ws["B2"] = r"C:\Users\evil\secret"
    p = tmp_path / "bad.xlsx"
    wb.save(p)

    assert list_publication_path_violations_in_excel(p)
    sanitize_excel_workbook_for_publication(p)
    assert not list_publication_path_violations_in_excel(p)


def test_g_compiled_validator_flags_path(tmp_path: Path) -> None:
    """Synthetic compiled_density_metrics.xlsx with a local path must fail validation."""
    dm = pd.DataFrame(
        [
            {
                "Note": "A4",
                "effective_partial_density": 1.5,
                "harmonic_energy_sum": 1.0,
                "inharmonic_energy_sum": 0.2,
                "subbass_energy_sum": 0.1,
                "total_component_energy": 1.3,
                "harmonic_energy_ratio": 0.7,
                "inharmonic_energy_ratio": 0.2,
                "subbass_energy_ratio": 0.1,
                "harmonic_order_count": 5,
                "spectral_entropy": 0.5,
            }
        ]
    )
    am = pd.DataFrame([{"Parameter": "folder_path", "Value": r"C:\Users\x\Desktop\secret"}])
    outp = tmp_path / "dirty_compiled.xlsx"
    with pd.ExcelWriter(outp, engine="openpyxl") as w:
        dm.to_excel(w, sheet_name="Density_Metrics", index=False)
        am.to_excel(w, sheet_name="Analysis_Metadata", index=False)
    errs = validate_compiled_density_workbook(outp)
    assert errs, "expected validation errors for forbidden path text"
    assert any("forbidden" in e.lower() for e in errs)


def test_h_spectral_workbook_sanitised(tmp_path: Path) -> None:
    """Per-note style workbook: path in Analysis_Metadata is removed by sanitizer."""
    from metadata_sanitizer import list_publication_path_violations_in_excel

    wb = Workbook()
    ws = wb.active
    ws.title = "Analysis_Metadata"
    ws["A1"] = "Parameter"
    ws["B1"] = "Value"
    ws["A2"] = "folder_path"
    ws["B2"] = r"C:\Users\lmr20\Desktop\analysis_results\file\note"
    p = tmp_path / "spectral_analysis.xlsx"
    wb.save(p)
    assert list_publication_path_violations_in_excel(p)
    sanitize_excel_workbook_for_publication(p)
    assert not list_publication_path_violations_in_excel(p)


def test_i_batch_enrich_redacts(tmp_path: Path) -> None:
    p = tmp_path / "dummy.wav"
    p.write_bytes(b"RIFF....")
    row = {"file_path": str(p), "file_name": p.name, "output_dir": str(tmp_path), "success": True}
    out = enrich_and_redact_batch_audio_result(row, p, "A4")
    assert out["file_path"] == REDACT_TOKEN
    assert out["output_dir"] == REDACT_TOKEN
    assert "public_audio_id" in out
