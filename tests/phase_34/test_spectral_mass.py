"""F-061 spectral_mass: derived-column axioms, export placement, formatting."""
from __future__ import annotations

import json
from pathlib import Path

import numpy as np
import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from metric_contract import get_metric_definition
from metric_formula_versions import SPECTRAL_MASS_FORMULA_VERSION, build_column_registry
from tests.phase_33.test_column_formula_versions import (
    test_every_exported_column_has_formula_stamp,
    test_new_column_without_stamp_is_rejected,
)
from tools.backfill_spectral_mass import backfill_spectral_mass
from tools.export_research_density_workbook import export_research_workbook
from tools.spectral_mass import (
    MASS_LEVEL_EXPONENT,
    SPECTRAL_MASS_COLUMN,
    SPECTRAL_MASS_COUNT_COLUMN,
    SPECTRAL_MASS_DATA_BAR_COLOR,
    SPECTRAL_MASS_FORMULA_VERSION,
    add_spectral_mass_column,
    compartment_count,
    compartment_count_contribution,
    compute_spectral_mass,
    place_spectral_mass_right_of_ewsd,
)

GOLDEN_PATH = Path(__file__).parent / "golden" / "clarinet_spectral_mass.json"


def _golden() -> dict:
    return json.loads(GOLDEN_PATH.read_text(encoding="utf-8"))


def _acd_frame(rows: list[dict]) -> pd.DataFrame:
    return pd.DataFrame(rows)


def _mass_from_golden_row(row: dict) -> tuple[float, float]:
    return compute_spectral_mass(
        {
            "harmonic": row["ACD_D0_harmonic"],
            "inharmonic": row["ACD_D0_inharmonic"],
            "subbass": row["ACD_D0_subbass"],
        },
        {
            "harmonic": row["ACD_D1_harmonic"],
            "inharmonic": row["ACD_D1_inharmonic"],
            "subbass": row["ACD_D1_subbass"],
        },
        {
            "harmonic": row["ACD_r_harmonic"],
            "inharmonic": row["ACD_r_inharmonic"],
            "subbass": row["ACD_r_subbass"],
        },
        row["ACD_energy_total"],
        status=row["ACD_status"],
    )


def _h_only(d0: float, d1: float, e_total: float, *, lam: float | None = None):
    return compute_spectral_mass(
        {"harmonic": d0, "inharmonic": 1.0, "subbass": 1.0},
        {"harmonic": d1, "inharmonic": 1.0, "subbass": 1.0},
        {"harmonic": 1.0, "inharmonic": 0.0, "subbass": 0.0},
        e_total,
        lam=lam,
    )


def test_extensivity_doubling_counts_at_fixed_lam_doubles_mass() -> None:
    d0, d1, lam, energy = 8.0, 2.0, 400.0, 800.0
    mass, _ = _h_only(d0, d1, energy, lam=lam)
    doubled, _ = _h_only(2.0 * d0, 2.0 * d1, energy, lam=lam)
    assert doubled == pytest.approx(2.0 * mass, abs=1e-9)


def test_bounded_level_lam_times_ten_scales_by_exponent() -> None:
    d0, d1, lam, energy = 6.0, 3.0, 250.0, 750.0
    mass, _ = _h_only(d0, d1, energy, lam=lam)
    scaled, _ = _h_only(d0, d1, energy, lam=10.0 * lam)
    assert scaled == pytest.approx(mass * (10.0 ** MASS_LEVEL_EXPONENT), abs=1e-12)


def test_ordering_fsharp4_above_gsharp6() -> None:
    payload = _golden()
    by_note = {row["Note"]: row for row in payload["notes"]}
    mass_1, _ = _mass_from_golden_row(by_note["F#4"])
    mass_2, _ = _mass_from_golden_row(by_note["G#6"])
    assert mass_1 > mass_2


def test_inharmonic_contribution_capped_by_energy_share() -> None:
    """r_I = 0.01 with many dominated I entities: I share of count <= 2%."""
    d0 = {"harmonic": 10.0, "inharmonic": 200.0, "subbass": 1.0}
    d1 = {"harmonic": 8.0, "inharmonic": 1.2, "subbass": 1.0}
    r = {"harmonic": 0.99, "inharmonic": 0.01, "subbass": 0.0}
    mass, count = compute_spectral_mass(d0, d1, r, e_total=1000.0)
    assert np.isfinite(mass) and np.isfinite(count)
    i_contrib = compartment_count_contribution(d0, d1, r, "inharmonic")
    count_i = compartment_count(d0["inharmonic"], d1["inharmonic"])
    assert i_contrib == pytest.approx(r["inharmonic"] * count_i, abs=1e-12)
    assert i_contrib <= r["inharmonic"] * count_i + 1e-12
    assert i_contrib / count <= 0.02 + 1e-12


def test_nan_propagation_never_zero() -> None:
    base = {
        "ACD_D0_harmonic": 8.0,
        "ACD_D1_harmonic": 2.0,
        "ACD_r_harmonic": 0.9,
        "ACD_D0_inharmonic": 4.0,
        "ACD_D1_inharmonic": 1.5,
        "ACD_r_inharmonic": 0.1,
        "ACD_D0_subbass": 1.0,
        "ACD_D1_subbass": 1.0,
        "ACD_r_subbass": 0.0,
        "ACD_energy_total": 100.0,
        "ACD_status": "ok",
    }
    cases = [
        {**base, "ACD_D0_harmonic": np.nan},
        {**base, "ACD_D1_inharmonic": np.nan},
        {**base, "ACD_r_harmonic": np.nan},
        {**base, "ACD_energy_total": np.nan, "ACD_score": np.nan, "ACD_magnitude_per_component": np.nan},
        {**base, "ACD_status": "empty"},
        {**base, "ACD_status": "ok_with_unused:x"},
    ]
    out = add_spectral_mass_column(_acd_frame(cases))
    values = out["spectral_mass"].to_numpy(dtype=float)
    assert np.isnan(values).all()
    assert not np.any(values == 0.0)
    counts = out["spectral_mass_count"].to_numpy(dtype=float)
    assert np.isnan(counts).all()
    assert not np.any(counts == 0.0)


def test_missing_status_is_nan() -> None:
    frame = _acd_frame(
        [
            {
                "ACD_D0_harmonic": 4.0,
                "ACD_D1_harmonic": 2.0,
                "ACD_r_harmonic": 1.0,
                "ACD_D0_inharmonic": 1.0,
                "ACD_D1_inharmonic": 1.0,
                "ACD_r_inharmonic": 0.0,
                "ACD_D0_subbass": 1.0,
                "ACD_D1_subbass": 1.0,
                "ACD_r_subbass": 0.0,
                "ACD_energy_total": 10.0,
            }
        ]
    )
    out = add_spectral_mass_column(frame)
    assert np.isnan(float(out["spectral_mass"].iloc[0]))


def test_contract_f061_mass_descriptor() -> None:
    definition = get_metric_definition("spectral_mass")
    assert definition is not None
    assert definition.formula_id == "F-061"
    assert definition.ontology_family == "mass_descriptor"
    assert definition.ontology_family not in {"legacy_only", "partial_count_descriptor"}
    assert "presence constitutes richness; loudness modulates it but must not overturn it" in (
        definition.formula
    )
    assert "Cross-compartment contributions are bounded by energy share" in definition.formula
    assert definition.formula_version == "2.0"
    assert SPECTRAL_MASS_FORMULA_VERSION == "2.0"
    assert definition.not_valid_for == (
        "Level-inclusive by design. Valid within level-controlled "
        "corpora (uniform recording conditions). Not valid for comparison across recording "
        "sessions, microphone distances, or gain settings. Decomposes exactly into "
        "spectral_mass_count and a size factor."
    )


def test_column_formula_version_and_phase_33_gate() -> None:
    registry = build_column_registry()
    assert registry["spectral_mass"]["formula_id"] == "F-061"
    assert registry["spectral_mass"]["formula_version"] == SPECTRAL_MASS_FORMULA_VERSION
    assert registry["spectral_mass_count"]["formula_id"] == "F-061"
    test_every_exported_column_has_formula_stamp()
    test_new_column_without_stamp_is_rejected()


def test_clarinet_golden_vectors() -> None:
    payload = _golden()
    assert payload["formula_version"] == "2.0"
    assert "synthetic" in payload["note"]
    assert len(payload["notes"]) == 5
    for row in payload["notes"]:
        mass, count = _mass_from_golden_row(row)
        assert mass == pytest.approx(row["spectral_mass"], abs=1e-12)
        assert count == pytest.approx(row["spectral_mass_count"], abs=1e-12)
        frame = add_spectral_mass_column(_acd_frame([row]))
        assert float(frame["spectral_mass"].iloc[0]) == pytest.approx(
            row["spectral_mass"], abs=1e-12
        )


def _write_per_note_workbook(path: Path, *, note: str = "D3") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    harmonic = pd.DataFrame(
        {
            "Frequency (Hz)": [146.83, 293.66, 440.0],
            "Amplitude_raw": [1.0, 0.7, 0.5],
            "include_for_density": [True, True, True],
        }
    )
    inharmonic = pd.DataFrame(
        {
            "Frequency (Hz)": [220.0, 330.0],
            "Amplitude_raw": [0.15, 0.10],
        }
    )
    subbass = pd.DataFrame(
        {
            "Frequency (Hz)": [55.0],
            "Amplitude_raw": [0.05],
        }
    )
    metrics = pd.DataFrame(
        {
            "Note": [note],
            "weight_function": ["log"],
            "component_harmonic_energy_ratio": [0.80],
            "component_inharmonic_energy_ratio": [0.15],
            "component_subbass_energy_ratio": [0.05],
        }
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        harmonic.to_excel(writer, sheet_name="Harmonic Spectrum", index=False)
        inharmonic.to_excel(writer, sheet_name="Inharmonic Spectrum", index=False)
        subbass.to_excel(writer, sheet_name="Sub-bass band", index=False)
        metrics.to_excel(writer, sheet_name="Metrics", index=False)


def _write_compiled_workbook(path: Path, *, note: str = "D3") -> None:
    density = pd.DataFrame(
        {
            "Note": [note],
            "source_file_name": ["Viola-D3-mf.wav"],
            "density_metric_raw": [0.42],
            "density_metric_normalized": [1.0],
            "harmonic_density_sum": [1.0],
            "inharmonic_density_sum": [0.1],
            "subbass_density_sum": [0.01],
            "component_harmonic_energy_ratio": [0.80],
            "component_inharmonic_energy_ratio": [0.15],
            "component_subbass_energy_ratio": [0.05],
            "density_frequency_ceiling_hz": [20000.0],
            "f0_final_hz": [146.83],
            "f0_source": ["nominal_guided"],
            "f0_final_source": ["nominal_guided"],
            "acoustic_f0_status": ["nominal_guided_acoustically_verified"],
            "f0_fit_accepted": [True],
        }
    )
    meta = pd.DataFrame(
        {
            "analysis_version": ["test"],
            "weight_function": ["log"],
            "density_salience_threshold_db": [-60.0],
            "density_frequency_ceiling_hz": [20000.0],
        }
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        density.to_excel(writer, sheet_name="Density_Metrics", index=False)
        meta.to_excel(writer, sheet_name="Analysis_Metadata", index=False)


def _headers(ws) -> list:
    return [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]


def _data_bar_rules_for_column(ws, column_name: str) -> list:
    headers = _headers(ws)
    if column_name not in headers:
        return []
    letter = get_column_letter(headers.index(column_name) + 1)
    found = []
    for cf_range in ws.conditional_formatting:
        if letter not in str(cf_range):
            continue
        for rule in ws.conditional_formatting._cf_rules[cf_range]:
            if getattr(rule, "type", None) == "dataBar" and rule.dataBar is not None:
                found.append(rule)
    return found


def test_research_export_column_order_and_blue_data_bar(tmp_path: Path) -> None:
    _write_per_note_workbook(tmp_path / "D3" / "spectral_analysis.xlsx")
    compiled = tmp_path / "compiled_density_metrics.xlsx"
    _write_compiled_workbook(compiled)
    output = tmp_path / "compiled_density_metrics_research.xlsx"
    export_research_workbook(
        input_path=compiled,
        output_path=output,
        overwrite=True,
        no_charts=True,
        include_ewsd=True,
    )
    wb = load_workbook(output)
    ws = wb["Spectral_Density_Metrics"]
    headers = _headers(ws)
    assert SPECTRAL_MASS_COLUMN in headers
    assert SPECTRAL_MASS_COUNT_COLUMN in headers
    ewsd_i = headers.index("EWSD_score_acoustic_balanced")
    assert headers[ewsd_i + 1] == SPECTRAL_MASS_COLUMN
    assert headers[ewsd_i + 2] == SPECTRAL_MASS_COUNT_COLUMN
    rules = _data_bar_rules_for_column(ws, SPECTRAL_MASS_COLUMN)
    assert rules, "expected blue data bar on spectral_mass"
    color = str(rules[0].dataBar.color.rgb).upper()
    assert color.endswith("4472C4")
    assert SPECTRAL_MASS_DATA_BAR_COLOR.endswith("4472C4")
    assert "spectral_mass_formula_id" in headers
    assert "spectral_mass_formula_version" in headers


def test_backfill_refuses_workbook_without_compartment_d1(tmp_path: Path) -> None:
    src = tmp_path / "old_research.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Spectral_Density_Metrics"
    ws.append(["Note", "ACD_D0", "ACD_score", "ACD_magnitude_per_component", "ACD_status"])
    ws.append(["F#4", 24.9, 1.15, 9208.02, "ok"])
    wb.save(src)
    with pytest.raises(ValueError, match="ACD_D1_harmonic"):
        backfill_spectral_mass(src)


def test_backfill_writes_sidecar_with_placement_and_bars(tmp_path: Path) -> None:
    src = tmp_path / "compiled_density_metrics_research.xlsx"
    frame = place_spectral_mass_right_of_ewsd(
        add_spectral_mass_column(
            pd.DataFrame(
                {
                    "Note": ["F#4"],
                    "EWSD_score_acoustic_balanced": [12.0],
                    "ACD_D0_harmonic": [24.0],
                    "ACD_D1_harmonic": [1.10],
                    "ACD_r_harmonic": [0.96],
                    "ACD_D0_inharmonic": [40.0],
                    "ACD_D1_inharmonic": [2.5],
                    "ACD_r_inharmonic": [0.03],
                    "ACD_D0_subbass": [15.0],
                    "ACD_D1_subbass": [1.8],
                    "ACD_r_subbass": [0.01],
                    "ACD_energy_total": [10589.223],
                    "ACD_status": ["ok"],
                    "trailing": ["keep"],
                }
            )
        )
    )
    # Write a pre-mass workbook (ACD present, mass absent).
    raw = frame.drop(
        columns=[
            c
            for c in frame.columns
            if c.startswith("spectral_mass")
        ]
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Spectral_Density_Metrics"
    ws.append(list(raw.columns))
    ws.append([raw.iloc[0][c] for c in raw.columns])
    wb.create_sheet("README").append(["old"])
    wb.save(src)

    dest = backfill_spectral_mass(src)
    assert dest.name == "compiled_density_metrics_research_massfilled.xlsx"
    assert dest.exists()
    assert src.exists()
    with pytest.raises(FileExistsError):
        backfill_spectral_mass(src)

    filled = load_workbook(dest)
    ws = filled["Spectral_Density_Metrics"]
    headers = _headers(ws)
    ewsd_i = headers.index("EWSD_score_acoustic_balanced")
    assert headers[ewsd_i + 1] == SPECTRAL_MASS_COLUMN
    assert headers[ewsd_i + 2] == SPECTRAL_MASS_COUNT_COLUMN
    assert float(ws.cell(2, ewsd_i + 2).value) == pytest.approx(16.530465555316866, abs=1e-9)
    assert _data_bar_rules_for_column(ws, SPECTRAL_MASS_COLUMN)
    assert "README" in filled.sheetnames
