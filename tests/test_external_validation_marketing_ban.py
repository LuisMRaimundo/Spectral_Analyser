# -*- coding: utf-8 -*-
"""Regression: active tree must not advertise external symbolic-engine or MCP-based 'validation' marketing."""

from __future__ import annotations

import json
import sys
from io import StringIO
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parent.parent
THIS_FILE = Path(__file__).resolve()
ANALYZE_DIR = ROOT / "audio_analysis"
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ANALYZE_DIR))

from super_audio_analyzer import (  # noqa: E402
    _METRICS_SUMMARY_NOTES,
    _write_metrics_summary_mapping,
)


def _forbidden_substrings() -> tuple[str, ...]:
    # Built from code points so this module's source does not embed vendor trademarks.
    w = "".join(chr(c) for c in (119, 111, 108, 102, 114, 97, 109))
    use_mcp_tool = "".join(chr(c) for c in (117, 115, 101, 32, 109, 99, 112, 32, 116, 111, 111, 108))
    return (
        w,
        w + " alpha",
        "mcp_" + w,
        "query_" + w,
        use_mcp_tool,
        w + "-alpha",
    )


def _legacy_json_key_reasoning_attr() -> str:
    return "".join(
        chr(c)
        for c in (
            119,
            111,
            108,
            102,
            114,
            97,
            109,
            95,
            114,
            101,
            97,
            115,
            111,
            110,
            105,
            110,
            103,
        )
    )


def _legacy_json_key_available_attr() -> str:
    return "".join(
        chr(c)
        for c in (
            119,
            111,
            108,
            102,
            114,
            97,
            109,
            95,
            97,
            118,
            97,
            105,
            108,
            97,
            98,
            108,
            101,
        )
    )


def _assert_text_clean(label: str, text: str) -> None:
    low = text.lower()
    for frag in _forbidden_substrings():
        assert frag.lower() not in low, f"{label}: forbidden substring {frag!r}"


def _path_is_active_scan_target(path: Path) -> bool:
    parts = path.parts
    if "Backup" in parts:
        return False
    if ".venv" in parts:
        return False
    if path.suffix == ".egg-info" or any(part.endswith(".egg-info") for part in parts):
        return False
    if ".git" in parts:
        return False
    if "__pycache__" in parts:
        return False
    if ".pytest_cache" in parts:
        return False
    if "history" in parts and "Backup" in parts:
        return False
    return True


def _iter_scannable_files() -> list[Path]:
    exts = {".py", ".md", ".txt", ".json", ".yml", ".yaml", ".csv", ".toml"}
    out: list[Path] = []
    for p in ROOT.rglob("*"):
        if not p.is_file():
            continue
        if p.resolve() == THIS_FILE:
            continue
        if not _path_is_active_scan_target(p):
            continue
        if p.suffix.lower() not in exts:
            continue
        out.append(p)
    return out


def test_active_repo_text_files_exclude_backup_have_no_external_engine_marketing() -> None:
    failures: list[str] = []
    forbidden = _forbidden_substrings()
    for path in _iter_scannable_files():
        try:
            text = path.read_text(encoding="utf-8", errors="replace")
        except OSError as e:
            failures.append(f"{path}: read error {e}")
            continue
        low = text.lower()
        for frag in forbidden:
            if frag.lower() in low:
                failures.append(f"{path}: contains {frag!r}")
                break
    assert not failures, (
        "External symbolic-engine / MCP marketing strings in active files:\n" + "\n".join(failures[:80])
    )


def test_super_metrics_summary_template_has_no_external_engine_marketing() -> None:
    data = {
        "harmonic_energy_percentage": 71.0,
        "inharmonic_energy_percentage": 29.0,
        "harmonic_energy_percentage_peak_based": 68.0,
        "inharmonic_energy_percentage_peak_based": 32.0,
        "harmonic_density": 0.5,
        "dissonance_curve": {0.0: 0.1, 0.1: 0.2},
    }
    buf = StringIO()
    buf.write(_METRICS_SUMMARY_NOTES)
    _write_metrics_summary_mapping(buf, data, str)
    _assert_text_clean("metrics summary mapping", buf.getvalue())


def test_batch_super_analysis_json_samples_clean() -> None:
    samples = sorted((ANALYZE_DIR / "batch_results").glob("*/super_analysis_results.json"))
    if not samples:
        pytest.skip(
            "Optional batch_results fixtures (super_analysis_results.json) are absent in this "
            "curated GitHub export; they are intentionally omitted from the curated export."
        )
    k_reason = _legacy_json_key_reasoning_attr()
    k_avail = _legacy_json_key_available_attr()
    for p in samples:
        data = json.loads(p.read_text(encoding="utf-8"))
        dumped = json.dumps(data, ensure_ascii=False)
        _assert_text_clean(str(p), dumped)
        assert "physical_validation" not in data
        assert k_reason not in dumped
        assert k_avail not in dumped


def test_batch_metrics_summary_txt_samples_clean() -> None:
    txts = sorted((ANALYZE_DIR / "batch_results").glob("*/metrics_summary.txt"))
    if not txts:
        pytest.skip(
            "Optional batch_results fixtures (metrics_summary.txt) are absent in this curated "
            "GitHub export; they are intentionally omitted from the curated export."
        )
    for p in txts:
        _assert_text_clean(str(p), p.read_text(encoding="utf-8", errors="replace"))


@pytest.mark.parametrize(
    "pattern",
    ["*.xlsx", "*.xls"],
)
def test_workbook_metadata_cells_if_present_have_no_external_engine_marketing(pattern: str) -> None:
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError:  # pragma: no cover
        pytest.skip("openpyxl not installed")
    paths = [p for p in ROOT.rglob(pattern) if _path_is_active_scan_target(p)]
    if not paths:
        return
    for wb_path in paths:
        wb = load_workbook(wb_path, read_only=True, data_only=True)
        try:
            chunks: list[str] = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        if isinstance(cell, str):
                            chunks.append(cell)
            blob = "\n".join(chunks)
            _assert_text_clean(str(wb_path), blob)
        finally:
            wb.close()
