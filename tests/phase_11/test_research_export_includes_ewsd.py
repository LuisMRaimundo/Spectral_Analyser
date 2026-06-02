from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from tools import export_research_density_workbook as research_export


def _write_per_note_workbook(path: Path, *, note: str = "D3") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    harmonic = pd.DataFrame(
        {
            "Frequency (Hz)": [146.83, 293.66, 440.0],
            "Amplitude_raw": [1.0, 0.7, 0.5],
            "Magnitude (dB)": [0.0, -3.0, -6.0],
            "include_for_density": [True, True, True],
        }
    )
    inharmonic = pd.DataFrame(
        {
            "Frequency (Hz)": [220.0, 330.0],
            "Amplitude_raw": [0.15, 0.10],
            "Magnitude (dB)": [-16.0, -20.0],
        }
    )
    subbass = pd.DataFrame(
        {
            "Frequency (Hz)": [55.0],
            "Amplitude_raw": [0.05],
            "Magnitude (dB)": [-26.0],
        }
    )
    metrics = pd.DataFrame(
        {
            "Note": [note],
            "weight_function": ["log"],
            "pure_observation_w_h": [0.80],
            "pure_observation_w_i": [0.15],
            "pure_observation_w_s": [0.05],
            "component_harmonic_energy_ratio": [0.80],
            "component_inharmonic_energy_ratio": [0.15],
            "component_subbass_energy_ratio": [0.05],
        }
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        harmonic.to_excel(writer, sheet_name="Harmonic Spectrum", index=False)
        inharmonic.to_excel(writer, sheet_name="Inharmonic Spectrum", index=False)
        subbass.to_excel(writer, sheet_name="Sub-bass band", index=False)
        metrics.to_excel(writer, sheet_name="Metrics", index=False)


def _write_compiled_workbook(path: Path, *, note: str = "D3") -> None:
    density = pd.DataFrame(
        {
            "Note": [note],
            "source_file_name": ["Viola-D3-mf.wav"],
            "density_metric_raw": [0.42],
            "density_metric_normalized": [1.0],
            "harmonic_density_sum": [1.0],
            "inharmonic_density_sum": [0.1],
            "subbass_density_sum": [0.01],
            "component_harmonic_energy_ratio": [0.80],
            "component_inharmonic_energy_ratio": [0.15],
            "component_subbass_energy_ratio": [0.05],
            "density_frequency_ceiling_hz": [20000.0],
            "f0_final_hz": [146.83],
            "f0_source": ["nominal_guided"],
            "f0_final_source": ["nominal_guided"],
            "acoustic_f0_status": ["nominal_guided_acoustically_verified"],
            "f0_fit_accepted": [True],
        }
    )
    meta = pd.DataFrame(
        {
            "analysis_version": ["test"],
            "weight_function": ["log"],
            "density_salience_threshold_db": [-60.0],
            "density_frequency_ceiling_hz": [20000.0],
        }
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        density.to_excel(writer, sheet_name="Density_Metrics", index=False)
        meta.to_excel(writer, sheet_name="Analysis_Metadata", index=False)


def test_research_export_includes_ewsd_stage3_columns(tmp_path: Path) -> None:
    _write_per_note_workbook(tmp_path / "D3" / "spectral_analysis.xlsx", note="D3")
    compiled = tmp_path / "compiled_density_metrics.xlsx"
    _write_compiled_workbook(compiled, note="D3")
    output = tmp_path / "compiled_density_metrics_research.xlsx"

    out = research_export.export_research_workbook(
        input_path=compiled,
        output_path=output,
        overwrite=True,
        no_charts=True,
        include_ewsd=True,
    )
    assert out.exists()

    sdm = pd.read_excel(out, sheet_name="Spectral_Density_Metrics")
    row = sdm.iloc[0]
    assert str(row["Note"]).strip() == "D3"
    assert pd.notna(row["EWSD_score_total"])
    assert pd.notna(row["EWSD_score_acoustic_balanced"])
    assert float(row["EWSD_score_total"]) >= 0.0
    assert float(row["EWSD_score_acoustic_balanced"]) >= 0.0
    assert str(row["ewsd_mode"]).strip() == "individual_exact"
    assert bool(row["ewsd_primary_analysis_eligible"]) is True
    assert str(row["ewsd_merge_status"]).strip() == "merged_individual_exact"
    assert str(row["ewsd_stage3_version"]).strip().startswith("EWSD-R v18")
    assert "EWSD_score_acoustic_balanced_ci_low" in sdm.columns
    assert "EWSD_score_acoustic_balanced_ci_high" in sdm.columns
    assert pd.notna(row["EWSD_score_acoustic_balanced_ci_low"])
    assert pd.notna(row["EWSD_score_acoustic_balanced_ci_high"])
    assert float(row["EWSD_score_acoustic_balanced_ci_low"]) <= float(row["EWSD_score_acoustic_balanced"])
    assert float(row["EWSD_score_acoustic_balanced"]) <= float(row["EWSD_score_acoustic_balanced_ci_high"])
    assert str(row.get("ewsd_uncertainty_sources", "")).strip() in {"partials+ratios", "partials"}


def test_research_export_ewsd_acoustic_balanced_has_red_data_bar(tmp_path: Path) -> None:
    _write_per_note_workbook(tmp_path / "D3" / "spectral_analysis.xlsx", note="D3")
    compiled = tmp_path / "compiled_density_metrics.xlsx"
    _write_compiled_workbook(compiled, note="D3")
    output = tmp_path / "compiled_density_metrics_research.xlsx"

    research_export.export_research_workbook(
        input_path=compiled,
        output_path=output,
        overwrite=True,
        no_charts=True,
        include_ewsd=True,
    )

    wb = load_workbook(output)
    ws = wb["Spectral_Density_Metrics"]
    assert "EWSD_score_acoustic_balanced" in {
        ws.cell(1, c).value for c in range(1, ws.max_column + 1)
    }
    data_bar_rules = [
        rule
        for cf in ws.conditional_formatting
        for rule in ws.conditional_formatting._cf_rules[cf]
        if getattr(rule, "type", None) == "dataBar" and rule.dataBar is not None
    ]
    assert data_bar_rules, "expected red data bar on EWSD_score_acoustic_balanced"
    assert str(data_bar_rules[0].dataBar.color.rgb).upper().endswith("C00000")


def test_research_export_ewsd_skipped_without_per_note_workbooks(tmp_path: Path) -> None:
    compiled = tmp_path / "compiled_density_metrics.xlsx"
    _write_compiled_workbook(compiled, note="D3")
    output = tmp_path / "compiled_density_metrics_research.xlsx"

    research_export.export_research_workbook(
        input_path=compiled,
        output_path=output,
        overwrite=True,
        no_charts=True,
        include_ewsd=True,
    )
    sdm = pd.read_excel(output, sheet_name="Spectral_Density_Metrics")
    assert "EWSD_score_total" in sdm.columns
    assert pd.isna(sdm.iloc[0]["EWSD_score_total"])
    assert str(sdm.iloc[0]["ewsd_merge_status"]).strip() == "no_per_note_workbooks_found"
