"""Tests for ``tools/export_research_density_workbook.py``."""

from __future__ import annotations

import hashlib
import subprocess
import sys
import zipfile
from pathlib import Path

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook

REPO = Path(__file__).resolve().parents[1]


def test_research_workbook_has_no_xl_table_parts_and_loads(tmp_path: Path) -> None:
    """Formal Excel Table parts (xl/tables/table*.xml) trigger Excel repair; export must not create them."""
    src = tmp_path / "compiled_density_metrics.xlsx"
    dst = tmp_path / "compiled_density_metrics_research.xlsx"
    _write_minimal_compiled_workbook(src)
    proc = _run_export(src, dst)
    assert proc.returncode == 0, proc.stderr + proc.stdout
    with zipfile.ZipFile(dst) as z:
        table_parts = [n for n in z.namelist() if n.startswith("xl/tables/")]
        assert table_parts == []
    wb = load_workbook(dst)
    assert "Dashboard" in wb.sheetnames
    assert "Spectral_Density_Metrics" in wb.sheetnames
    sdm = wb["Spectral_Density_Metrics"]
    assert sdm.auto_filter.ref is not None
    assert str(sdm.auto_filter.ref).startswith("A1:")


def test_sanitize_dataframe_columns_blank_and_duplicates() -> None:
    from tools.export_research_density_workbook import _sanitize_dataframe_columns

    df = pd.DataFrame([[1, 2, 3]], columns=["", "Note", "Note"])
    out = _sanitize_dataframe_columns(df)
    assert list(out.columns) == ["column_1", "Note", "Note_2"]


def _write_minimal_compiled_workbook(path: Path, *, sparse: bool = False) -> None:
    """Minimal ``compiled_density_metrics``-style workbook for export tests."""
    dm = pd.DataFrame(
        {
            "Note": ["A4", "D3"],
            "density_metric_raw": [1.0, 1.0],
            "density_metric_normalized": [1.0, 1.0],
            "weighted_harmonic_density_contribution": [0.2, 0.8],
            "weighted_inharmonic_density_contribution": [0.3, 0.15],
            "weighted_subbass_density_contribution": [0.5, 0.05],
            "component_harmonic_energy_ratio": [0.2, 0.8],
            "component_inharmonic_energy_ratio": [0.3, 0.15],
            "component_subbass_energy_ratio": [0.5, 0.05],
            "Harmonic Partials sum": [1.0, 10.0],
            "Inharmonic Partials sum": [2.0, 20.0],
            "Sub-bass sum": [3.0, 30.0],
            "Total sum": [6.0, 60.0],
            "source_file_name": ["Clarinet_A4_pp.wav", "Bassoon_D3_mf.wav"],
            "weight_function": ["linear", "linear"],
            "density_weighted_sum": [1.5, 2.5],
            "density_log_weighted": [0.1, 0.2],
            "effective_partial_density": [0.4, 0.5],
            "spectral_entropy": [0.7, 0.8],
        }
    )
    if sparse:
        dm = pd.DataFrame({"Note": ["C4"], "density_metric_raw": [0.5]})

    am = pd.DataFrame(
        [
            ("pipeline_contract_version", "test-contract"),
            ("ANALYSIS_SCHEMA_VERSION", "99"),
            ("weight_function", "linear"),
        ],
        columns=["Parameter", "Value"],
    )

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        dm.to_excel(writer, sheet_name="Density_Metrics", index=False)
        am.to_excel(writer, sheet_name="Analysis_Metadata", index=False)


def _run_export(inp: Path, out: Path, extra: list[str] | None = None) -> subprocess.CompletedProcess[str]:
    cmd = [
        sys.executable,
        str(REPO / "tools" / "export_research_density_workbook.py"),
        str(inp),
        "--output",
        str(out),
        "--overwrite",
    ]
    if extra:
        cmd.extend(extra)
    return subprocess.run(cmd, capture_output=True, text=True, check=False)


def test_export_creates_research_workbook(tmp_path: Path) -> None:
    src = tmp_path / "compiled_density_metrics.xlsx"
    dst = tmp_path / "compiled_density_metrics_research.xlsx"
    _write_minimal_compiled_workbook(src)
    proc = _run_export(src, dst)
    assert proc.returncode == 0, proc.stderr + proc.stdout
    assert dst.is_file()
    xl = pd.ExcelFile(dst, engine="openpyxl")
    expected = {
        "README",
        "Dashboard",
        "Spectral_Density_Metrics",
        "Component_Balance",
        "Validation_Summary",
        "Charts_Data",
        "Metadata",
    }
    assert set(xl.sheet_names) == expected


def test_spectral_density_metrics_columns(tmp_path: Path) -> None:
    src = tmp_path / "in.xlsx"
    dst = tmp_path / "out.xlsx"
    _write_minimal_compiled_workbook(src)
    assert _run_export(src, dst).returncode == 0
    df = pd.read_excel(dst, sheet_name="Spectral_Density_Metrics", engine="openpyxl")
    for col in (
        "Note",
        "MIDI",
        "density_metric_raw",
        "density_weighted_sum",
        "Total sum",
        "effective_partial_density",
        "spectral_entropy",
    ):
        assert col in df.columns


def test_component_balance_recomputes(tmp_path: Path) -> None:
    src = tmp_path / "in.xlsx"
    dst = tmp_path / "out.xlsx"
    _write_minimal_compiled_workbook(src)
    assert _run_export(src, dst).returncode == 0
    cb = pd.read_excel(dst, sheet_name="Component_Balance", engine="openpyxl")
    wsum = (
        pd.to_numeric(cb["weighted_harmonic_density_contribution"], errors="coerce")
        + pd.to_numeric(cb["weighted_inharmonic_density_contribution"], errors="coerce")
        + pd.to_numeric(cb["weighted_subbass_density_contribution"], errors="coerce")
    )
    assert np.allclose(
        pd.to_numeric(cb["density_metric_raw_recomputed"], errors="coerce"),
        wsum,
        equal_nan=True,
    )
    tsum = (
        pd.to_numeric(cb["harmonic_density_sum"], errors="coerce")
        + pd.to_numeric(cb["inharmonic_density_sum"], errors="coerce")
        + pd.to_numeric(cb["subbass_density_sum"], errors="coerce")
    )
    assert np.allclose(
        pd.to_numeric(cb["total_sum_recomputed"], errors="coerce"),
        tsum,
        equal_nan=True,
    )


def test_source_workbook_not_modified(tmp_path: Path) -> None:
    src = tmp_path / "compiled_density_metrics.xlsx"
    dst = tmp_path / "out.xlsx"
    _write_minimal_compiled_workbook(src)
    h_before = hashlib.sha256(src.read_bytes()).hexdigest()
    assert _run_export(src, dst).returncode == 0
    h_after = hashlib.sha256(src.read_bytes()).hexdigest()
    assert h_before == h_after


def test_sparse_workbook_no_crash_and_readme_warnings(tmp_path: Path) -> None:
    src = tmp_path / "sparse.xlsx"
    dst = tmp_path / "out.xlsx"
    _write_minimal_compiled_workbook(src, sparse=True)
    proc = _run_export(src, dst)
    assert proc.returncode == 0, proc.stderr
    assert "WARNING" in proc.stderr or "Min-max" in proc.stderr or "Component_Balance" in proc.stderr
    readme = pd.read_excel(dst, sheet_name="README", engine="openpyxl", header=None)
    text = "\n".join(str(x) for x in readme.iloc[:, 0].tolist() if pd.notna(x))
    assert "Warnings" in text


def test_charts_data_sorted_by_midi(tmp_path: Path) -> None:
    src = tmp_path / "in.xlsx"
    dst = tmp_path / "out.xlsx"
    _write_minimal_compiled_workbook(src)
    assert _run_export(src, dst).returncode == 0
    cd = pd.read_excel(dst, sheet_name="Charts_Data", engine="openpyxl")
    midi = pd.to_numeric(cd["MIDI"], errors="coerce")
    assert midi.is_monotonic_increasing


def test_no_charts_dashboard_has_no_chart_objects(tmp_path: Path) -> None:
    src = tmp_path / "in.xlsx"
    dst = tmp_path / "out.xlsx"
    _write_minimal_compiled_workbook(src)
    proc = _run_export(src, dst, extra=["--no-charts"])
    assert proc.returncode == 0
    wb = load_workbook(dst)
    assert "Dashboard" in wb.sheetnames
    dash = wb["Dashboard"]
    assert len(getattr(dash, "_charts", [])) == 0


def test_default_output_path_fails_if_exists(tmp_path: Path) -> None:
    src = tmp_path / "compiled_density_metrics.xlsx"
    _write_minimal_compiled_workbook(src)
    default_out = tmp_path / "compiled_density_metrics_research.xlsx"
    default_out.write_text("block")
    proc = subprocess.run(
        [sys.executable, str(REPO / "tools" / "export_research_density_workbook.py"), str(src)],
        capture_output=True,
        text=True,
        check=False,
    )
    assert proc.returncode != 0
    assert "already exists" in proc.stderr.lower() or "already exists" in proc.stdout.lower()


def _write_compiled_with_instrument_dynamic(path: Path) -> None:
    dm = pd.DataFrame(
        {
            "Note": ["A4", "D3"],
            "density_metric_raw": [1.0, 1.0],
            "density_metric_normalized": [1.0, 1.0],
            "weighted_harmonic_density_contribution": [0.2, 0.8],
            "weighted_inharmonic_density_contribution": [0.3, 0.15],
            "weighted_subbass_density_contribution": [0.5, 0.05],
            "component_harmonic_energy_ratio": [0.2, 0.8],
            "component_inharmonic_energy_ratio": [0.3, 0.15],
            "component_subbass_energy_ratio": [0.5, 0.05],
            "Harmonic Partials sum": [1.0, 10.0],
            "Inharmonic Partials sum": [2.0, 20.0],
            "Sub-bass sum": [3.0, 30.0],
            "Total sum": [6.0, 60.0],
            "source_file_name": ["take1.wav", "take2.wav"],
            "Instrument": ["Bassoon", "Bassoon"],
            "Dynamic": ["pp", "pp"],
            "weight_function": ["linear", "linear"],
            "density_weighted_sum": [1.5, 2.5],
            "density_log_weighted": [0.1, 0.2],
            "effective_partial_density": [0.4, 0.5],
            "spectral_entropy": [0.7, 0.8],
        }
    )
    am = pd.DataFrame(
        [
            ("pipeline_contract_version", "test-contract"),
            ("ANALYSIS_SCHEMA_VERSION", "99"),
            ("weight_function", "linear"),
        ],
        columns=["Parameter", "Value"],
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        dm.to_excel(writer, sheet_name="Density_Metrics", index=False)
        am.to_excel(writer, sheet_name="Analysis_Metadata", index=False)


def test_cli_instrument_dynamic_populate_when_absent(tmp_path: Path) -> None:
    src = tmp_path / "compiled_density_metrics.xlsx"
    dst = tmp_path / "out.xlsx"
    _write_minimal_compiled_workbook(src)
    proc = _run_export(
        src,
        dst,
        extra=["--instrument", "Oboe", "--dynamic", "pp"],
    )
    assert proc.returncode == 0, proc.stderr + proc.stdout
    df = pd.read_excel(dst, sheet_name="Spectral_Density_Metrics", engine="openpyxl")
    assert (df["Instrument"].astype(str) == "Oboe").all()
    assert (df["Dynamic"].astype(str) == "pp").all()
    cb = pd.read_excel(dst, sheet_name="Component_Balance", engine="openpyxl")
    assert (cb["Instrument"].astype(str) == "Oboe").all()
    assert (cb["Dynamic"].astype(str) == "pp").all()
    vs = pd.read_excel(dst, sheet_name="Validation_Summary", engine="openpyxl")
    assert (vs["Instrument"].astype(str) == "Oboe").all()


def test_cli_without_force_does_not_override_existing_instrument_dynamic(tmp_path: Path) -> None:
    src = tmp_path / "compiled_density_metrics.xlsx"
    dst = tmp_path / "out.xlsx"
    _write_compiled_with_instrument_dynamic(src)
    proc = _run_export(
        src,
        dst,
        extra=["--instrument", "Flute", "--dynamic", "fff"],
    )
    assert proc.returncode == 0, proc.stderr + proc.stdout
    df = pd.read_excel(dst, sheet_name="Spectral_Density_Metrics", engine="openpyxl")
    assert (df["Instrument"].astype(str) == "Bassoon").all()
    assert (df["Dynamic"].astype(str) == "pp").all()


def test_force_metadata_overrides_existing(tmp_path: Path) -> None:
    src = tmp_path / "compiled_density_metrics.xlsx"
    dst = tmp_path / "out.xlsx"
    _write_compiled_with_instrument_dynamic(src)
    proc = _run_export(
        src,
        dst,
        extra=["--instrument", "Clarinet", "--dynamic", "mf", "--force-metadata"],
    )
    assert proc.returncode == 0, proc.stderr + proc.stdout
    df = pd.read_excel(dst, sheet_name="Spectral_Density_Metrics", engine="openpyxl")
    assert (df["Instrument"].astype(str) == "Clarinet").all()
    assert (df["Dynamic"].astype(str) == "mf").all()


def test_inference_from_filename_tokens(tmp_path: Path) -> None:
    src = tmp_path / "compiled_density_metrics.xlsx"
    dst = tmp_path / "out.xlsx"
    _write_minimal_compiled_workbook(src)
    assert _run_export(src, dst).returncode == 0
    df = pd.read_excel(dst, sheet_name="Spectral_Density_Metrics", engine="openpyxl")
    assert df.loc[df["Note"] == "A4", "Instrument"].iloc[0] == "Clarinet"
    assert df.loc[df["Note"] == "A4", "Dynamic"].iloc[0] == "pp"
    assert df.loc[df["Note"] == "D3", "Instrument"].iloc[0] == "Bassoon"
    assert df.loc[df["Note"] == "D3", "Dynamic"].iloc[0] == "mf"


def test_plain_filenames_yield_blank_metadata_and_readme_warnings(tmp_path: Path) -> None:
    src = tmp_path / "compiled_density_metrics.xlsx"
    dst = tmp_path / "out.xlsx"
    dm = pd.DataFrame(
        {
            "Note": ["C4"],
            "density_metric_raw": [0.5],
            "density_metric_normalized": [1.0],
            "weighted_harmonic_density_contribution": [0.2],
            "weighted_inharmonic_density_contribution": [0.3],
            "weighted_subbass_density_contribution": [0.5],
            "component_harmonic_energy_ratio": [0.2],
            "component_inharmonic_energy_ratio": [0.3],
            "component_subbass_energy_ratio": [0.5],
            "Harmonic Partials sum": [1.0],
            "Inharmonic Partials sum": [2.0],
            "Sub-bass sum": [3.0],
            "Total sum": [6.0],
            "source_file_name": ["recording_plain.wav"],
            "weight_function": ["linear"],
            "density_weighted_sum": [1.5],
            "density_log_weighted": [0.1],
            "effective_partial_density": [0.4],
            "spectral_entropy": [0.7],
        }
    )
    am = pd.DataFrame([("pipeline_contract_version", "test-contract")], columns=["Parameter", "Value"])
    with pd.ExcelWriter(src, engine="openpyxl") as writer:
        dm.to_excel(writer, sheet_name="Density_Metrics", index=False)
        am.to_excel(writer, sheet_name="Analysis_Metadata", index=False)
    proc = _run_export(src, dst)
    assert proc.returncode == 0, proc.stderr + proc.stdout
    sdf = pd.read_excel(dst, sheet_name="Spectral_Density_Metrics", engine="openpyxl")
    if "Instrument" in sdf.columns:
        assert pd.isna(sdf.loc[0, "Instrument"]) or str(sdf.loc[0, "Instrument"]).strip() == ""
    if "Dynamic" in sdf.columns:
        assert pd.isna(sdf.loc[0, "Dynamic"]) or str(sdf.loc[0, "Dynamic"]).strip() == ""
    readme = pd.read_excel(dst, sheet_name="README", engine="openpyxl", header=None)
    text = "\n".join(str(x) for x in readme.iloc[:, 0].tolist() if pd.notna(x))
    assert "Instrument column missing" in text or "could not be inferred" in text
    assert "Dynamic column missing" in text or "could not be inferred" in text


def test_inference_ff_from_filename(tmp_path: Path) -> None:
    src = tmp_path / "compiled_density_metrics.xlsx"
    dst = tmp_path / "out.xlsx"
    dm = pd.DataFrame(
        {
            "Note": ["E5"],
            "density_metric_raw": [1.0],
            "density_metric_normalized": [1.0],
            "weighted_harmonic_density_contribution": [0.2],
            "weighted_inharmonic_density_contribution": [0.3],
            "weighted_subbass_density_contribution": [0.5],
            "component_harmonic_energy_ratio": [0.2],
            "component_inharmonic_energy_ratio": [0.3],
            "component_subbass_energy_ratio": [0.5],
            "Harmonic Partials sum": [1.0],
            "Inharmonic Partials sum": [2.0],
            "Sub-bass sum": [3.0],
            "Total sum": [6.0],
            "source_file_name": ["Trumpet_E5_ff.wav"],
            "weight_function": ["linear"],
            "density_weighted_sum": [1.5],
            "density_log_weighted": [0.1],
            "effective_partial_density": [0.4],
            "spectral_entropy": [0.7],
        }
    )
    am = pd.DataFrame([("pipeline_contract_version", "test-contract")], columns=["Parameter", "Value"])
    with pd.ExcelWriter(src, engine="openpyxl") as writer:
        dm.to_excel(writer, sheet_name="Density_Metrics", index=False)
        am.to_excel(writer, sheet_name="Analysis_Metadata", index=False)
    assert _run_export(src, dst).returncode == 0
    df = pd.read_excel(dst, sheet_name="Spectral_Density_Metrics", engine="openpyxl")
    assert df.loc[df["Note"] == "E5", "Instrument"].iloc[0] == "Trumpet"
    assert df.loc[df["Note"] == "E5", "Dynamic"].iloc[0] == "ff"


def test_chart_paths_relative_to_compiled_parent(tmp_path: Path) -> None:
    root = tmp_path
    src = root / "compiled_density_metrics.xlsx"
    _write_minimal_compiled_workbook(src)
    stem_dir = root / "Clarinet_A4_pp" / "A4"
    stem_dir.mkdir(parents=True)
    (stem_dir / "component_amplitude_mass_pie.png").write_bytes(b"\x89PNG\r\n\x1a\n")
    (stem_dir / "component_energy_ratio_pie.png").write_bytes(b"\x89PNG\r\n\x1a\n")
    dst = root / "compiled_density_metrics_research.xlsx"
    assert _run_export(src, dst).returncode == 0
    df = pd.read_excel(dst, sheet_name="Spectral_Density_Metrics", engine="openpyxl")
    row = df.loc[df["Note"] == "A4"].iloc[0]
    amp = str(row["amplitude_mass_chart_file"]).replace("\\", "/")
    erg = str(row["energy_ratio_chart_file"]).replace("\\", "/")
    assert "Clarinet_A4_pp/A4/component_amplitude_mass_pie.png" in amp
    assert "Clarinet_A4_pp/A4/component_energy_ratio_pie.png" in erg


def test_research_export_no_path_columns_canonical_alias_from_v5(tmp_path: Path) -> None:
    """Omit path-like columns; ``canonical_density_v5_adapted`` appears only as ``canonical_density``."""
    src = tmp_path / "compiled_density_metrics.xlsx"
    dst = tmp_path / "out.xlsx"
    dm = pd.DataFrame(
        {
            "Note": ["A4", "D3"],
            "density_metric_raw": [1.0, 1.0],
            "density_metric_normalized": [1.0, 1.0],
            "weighted_harmonic_density_contribution": [0.2, 0.8],
            "weighted_inharmonic_density_contribution": [0.3, 0.15],
            "weighted_subbass_density_contribution": [0.5, 0.05],
            "component_harmonic_energy_ratio": [0.2, 0.8],
            "component_inharmonic_energy_ratio": [0.3, 0.15],
            "component_subbass_energy_ratio": [0.5, 0.05],
            "Harmonic Partials sum": [1.0, 10.0],
            "Inharmonic Partials sum": [2.0, 20.0],
            "Sub-bass sum": [3.0, 30.0],
            "Total sum": [6.0, 60.0],
            "source_file_name": ["Clarinet_A4_pp.wav", "Bassoon_D3_mf.wav"],
            "weight_function": ["linear", "linear"],
            "density_weighted_sum": [1.5, 2.5],
            "density_log_weighted": [0.1, 0.2],
            "effective_partial_density": [0.4, 0.5],
            "spectral_entropy": [0.7, 0.8],
        }
    )
    canon = pd.DataFrame(
        {
            "Note": ["A4", "D3"],
            "canonical_density_v5_adapted": [1.25, 2.5],
        }
    )
    am = pd.DataFrame(
        [
            ("pipeline_contract_version", "test-contract"),
            ("ANALYSIS_SCHEMA_VERSION", "99"),
            ("weight_function", "linear"),
        ],
        columns=["Parameter", "Value"],
    )
    with pd.ExcelWriter(src, engine="openpyxl") as writer:
        dm.to_excel(writer, sheet_name="Density_Metrics", index=False)
        canon.to_excel(writer, sheet_name="Canonical_Metrics", index=False)
        am.to_excel(writer, sheet_name="Analysis_Metadata", index=False)
    proc = _run_export(src, dst)
    assert proc.returncode == 0, proc.stderr + proc.stdout
    sdm = pd.read_excel(dst, sheet_name="Spectral_Density_Metrics", engine="openpyxl")
    vs = pd.read_excel(dst, sheet_name="Validation_Summary", engine="openpyxl")
    assert "Source_File" not in sdm.columns
    assert "Source_Workbook" not in sdm.columns
    assert "canonical_density_v5_adapted" not in sdm.columns
    assert "canonical_density" in sdm.columns
    assert float(sdm.loc[sdm["Note"] == "A4", "canonical_density"].iloc[0]) == pytest.approx(1.25)
    assert "Source_File" not in vs.columns
