# -*- coding: utf-8 -*-
"""
Publication-safe metadata: redact local absolute paths from exported JSON/Excel/CSV/text.

Runtime code may still use full paths on disk; only values written to scientific exports are redacted.
"""

from __future__ import annotations

import getpass
import hashlib
import json
import os
import re
import zipfile
from pathlib import Path, PureWindowsPath
from datetime import datetime, timezone
from typing import Any, Dict, Iterable, List, Mapping, MutableMapping, Optional, Set, Tuple, Union

import pandas as pd

try:
    from constants import REDACT_LOCAL_PATHS_FOR_PUBLICATION
except Exception:  # pragma: no cover - constants must exist in normal runs
    REDACT_LOCAL_PATHS_FOR_PUBLICATION = True  # type: ignore[misc]

try:
    from constants import PUBLICATION_CLEAN_EXPORT
except Exception:  # pragma: no cover
    PUBLICATION_CLEAN_EXPORT = True  # type: ignore[misc]

REDACT_TOKEN = "redacted_for_publication"

# Keys whose values are always treated as path-like for publication exports
_PATH_KEY_FRAGMENTS: tuple[str, ...] = (
    "_path",
    "_dir",
    "folder_path",
    "output_path",
    "output_directory",
    "analysis_results_path",
    "results_directory",
    "source_path",
    "audio_path",
    "local_path",
    "parent_directory",
    "full_path",
    "compiled_excel_path",
    "excel_summary_path",
    "audio_dir",
    "output_dir",
    "file_path",
)

_PATH_KEY_EXACT_LOWER: frozenset[str] = frozenset(
    {
        "file_path",
        "output_dir",
        "folder_path",
        "results_directory",
        "source_path",
        "audio_path",
        "local_path",
        "parent_directory",
        "full_path",
    }
)

_WIN_ABS = re.compile(r"(?:[A-Za-z]:\\(?:[^\\/:*?\"<>|\r\n]+\\)*[^\\/:*?\"<>|\r\n]*)")
_WIN_USERS = re.compile(r"[A-Za-z]:\\Users\\", re.IGNORECASE)
_POSIX_USERS = re.compile(r"/Users/[^/\s]+")
_POSIX_HOME = re.compile(r"/home/[^/\s]+")
_POSIX_MNT = re.compile(r"/mnt/[^/\s]+")
_DESKTOP_SEG = re.compile(r"(?:^|[\\/])Desktop[\\/]", re.IGNORECASE)


def publication_redaction_enabled() -> bool:
    return bool(REDACT_LOCAL_PATHS_FOR_PUBLICATION)


def publication_clean_export_enabled() -> bool:
    """Publication-facing workbooks strip paths, redacted-only columns, and dev metadata."""
    return bool(PUBLICATION_CLEAN_EXPORT)


def format_utc_publication_timestamp() -> str:
    """Locale-neutral UTC timestamp for README / Metadata / Dashboard (no %Z / tzname)."""
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


# Parameters dropped from per-note ``Analysis_Metadata`` (long rows) when publication-clean is on.
PUBLICATION_ANALYSIS_META_PARAMETERS_OMIT: frozenset[str] = frozenset(
    {
        "compile_metrics_file",
        "publication_chart_policy_file",
        "sys_executable",
        "cwd",
        "folder_path",
        "source_compiled_workbook",
        "Source workbook",
        "Source_File",
        "proc_audio_file",
    }
)

# Keys removed from compiled ``Analysis_Metadata`` wide dict before Excel write.
PUBLICATION_META_FLAT_KEYS_OMIT: frozenset[str] = frozenset(
    {
        "compile_metrics_file",
        "publication_chart_policy_file",
        "sys_executable",
        "cwd",
        "folder_path",
        "source_compiled_workbook",
        "Source workbook",
        "Source_File",
        "density_metrics_publication_warning",
        "proc_audio_file",
    }
)


def generalize_platform_for_publication(value: Any) -> Any:
    """Replace detailed ``platform.platform()`` strings with a coarse OS label."""
    if value is None:
        return value
    s = str(value).strip()
    if not s:
        return value
    low = s.lower()
    if low.startswith("win"):
        return "Windows"
    if low.startswith("darwin") or "macosx" in low:
        return "macOS"
    if low.startswith("linux"):
        return "Linux"
    return "unknown_platform"


def neutralize_orchestrator_validation_label(value: Any) -> Any:
    """Replace internal GUI orchestrator audit tokens in exported metadata."""
    if value is None:
        return value
    s = str(value).strip()
    if not s:
        return value
    if "not_validated_orchestrator" in s:
        return "gui_orchestrator_pipeline"
    return value


def filter_analysis_meta_rows_publication_clean(
    rows: List[Tuple[str, Any]],
) -> List[Tuple[str, Any]]:
    """Filter long-format Analysis_Metadata rows for publication exports."""
    if not publication_clean_export_enabled():
        return rows
    out: List[Tuple[str, Any]] = []
    for param, val in rows:
        ps = str(param)
        if ps in PUBLICATION_ANALYSIS_META_PARAMETERS_OMIT:
            continue
        v: Any = val
        if ps == "platform":
            v = generalize_platform_for_publication(v)
        elif ps == "input_schema_validation_status":
            v = neutralize_orchestrator_validation_label(v)
        elif ps.endswith("_file") or ps.endswith("_path") or ps.endswith("_dir"):
            if isinstance(v, str) and v.strip() and (
                detect_absolute_local_path(v) or string_contains_forbidden_local_path(v)
            ):
                v = sanitize_path_for_publication(v, dataset_root=None)
        out.append((ps, v))
    return out


def apply_publication_clean_meta_flat(meta_flat: Mapping[str, Any]) -> Dict[str, Any]:
    """Return a copy of compiled workbook metadata with publication-only keys removed or rewritten."""
    if not publication_clean_export_enabled():
        return dict(meta_flat)
    out: Dict[str, Any] = dict(meta_flat)
    for k in PUBLICATION_META_FLAT_KEYS_OMIT:
        out.pop(k, None)
    if "platform" in out:
        out["platform"] = generalize_platform_for_publication(out.get("platform"))
    if "input_schema_validation_status" in out:
        out["input_schema_validation_status"] = neutralize_orchestrator_validation_label(
            out.get("input_schema_validation_status")
        )
    return out


def _column_values_all_blank_or_redacted(series: pd.Series) -> bool:
    if series is None:
        return True
    s = series
    if s.isna().all():
        return True
    if pd.api.types.is_numeric_dtype(s):
        return bool(s.isna().all())
    stripped = s.astype(str).str.strip()
    blank_like = stripped.isin(("", "nan", "None", "NaN", "<NA>"))
    if blank_like.all():
        return True
    rest = stripped[~blank_like]
    if rest.empty:
        return True
    return bool((rest == REDACT_TOKEN).all())


def drop_publication_noise_columns_from_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Drop columns that are entirely empty/NaN or only ``REDACT_TOKEN`` (publication polish)."""
    if df is None or df.empty or not publication_clean_export_enabled():
        return df
    protected = frozenset(
        {
            "Note",
            "MIDI",
            "Pitch_Class",
            "Octave",
            "Register",
        }
    )
    keep: List[Any] = []
    for c in df.columns:
        cs = str(c)
        if cs in protected:
            keep.append(c)
            continue
        if _column_values_all_blank_or_redacted(df[c]):
            continue
        keep.append(c)
    if len(keep) == len(df.columns):
        return df
    return df.loc[:, keep].copy()


def apply_publication_clean_research_metadata_fields(
    rows: Mapping[str, Any],
    *,
    workbook_basename: str,
) -> Dict[str, Any]:
    """Strip absolute paths and coarse-label platform for the research workbook Metadata sheet."""
    if not publication_clean_export_enabled():
        return dict(rows)
    out: Dict[str, Any] = dict(rows)
    out.pop("source_compiled_workbook", None)
    out["research_export_source_workbook"] = workbook_basename
    for key in ("compiled_from", "output_path", "folder_path", "results_directory"):
        v = out.get(key)
        if isinstance(v, str) and v.strip() and (
            detect_absolute_local_path(v) or string_contains_forbidden_local_path(v)
        ):
            out[key] = sanitize_path_for_publication(v, dataset_root=None)
    if "platform" in out:
        out["platform"] = generalize_platform_for_publication(out.get("platform"))
    if "input_schema_validation_status" in out:
        out["input_schema_validation_status"] = neutralize_orchestrator_validation_label(
            out.get("input_schema_validation_status")
        )
    return out


def publication_clean_drop_known_sparse_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Remove known diagnostic columns when publication-clean and entirely uninformative."""
    if df is None or df.empty or not publication_clean_export_enabled():
        return df
    candidates = (
        "density_normalization_denominator",
        "accepted_inharmonic_peak_count",
        "accepted_inharmonic_partial_count",
        "debug_counts_invariant_failures",
        "f0_detuning_cents_from_nominal",
        "leakage_guard_cutoff_hz",
        "input_schema_validation_status",
        "publication_output_allowed",
        "Source_File",
        "Source_Workbook",
        "density_formula_version",
    )
    drop: List[str] = []
    for name in candidates:
        if name not in df.columns:
            continue
        if _column_values_all_blank_or_redacted(df[name]):
            drop.append(name)
    if not drop:
        return df
    return df.drop(columns=drop, errors="ignore")


def publication_research_canonical_density_columns(df: pd.DataFrame) -> pd.DataFrame:
    """For research exports: keep ``canonical_density``; drop or hide ``canonical_density_v5_adapted``."""
    if df is None or df.empty or not publication_clean_export_enabled():
        return df
    if "canonical_density" in df.columns and "canonical_density_v5_adapted" in df.columns:
        return df.drop(columns=["canonical_density_v5_adapted"], errors="ignore")
    if "canonical_density_v5_adapted" in df.columns and "canonical_density" not in df.columns:
        return df.rename(columns={"canonical_density_v5_adapted": "canonical_density"})
    return df


def is_absolute_path_like(value: Any) -> bool:
    """True if *value* looks like a Windows or POSIX absolute filesystem path."""
    if value is None or not isinstance(value, str):
        return False
    s = value.strip()
    if not s:
        return False
    if len(s) >= 2 and s[1] == ":" and s[0].isalpha() and (len(s) == 2 or s[2] in "\\/"):
        return True
    if s.startswith("/Users/") or s.startswith("/home/") or s.startswith("/mnt/"):
        return True
    if _WIN_USERS.search(s):
        return True
    if _POSIX_USERS.search(s) or _POSIX_HOME.search(s) or _POSIX_MNT.search(s):
        return True
    return False


def string_contains_forbidden_local_path(s: str) -> bool:
    """Heuristic: substring patterns that should not appear in publication exports."""
    if not isinstance(s, str) or not s:
        return False
    if _WIN_USERS.search(s):
        return True
    if _POSIX_USERS.search(s) or _POSIX_HOME.search(s) or _POSIX_MNT.search(s):
        return True
    if _DESKTOP_SEG.search(s) and _WIN_ABS.search(s):
        return True
    if re.search(r"(?:^|[\s\"'`])/[A-Za-z0-9_.-]+/[^/\s]+", s) and ("/Users/" in s or "/home/" in s or "/mnt/" in s):
        return True
    return False


def _key_is_path_sensitive(key: str) -> bool:
    lk = str(key).lower().strip()
    if lk in _PATH_KEY_EXACT_LOWER:
        return True
    return any(lk.endswith(sfx) for sfx in ("_path", "_dir"))


def redact_path(value: Any, project_root: Optional[Path] = None) -> Any:
    """Replace absolute path strings with REDACT_TOKEN or a safe relative fragment."""
    if not publication_redaction_enabled():
        return value
    if value is None or not isinstance(value, str):
        return value
    s = value.strip()
    if not s:
        return value
    if project_root is not None:
        try:
            p = Path(s)
            pr = Path(project_root).resolve()
            if p.is_absolute():
                try:
                    rel = p.resolve().relative_to(pr)
                    rel_s = rel.as_posix()
                    if not rel_s.startswith("..") and ".." not in rel_s:
                        return f"<DATASET_ROOT>/{rel_s}"
                except Exception:
                    pass
        except Exception:
            pass
    if not (is_absolute_path_like(s) or string_contains_forbidden_local_path(s)):
        return value
    return REDACT_TOKEN


def sanitize_metadata_value(value: Any, project_root: Optional[Path] = None) -> Any:
    """Recursively sanitize strings inside dicts/lists; redact path-like strings."""
    if not publication_redaction_enabled():
        return value
    if isinstance(value, str):
        if not string_contains_forbidden_local_path(value) and not is_absolute_path_like(value):
            return value
        # Embedded path inside longer text: replace known path segments
        out = value
        for m in list(_WIN_ABS.finditer(out)):
            span = m.group(0)
            if is_absolute_path_like(span) or string_contains_forbidden_local_path(span):
                out = out.replace(span, REDACT_TOKEN)
        out = _POSIX_USERS.sub(REDACT_TOKEN, out)
        out = _POSIX_HOME.sub(REDACT_TOKEN, out)
        out = _POSIX_MNT.sub(REDACT_TOKEN, out)
        if string_contains_forbidden_local_path(out) or is_absolute_path_like(out.strip()):
            return REDACT_TOKEN
        return out
    if isinstance(value, dict):
        return sanitize_metadata_dict(value, project_root=project_root)
    if isinstance(value, list):
        return [sanitize_metadata_value(v, project_root=project_root) for v in value]
    if isinstance(value, tuple):
        return tuple(sanitize_metadata_value(v, project_root=project_root) for v in value)
    return value


def sanitize_metadata_dict(d: Mapping[str, Any], project_root: Optional[Path] = None) -> Dict[str, Any]:
    if not publication_redaction_enabled():
        return dict(d)
    out: Dict[str, Any] = {}
    for k, v in d.items():
        ks = str(k)
        if _key_is_path_sensitive(ks):
            # Never emit raw local directories / absolute paths under path-like keys.
            if isinstance(v, str) and v.strip() and v.strip() != REDACT_TOKEN:
                out[ks] = REDACT_TOKEN
            elif isinstance(v, (dict, list, tuple)):
                out[ks] = sanitize_metadata_value(v, project_root=project_root)
            else:
                out[ks] = v
            continue
        out[ks] = sanitize_metadata_value(v, project_root=project_root)
    return out


def enrich_and_redact_batch_audio_result(
    result: Dict[str, Any],
    file_path: Union[str, Path],
    note_name: Optional[str] = None,
) -> Dict[str, Any]:
    """Add publication-safe ids and strip absolute paths from one batch row dict."""
    out = dict(result)
    fp = Path(file_path)
    if publication_redaction_enabled():
        out["source_file_basename"] = fp.name
        h = short_sha256_for_file(fp)
        if h:
            out["source_file_hash_short"] = h
        out["public_audio_id"] = make_public_audio_id(fp, note_name, None)
    return sanitize_metadata_dict(out)


def sanitize_dataframe_for_publication(df: pd.DataFrame, project_root: Optional[Path] = None) -> pd.DataFrame:
    """Redact path-like strings in every cell (object columns and stray strings)."""
    if not publication_redaction_enabled() or df is None or df.empty:
        return df
    out = df.copy()
    for col in out.columns:
        ckey = str(col)
        if _key_is_path_sensitive(ckey):
            out[col] = out[col].map(
                lambda v: REDACT_TOKEN
                if isinstance(v, str) and str(v).strip() not in ("", REDACT_TOKEN)
                else sanitize_metadata_value(v, project_root=project_root)
            )
            continue
        out[col] = out[col].map(lambda v: sanitize_metadata_value(v, project_root=project_root))
    return out


def short_sha256_for_file(file_path: Union[str, Path], length: int = 12) -> Optional[str]:
    """Short hash of file bytes for reproducibility (None if unreadable)."""
    p = Path(file_path)
    try:
        if not p.is_file():
            return None
        h = hashlib.sha256()
        with p.open("rb") as f:
            for chunk in iter(lambda: f.read(1024 * 1024), b""):
                h.update(chunk)
        return h.hexdigest()[: int(length)]
    except Exception:
        return None


def make_public_audio_id(
    file_path: Union[str, Path],
    note: Optional[str] = None,
    index: Optional[int] = None,
) -> str:
    """Stable-ish public identifier without exposing local directories."""
    p = Path(file_path)
    stem = p.stem
    h = short_sha256_for_file(p, 8) or "nohash"
    parts = ["audio", h, stem.replace(" ", "_")[:48]]
    if note:
        parts.append(str(note).replace(" ", "_")[:24])
    if index is not None:
        parts.append(f"i{int(index)}")
    return "__".join(parts)


def sanitize_excel_workbook_for_publication(path: Union[str, Path]) -> None:
    """In-place redaction of all string cells in an existing .xlsx workbook."""
    if not publication_redaction_enabled():
        return
    path = Path(path)
    if not path.is_file():
        return
    try:
        from openpyxl import load_workbook
    except ImportError:
        return
    wb = load_workbook(path)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and (is_absolute_path_like(v) or string_contains_forbidden_local_path(v)):
                    cell.value = sanitize_metadata_value(v)
    wb.save(path)


def string_fails_publication_scan(text: str) -> bool:
    """True if *text* contains obvious local absolute path material (publication FAIL)."""
    if not isinstance(text, str) or not text.strip():
        return False
    low = text.lower()
    if "c:\\users\\" in low or "c:/users/" in low:
        return True
    if "/Users/" in text or "/home/" in low or "/mnt/" in low:
        return True
    if re.search(r"[a-z]:\\[^\n]*\\desktop\\", low):
        return True
    if re.search(r"[a-z]:\\users\\", low):
        return True
    return False


def scan_text_for_forbidden_paths(text: str) -> bool:
    """Alias for :func:`string_fails_publication_scan` (validator / scanner API)."""
    return string_fails_publication_scan(text)


def list_publication_path_violations_in_excel(path: Union[str, Path]) -> List[str]:
    """Return human-readable violations when any cell (any sheet) contains forbidden path text."""
    path = Path(path)
    if not path.is_file():
        return [f"not a file: {path}"]
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return [f"cannot scan workbook: {e}"]
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for val in row:
                    if isinstance(val, str) and string_fails_publication_scan(val):
                        return [f"{path.name} sheet={ws.title!r}: forbidden local path text in cell"]
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return []


# ---------------------------------------------------------------------------
# JSON / repository export tree sanitization (v7-compatible, additive API).
# Uses ``dataset_root`` for corpus-relative paths; separate from
# :func:`sanitize_metadata_dict` which follows REDACT_TOKEN rules for Excel/CSV.
# ---------------------------------------------------------------------------

_ENV_PROFILE_EXPORT = ("SOUNDSPECTRAN_EXPORT_PROFILE", "SOUNDSPECTRANALYSE_EXPORT_PROFILE")
_ENV_PRIVACY_EXPORT = "SOUNDSPECTRAN_METADATA_PRIVACY_MODE"
_ENV_LEGACY_ABS_EXPORT = "SOUNDSPECTRANALYSE_EXPORT_ABSOLUTE_PATHS"

_WIN_ABS_START_EXPORT = re.compile(r"^[A-Za-z]:[/\\]")
_UNC_EXPORT = re.compile(r"^\\\\(\?|\.)\\")
_UNIX_ABS_EXPORT = re.compile(r"^/(Users|home|tmp|var|mnt)(/|$)", re.I)
_WIN_USERS_EXPORT = re.compile(r"[A-Za-z]:\\Users\\", re.I)
_POSIX_USERS_EXPORT = re.compile(r"/Users/|/home/", re.I)
_TMP_WIN_EXPORT = re.compile(r"\\AppData\\Local\\Temp\\|\\Temp\\", re.I)
_EMBED_WIN_DRIVE_EXPORT = re.compile(r"(?<![A-Za-z0-9_])[A-Za-z]:\\Users\\", re.I)
_EMBED_POSIX_HOME_EXPORT = re.compile(r"/Users/[^/\s\"']+|/home/[^/\s\"']+", re.I)

_EXPORT_PATH_LIKE_KEYS = frozenset(
    {
        "audio_file",
        "file_path",
        "input_path",
        "source_path",
        "path",
        "output_dir",
        "output_folder",
        "run_folder",
        "folder_path",
        "working_directory",
        "cwd",
        "dataset_root",
    }
)


def get_export_profile() -> str:
    for k in _ENV_PROFILE_EXPORT:
        v = os.environ.get(k, "").strip()
        if v:
            return v.lower()
    return ""


def _username_hints_export() -> List[str]:
    out: List[str] = []
    for key in ("USERNAME", "USER"):
        u = os.environ.get(key, "").strip()
        if u and len(u) >= 2:
            out.append(u)
    try:
        gu = getpass.getuser()
        if gu and gu not in out and len(gu) >= 2:
            out.append(gu)
    except Exception:
        pass
    return [x for x in out if x]


def get_metadata_privacy_mode() -> str:
    """``public`` (default) or ``internal_debug`` for structured JSON-style exports."""
    if get_export_profile() == "public_repository":
        return "public"
    if os.environ.get(_ENV_LEGACY_ABS_EXPORT, "").strip().lower() in ("1", "true", "yes", "debug"):
        return "internal_debug"
    v = os.environ.get(_ENV_PRIVACY_EXPORT, "public").strip().lower()
    if v in ("internal", "internal_debug", "debug"):
        return "internal_debug"
    return "public"


def detect_absolute_local_path(value: Any) -> bool:
    """True if *value* looks like an absolute local path or embeds home-style segments."""
    if isinstance(value, Path):
        value = str(value)
    if not isinstance(value, str) or not value.strip():
        return False
    t = value.strip().strip('"').strip("'")
    if t.startswith("file://"):
        t = t[7:]
    if re.match(r"^[a-z]+://", t, re.I):
        return False
    if _UNC_EXPORT.match(t) or _WIN_ABS_START_EXPORT.match(t):
        return True
    if t.startswith("/") and not t.startswith("//"):
        if _UNIX_ABS_EXPORT.match(t) or t.startswith("/tmp/") or t.startswith("/var/"):
            return True
        if re.match(r"^/[^/\s]+/", t) and len(t) > 2:
            return True
    if _WIN_USERS_EXPORT.search(t) or _POSIX_USERS_EXPORT.search(t):
        return True
    if _TMP_WIN_EXPORT.search(t):
        return True
    if "~" in t[:3] and (t.startswith("~/") or t.startswith("~\\")):
        return True
    if _EMBED_WIN_DRIVE_EXPORT.search(t) or _EMBED_POSIX_HOME_EXPORT.search(t):
        return True
    if re.search(r"\\Desktop\\", t, re.I) or re.search(r"/Desktop/", t, re.I):
        return True
    return False


def publication_audio_path_fields(path: Union[str, Path], *, dataset_root: Optional[Path] = None) -> dict[str, Any]:
    """Publication-safe audio path decomposition (no host directories)."""
    path_text = str(path)
    is_windows_style = bool(_WIN_ABS_START_EXPORT.match(path_text)) or "\\" in path_text
    p = PureWindowsPath(path_text) if is_windows_style else Path(path_text)
    name = p.name or "unknown_audio"
    stem = p.stem
    ext = p.suffix
    if ext and not ext.startswith("."):
        ext = "." + ext
    rel = name
    if dataset_root is not None:
        try:
            rel = Path(path_text).resolve().relative_to(Path(dataset_root).resolve()).as_posix()
        except Exception:
            rel = name
    rel_posix = rel.replace("\\", "/")
    return {
        "audio_file_name": name,
        "audio_file_stem": stem,
        "audio_file_extension": ext if ext else "",
        "audio_file_basename": stem,
        "audio_file_ext": ext if ext else "",
        "audio_relative_path": rel_posix,
        "dataset_relative_path": rel_posix,
        "audio_path_policy": "sanitized_for_publication",
    }


def publication_output_dir_fields(path: Union[str, Path], *, dataset_root: Optional[Path] = None) -> dict[str, Any]:
    """Publication-safe output directory (relative to corpus root when possible)."""
    p = Path(str(path))
    if dataset_root is not None:
        try:
            rel = Path(p).resolve().relative_to(Path(dataset_root).resolve()).as_posix()
        except Exception:
            rel = p.name or "output"
    else:
        rel = p.name or "output"
    return {
        "output_dir_relative": rel.replace("\\", "/"),
        "output_path_policy": "sanitized_for_publication",
    }


def sanitize_path_for_publication(path: Union[str, Path, None], dataset_root: Optional[Path] = None) -> str:
    """Basename or POSIX path relative to *dataset_root* (never absolute host paths)."""
    if path is None:
        return ""
    path_text = str(path)
    # Treat Windows-formatted paths explicitly so basename extraction works on Linux runners.
    is_windows_style = bool(_WIN_ABS_START_EXPORT.match(path_text)) or "\\" in path_text
    p = Path(path_text)
    if dataset_root is not None:
        try:
            if is_windows_style:
                return PureWindowsPath(path_text).name
            return p.resolve().relative_to(Path(dataset_root).resolve()).as_posix()
        except Exception:
            pass
    if is_windows_style:
        return PureWindowsPath(path_text).name or "path_redacted"
    return p.name if p.name else "path_redacted"


def detect_private_path_leakage_fragment(text: Any) -> bool:
    """True when *text* embeds obvious local-only path material."""
    if not isinstance(text, str) or not text.strip():
        return False
    s = text
    low = s.lower()
    if detect_absolute_local_path(s):
        return True
    if "://" in low[:12]:
        return False
    if "/" not in s and "\\" not in s:
        return False
    if "desktop" in low or "/users/" in low or "\\users\\" in low or "/home/" in low or "\\appdata\\" in low:
        return True
    for u in _username_hints_export():
        if not u or len(u) < 2:
            continue
        ul = u.lower()
        if ul in low and ("/" in s or "\\" in s):
            return True
    if re.search(r"(?i)(^|[/\\])[^/\\]{2,}[/\\][^/\\]+[/\\][^/\\]+", s):
        if "desktop" in low or "users" in low or "appdata" in low:
            return True
    return False


def redact_public_path_leakage(value: str) -> str:
    """Strip host tails from free-text metadata in ``public`` mode."""
    if get_metadata_privacy_mode() != "public":
        return value
    if not isinstance(value, str) or not value.strip():
        return value
    t = value.strip()
    if ":" in t and not t.lower().startswith("http"):
        head, _, tail = t.partition(":")
        if detect_absolute_local_path(tail.strip()):
            return (head.strip() or "redacted").split("/")[-1].split("\\")[-1]
    if detect_private_path_leakage_fragment(t):
        return "redacted_local_path"
    return value


def _merge_audio_publication_export(
    out: MutableMapping[str, Any], raw_path: str, dataset_root: Optional[Path]
) -> None:
    fields = publication_audio_path_fields(raw_path, dataset_root=dataset_root)
    for fk, fv in fields.items():
        out.setdefault(fk, fv)
    out["audio_file"] = fields["audio_file_name"]


def _merge_output_publication_export(
    out: MutableMapping[str, Any], raw_path: str, dataset_root: Optional[Path]
) -> None:
    fields = publication_output_dir_fields(raw_path, dataset_root=dataset_root)
    for fk, fv in fields.items():
        out.setdefault(fk, fv)


def sanitize_export_metadata_value(value: Any, dataset_root: Optional[Path] = None) -> Any:
    """Sanitize one value for JSON/repository exports (v7 semantics)."""
    if value is None:
        return None
    if isinstance(value, Path):
        return sanitize_path_for_publication(value, dataset_root=dataset_root)
    if isinstance(value, str):
        if len(value) > 12 and (
            ":\\" in value or "/Users/" in value or "/home/" in value or value.strip().startswith("{")
        ):
            try:
                obj = json.loads(value)
                dumped = json.dumps(
                    sanitize_export_metadata_dict(obj, dataset_root=dataset_root),
                    ensure_ascii=False,
                    default=str,
                )
                if dumped != value:
                    return dumped
            except Exception:
                pass
        if detect_absolute_local_path(value):
            return sanitize_path_for_publication(value, dataset_root=dataset_root)
        return redact_public_path_leakage(value)
    if isinstance(value, Mapping):
        return sanitize_export_metadata_dict(dict(value), dataset_root=dataset_root)
    if isinstance(value, (list, tuple)):
        return [sanitize_export_metadata_value(v, dataset_root=dataset_root) for v in value]
    return value


def sanitize_export_metadata_dict(obj: Any, dataset_root: Optional[Path] = None) -> Any:
    """Deep sanitization for JSON-like trees (repository exports strip host paths)."""
    if obj is None:
        return None
    if isinstance(obj, Path):
        return sanitize_path_for_publication(obj, dataset_root=dataset_root)
    if isinstance(obj, str):
        return sanitize_export_metadata_value(obj, dataset_root=dataset_root)
    if isinstance(obj, Mapping):
        out: dict[str, Any] = {}
        for k, v in obj.items():
            ks = str(k)
            if ks == "dataset_root":
                continue
            if ks in _EXPORT_PATH_LIKE_KEYS and isinstance(v, str) and detect_absolute_local_path(v):
                if ks == "audio_file":
                    _merge_audio_publication_export(out, v, dataset_root)
                elif ks == "file_path":
                    fields = publication_audio_path_fields(v, dataset_root=dataset_root)
                    for fk, fv in fields.items():
                        out.setdefault(fk, fv)
                    out["file_path"] = fields["audio_file_name"]
                elif ks in ("output_dir", "output_folder", "run_folder", "folder_path"):
                    _merge_output_publication_export(out, v, dataset_root)
                    out[ks] = str(out.get("output_dir_relative", Path(v).name))
                else:
                    out[ks] = sanitize_path_for_publication(v, dataset_root=dataset_root)
            elif isinstance(v, (Mapping, list, tuple)):
                out[ks] = sanitize_export_metadata_dict(v, dataset_root=dataset_root)
            elif isinstance(v, str) and detect_absolute_local_path(v):
                out[ks] = sanitize_path_for_publication(v, dataset_root=dataset_root)
            else:
                out[ks] = sanitize_export_metadata_value(v, dataset_root=dataset_root)
        return out
    if isinstance(obj, (list, tuple)):
        return [sanitize_export_metadata_dict(v, dataset_root=dataset_root) for v in obj]
    return obj


def sanitize_run_parameters_json(json_string: str, dataset_root: Optional[Path] = None) -> str:
    """Parse JSON, sanitize path-like strings, re-serialize."""
    if not isinstance(json_string, str) or not json_string.strip():
        return json_string
    try:
        data = json.loads(json_string)
    except Exception:
        if detect_absolute_local_path(json_string):
            return sanitize_path_for_publication(json_string, dataset_root=dataset_root)
        return json_string
    return json.dumps(sanitize_export_metadata_dict(data, dataset_root=dataset_root), ensure_ascii=False, default=str)


def _forbidden_substrings_export(username: Optional[str]) -> List[str]:
    frags: List[str] = [
        "C:\\",
        "C:/",
        "C:\\Users",
        "C:/Users",
        "\\Users\\",
        "/Users/",
        "/home/",
        "\\Desktop\\",
        "/Desktop/",
        "Desktop",
    ]
    hints = _username_hints_export() if username is None else ([username] if username else [])
    for u in hints:
        if u and len(u) >= 2:
            frags.append(u)
            frags.append(u.lower())
    return frags


def validate_no_private_paths(
    output_root: Union[str, Path],
    *,
    username: Optional[str] = None,
    project_root: Optional[Path] = None,
    extra_forbidden: Optional[Iterable[str]] = None,
    scan_logs: bool = True,
    skip_relative_dir_parts: Tuple[str, ...] = ("local_debug_logs", "previous_runs"),
) -> Tuple[bool, List[str]]:
    """Scan *output_root* for private path leakage in text-like exports."""
    root = Path(output_root)
    errors: List[str] = []
    if not root.exists():
        return False, [f"output_root does not exist: {root}"]

    forbidden = list(_forbidden_substrings_export(username))
    if extra_forbidden:
        forbidden.extend(extra_forbidden)
    if project_root is not None:
        try:
            pr = project_root.resolve()
            forbidden.append(str(pr))
            if os.name == "nt":
                forbidden.append(str(pr).replace("/", "\\"))
        except Exception:
            pass

    def check_text(rel_display: str, text: str) -> None:
        tl = text
        tlower = tl.lower()
        if detect_private_path_leakage_fragment(tl):
            errors.append(f"{rel_display}: private_path_fragment")
        for frag in forbidden:
            if not frag:
                continue
            fl = frag.lower()
            if frag in tl or (len(frag) >= 3 and fl in tlower):
                errors.append(f"{rel_display}: forbidden fragment {frag!r}")

    exts_text = {".json", ".txt", ".csv", ".html", ".htm", ".xml", ".md"}
    if scan_logs:
        exts_text = set(exts_text) | {".log"}
    skip_parts = frozenset(skip_relative_dir_parts or ())
    for path in sorted(root.rglob("*")):
        if path.is_dir():
            continue
        rel = path.relative_to(root)
        if skip_parts and any(p in skip_parts for p in rel.parts):
            continue
        suf = path.suffix.lower()
        try:
            if suf in exts_text:
                check_text(rel.as_posix(), path.read_text(encoding="utf-8", errors="replace"))
            elif suf == ".xlsx":
                try:
                    import openpyxl

                    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                    for sheet in wb.worksheets:
                        for row in sheet.iter_rows(values_only=True):
                            for cell in row:
                                if isinstance(cell, str):
                                    check_text(f"{rel.as_posix()}#{sheet.title}", cell)
                    wb.close()
                except Exception as e:
                    errors.append(f"{rel}: xlsx scan failed: {e}")
            elif suf == ".zip":
                with zipfile.ZipFile(path, "r") as zf:
                    for name in zf.namelist():
                        if Path(name).suffix.lower() in exts_text:
                            with zf.open(name) as fh:
                                raw = fh.read()
                            check_text(f"{rel.as_posix()}:{name}", raw.decode("utf-8", errors="replace"))
        except Exception as e:
            errors.append(f"{rel}: read error {e}")

    return (len(errors) == 0, errors)


def assert_no_private_paths(output_root: Union[str, Path], **kwargs: Any) -> None:
    ok, errs = validate_no_private_paths(output_root, **kwargs)
    if not ok:
        raise AssertionError("Private path patterns found:\n" + "\n".join(errs[:200]))
