#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Batch Audio Analyzer - Process Multiple Audio Files (legacy batch wrapper)
==========================================================================

This script is not the canonical publication-facing pipeline. Prefer
``proc_audio.AudioProcessor`` followed by ``compile_metrics.compile_density_metrics_with_pca``.

Features (historical):
- Parallel processing for efficiency
- Aggregated results and statistics
- Comparative analysis across files
- Individual and summary reports
- Progress tracking

Author: AI Assistant
Version: 1.0.0
"""

import sys
import os
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple, Mapping
from datetime import datetime
import json
import pandas as pd
import numpy as np
from concurrent.futures import ProcessPoolExecutor, as_completed
from multiprocessing import cpu_count
import logging

# Ensure UTF-8 output on Windows consoles (avoids UnicodeEncodeError)
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")

# Import the SuperAudioAnalyzer
from super_audio_analyzer import SuperAudioAnalyzer

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('batch_analyzer.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

CANONICAL_PIPELINE_ROLE = "legacy_batch_wrapper"
PUBLICATION_OUTPUT_ALLOWED = False
# Batch scope: canonical linear power masses (Σ A² per STFT-bin class) and the matching three-class + total-inharmonic percents.
# Per-partial tables (dB / linear) stay in Phase-2 ``spectral_analysis.xlsx`` per note.
# ``success`` is kept in-memory for statistics only and is omitted from JSON/Excel exports.
BATCH_SUMMARY_COLUMNS: Tuple[str, ...] = (
    "file_name",
    "note",
    "harmonic_power_mass",
    "inharmonic_residual_power_mass",
    "subbass_noise_power_mass",
    "total_inharmonic_power_mass",
    "total_power_mass",
    "harmonic_power_percent",
    "inharmonic_residual_power_percent",
    "subbass_noise_power_percent",
    "total_inharmonic_power_percent",
)


def _extract_note_from_filename(filename: str) -> Optional[str]:
    import re

    patterns = (
        r"([A-G][#b]?)(\d+)",
        r"([A-G][♯♭]?)(\d+)",
    )
    for pattern in patterns:
        match = re.search(pattern, str(filename))
        if match:
            note = match.group(1).replace("♯", "#").replace("♭", "b")
            return f"{note}{match.group(2)}"
    return None


def _dedupe_batch_rows_by_file_name(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    seen: set[str] = set()
    out: List[Dict[str, Any]] = []
    for r in rows:
        fn = r.get("file_name")
        if fn is None:
            continue
        key = str(fn)
        if key in seen:
            continue
        seen.add(key)
        out.append(r)
    return out


def _batch_row_public_dict(row: Mapping[str, Any]) -> Dict[str, Any]:
    """Subset one batch row to exported columns only (no ``success``, paths, or side fields)."""
    return {k: row.get(k) for k in BATCH_SUMMARY_COLUMNS}


class BatchAudioAnalyzer:
    """
    Batch processor for analyzing multiple audio files.
    
    Supports up to 100 files with parallel processing for efficiency.
    """
    
    def __init__(
        self,
        audio_files: List[Path],
        output_dir: Optional[Path] = None,
        max_workers: Optional[int] = None,
        config_file: Optional[Path] = None,
        **analyzer_kwargs
    ):
        """
        Initialize batch analyzer.
        
        Args:
            audio_files: List of audio file paths to analyze
            output_dir: Directory for batch results (default: batch_results/)
            max_workers: Maximum parallel workers (default: min(cpu_count(), len(files)))
            config_file: Path to batch configuration JSON file (optional)
            **analyzer_kwargs: Arguments to pass to SuperAudioAnalyzer (overrides config)
        """
        if len(audio_files) > 100:
            raise ValueError(f"Maximum 100 files supported. Got {len(audio_files)} files.")
        
        # Remove duplicate files (based on absolute path) before processing
        unique_files = []
        seen_paths = set()
        for f in audio_files:
            file_path = Path(f).resolve()  # Use absolute path for comparison
            if file_path not in seen_paths:
                unique_files.append(file_path)
                seen_paths.add(file_path)
        
        if len(unique_files) < len(audio_files):
            removed = len(audio_files) - len(unique_files)
            logger.warning(f"Removed {removed} duplicate file(s) from input list")
        
        self.audio_files = unique_files
        self.output_dir = Path(output_dir) if output_dir else Path("batch_results")
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Load configuration file if provided
        self.config = self._load_config(config_file) if config_file else {}
        
        # Merge config with analyzer_kwargs (kwargs take precedence)
        if 'batch_analysis_config' in self.config:
            config_params = self.config['batch_analysis_config']
            
            # Handle weight mode
            weight_mode = config_params.get('weight_mode', 'auto_extract')
            if weight_mode == 'fixed':
                fixed_weights = config_params.get('fixed_weights', {})
                analyzer_kwargs['auto_extract_weights'] = False
                analyzer_kwargs['harmonic_weight'] = fixed_weights.get('harmonic_weight', 0.90)
                analyzer_kwargs['inharmonic_weight'] = fixed_weights.get('inharmonic_weight', 0.10)
                logger.info(f"Using FIXED weights from config: harmonic={analyzer_kwargs['harmonic_weight']}, inharmonic={analyzer_kwargs['inharmonic_weight']}")
            else:
                analyzer_kwargs['auto_extract_weights'] = True
                logger.info("Using AUTO-EXTRACT weights (per-file energy distribution)")
            
            # Merge analysis parameters from config
            if 'analysis_parameters' in config_params:
                for key, value in config_params['analysis_parameters'].items():
                    if key not in analyzer_kwargs:  # Don't override explicit kwargs
                        analyzer_kwargs[key] = value
        
        # Determine number of workers
        if max_workers is None:
            self.max_workers = min(cpu_count(), len(self.audio_files), 8)  # Cap at 8
        else:
            self.max_workers = min(max_workers, len(self.audio_files))
        
        self.analyzer_kwargs = analyzer_kwargs
        self.results = []
        self.summary_stats = {}
        
        logger.info(f"Initialized BatchAudioAnalyzer with {len(self.audio_files)} files")
        logger.info(f"Using {self.max_workers} parallel workers")
        logger.info(f"Weight mode: {'FIXED' if not analyzer_kwargs.get('auto_extract_weights', True) else 'AUTO-EXTRACT'}")

    @staticmethod
    def _hepd_fields_from_analyzer_metrics(analyzer: Any) -> Dict[str, Any]:
        """Extract harmonic effective power density fields from ``analyzer.metrics`` (unit tests / diagnostics)."""
        m = getattr(analyzer, "metrics", None)
        out: Dict[str, Any] = {
            "harmonic_effective_power_density": None,
            "harmonic_effective_power_density_component_count": None,
            "harmonic_effective_power_density_status": None,
            "harmonic_effective_power_density_max_amplitude": None,
            "harmonic_effective_power_density_total_power": None,
            "harmonic_effective_power_density_normalized_by_harmonic_count": None,
        }
        if not isinstance(m, dict):
            return out
        st = m.get("harmonic_effective_power_density_status")
        out["harmonic_effective_power_density_status"] = str(st) if st is not None else None
        try:
            if m.get("harmonic_effective_power_density_component_count") is not None:
                out["harmonic_effective_power_density_component_count"] = int(
                    m.get("harmonic_effective_power_density_component_count")
                )
        except (TypeError, ValueError):
            pass
        for k in (
            "harmonic_effective_power_density",
            "harmonic_effective_power_density_max_amplitude",
            "harmonic_effective_power_density_total_power",
            "harmonic_effective_power_density_normalized_by_harmonic_count",
        ):
            if m.get(k) is None:
                continue
            try:
                fv = float(m[k])
                if np.isfinite(fv):
                    out[k] = fv
            except (TypeError, ValueError):
                continue
        return out

    @staticmethod
    def _hepm_fields_from_analyzer_metrics(analyzer: Any) -> Dict[str, Any]:
        """Extract harmonic effective power mass fields from ``analyzer.metrics`` (unit tests / diagnostics)."""
        m = getattr(analyzer, "metrics", None)
        out: Dict[str, Any] = {
            "harmonic_effective_power_mass": None,
            "harmonic_effective_power_mean": None,
            "harmonic_effective_power_rms": None,
            "harmonic_effective_power_component_count": None,
            "harmonic_effective_power_mass_status": None,
        }
        if not isinstance(m, dict):
            return out
        st = m.get("harmonic_effective_power_mass_status")
        out["harmonic_effective_power_mass_status"] = str(st) if st is not None else None
        try:
            if m.get("harmonic_effective_power_component_count") is not None:
                out["harmonic_effective_power_component_count"] = int(
                    m.get("harmonic_effective_power_component_count")
                )
        except (TypeError, ValueError):
            pass
        for k in ("harmonic_effective_power_mass", "harmonic_effective_power_mean", "harmonic_effective_power_rms"):
            if m.get(k) is None:
                continue
            try:
                fv = float(m[k])
                if np.isfinite(fv):
                    out[k] = fv
            except (TypeError, ValueError):
                continue
        return out

    @staticmethod
    def _load_config(config_file: Path) -> Dict[str, Any]:
        """Load configuration from JSON file."""
        try:
            with open(config_file, 'r', encoding='utf-8') as f:
                config = json.load(f)
            logger.info(f"Loaded configuration from: {config_file}")
            return config
        except Exception as e:
            logger.warning(f"Failed to load config file {config_file}: {e}. Using defaults.")
            return {}
    
    @staticmethod
    def _analyze_single_file(file_path: Path, output_subdir: Path, **kwargs) -> Dict[str, Any]:
        """
        Analyze a single audio file (static method for parallel processing).

        Produces the canonical ``final_batch_summary`` block from ``SuperAudioAnalyzer`` (linear power masses and percents).

        Returns a flat dict (one row per file) with keys in ``BATCH_SUMMARY_COLUMNS`` plus ``success``.
        """
        try:
            logger.info(f"Analyzing: {file_path.name}")

            # Create analyzer instance
            analyzer = SuperAudioAnalyzer(
                audio_path=file_path,
                output_dir=output_subdir,
                **kwargs
            )

            # CRITICAL FIX: Ensure output directory exists before analysis
            output_subdir.mkdir(parents=True, exist_ok=True)
            logger.info(f"Output directory created/verified: {output_subdir}")
            
            # Run complete analysis (must populate analyzer.results["spectral_component_stats"])
            try:
                analyzer.run_complete_analysis()
            except Exception as analysis_error:
                # Log the error but don't fail completely - try to save what we have
                logger.error(f"Error during run_complete_analysis for {file_path.name}: {analysis_error}", exc_info=True)
                # Try to save partial results if possible
                if hasattr(analyzer, 'save_results'):
                    try:
                        analyzer.save_results()
                        logger.info(f"Saved partial results for {file_path.name}")
                    except Exception as save_error:
                        logger.error(f"Failed to save partial results: {save_error}")
                raise  # Re-raise to be caught by outer exception handler
            
            # Validate that analysis completed successfully
            if not hasattr(analyzer, "results") or not isinstance(analyzer.results, dict):
                raise ValueError("Analysis failed: analyzer.results is missing or invalid")
            
            if "spectral_component_stats" not in analyzer.results:
                raise ValueError("Analysis failed: spectral_component_stats not found in results. "
                               "This may indicate that separate_harmonic_inharmonic() was not called.")
            if "final_batch_summary" not in analyzer.results:
                raise ValueError(
                    "Analysis failed: final_batch_summary not found in results "
                    "(expected linear power-mass batch block)."
                )
            
            # CRITICAL FIX: Explicitly ensure save_results() was called
            # Check if output files exist, if not, call save_results() again
            expected_json = output_subdir / "super_analysis_results.json"
            if not expected_json.exists():
                logger.warning(f"Expected output file not found: {expected_json}. Calling save_results()...")
                try:
                    analyzer.save_results()
                    logger.info(f"Successfully saved results for {file_path.name}")
                except Exception as save_error:
                    logger.error(f"Failed to save results for {file_path.name}: {save_error}", exc_info=True)

            note_name = _extract_note_from_filename(file_path.name)
            fb = analyzer.results.get("final_batch_summary")
            if not isinstance(fb, dict) or fb.get("total_power_mass") is None:
                raise ValueError("Analysis failed: final_batch_summary is missing or incomplete.")

            result: Dict[str, Any] = {
                "success": True,
                "file_name": file_path.name,
                "note": note_name or _extract_note_from_filename(file_path.name),
            }
            for _col in BATCH_SUMMARY_COLUMNS[2:]:
                result[_col] = fb.get(_col)
            logger.info(
                "Final batch power %% for %s (H+I+S, linear ΣA²): H=%.2f%%, I_residual=%.2f%%, S_noise=%.2f%%",
                file_path.name,
                float(fb.get("harmonic_power_percent") or 0.0),
                float(fb.get("inharmonic_residual_power_percent") or 0.0),
                float(fb.get("subbass_noise_power_percent") or 0.0),
            )

            try:
                from metadata_sanitizer import enrich_and_redact_batch_audio_result, publication_redaction_enabled

                if publication_redaction_enabled():
                    tmp = dict(result)
                    tmp["file_path"] = str(file_path)
                    tmp = enrich_and_redact_batch_audio_result(tmp, file_path, note_name)
                    result = {**{k: tmp.get(k) for k in BATCH_SUMMARY_COLUMNS}, "success": True}
            except Exception:
                pass

            logger.info(f"✓ Completed: {file_path.name}")
            return result

        except Exception as e:
            logger.error(f"✗ Error analyzing {file_path.name}: {e}", exc_info=True)
            note_guess = _extract_note_from_filename(file_path.name)
            fail: Dict[str, Any] = {"success": False, "file_name": file_path.name, "note": note_guess}
            for _col in BATCH_SUMMARY_COLUMNS[2:]:
                fail[_col] = None
            try:
                from metadata_sanitizer import enrich_and_redact_batch_audio_result, publication_redaction_enabled

                if publication_redaction_enabled():
                    tmp = dict(fail)
                    tmp["file_path"] = str(file_path)
                    tmp = enrich_and_redact_batch_audio_result(tmp, file_path, note_guess)
                    return {**{k: tmp.get(k) for k in BATCH_SUMMARY_COLUMNS}, "success": False}
            except Exception:
                pass
            return fail

    
    def run_batch_analysis(self) -> Dict[str, Any]:
        """
        Run batch analysis on all audio files.
        
        Returns:
            Dictionary with batch results and summary statistics
        """
        logger.info(f"\n{'='*60}")
        logger.info(f"Starting batch analysis of {len(self.audio_files)} files")
        logger.info(f"{'='*60}\n")
        
        # Prepare output subdirectories
        file_outputs = {}
        for i, file_path in enumerate(self.audio_files):
            subdir_name = f"{i+1:02d}_{file_path.stem}"
            file_outputs[file_path] = self.output_dir / subdir_name
        
        # Run parallel analysis
        results = []
        with ProcessPoolExecutor(max_workers=self.max_workers) as executor:
            # Submit all tasks
            futures = {
                executor.submit(
                    self._analyze_single_file,
                    file_path,
                    file_outputs[file_path],
                    **self.analyzer_kwargs
                ): file_path
                for file_path in self.audio_files
            }
            
            # Collect results as they complete
            completed = 0
            for future in as_completed(futures):
                file_path = futures[future]
                try:
                    result = future.result()
                    results.append(result)
                    completed += 1
                    logger.info(f"Progress: {completed}/{len(self.audio_files)} files completed")
                except Exception as e:
                    logger.error(f"Error processing {file_path}: {e}", exc_info=True)
                    note_guess = _extract_note_from_filename(file_path.name)
                    row_fail: Dict[str, Any] = {
                        "success": False,
                        "file_name": file_path.name,
                        "note": note_guess,
                    }
                    for _col in BATCH_SUMMARY_COLUMNS[2:]:
                        row_fail[_col] = None
                    results.append(row_fail)
        
        self.results = results
        
        # Calculate summary statistics
        self.summary_stats = self._calculate_summary_stats()
        
        # Save batch results
        self._save_batch_results()
        
        logger.info(f"\n{'='*60}")
        logger.info(f"Batch analysis complete!")
        logger.info(f"  Successful: {self.summary_stats['successful_count']}/{len(self.audio_files)}")
        logger.info(f"  Failed: {self.summary_stats['failed_count']}/{len(self.audio_files)}")
        logger.info(f"{'='*60}\n")
        
        return {
            'results': results,
            'summary': self.summary_stats
        }
    
    def _calculate_summary_stats(self) -> Dict[str, Any]:
        """Calculate summary statistics across all analyses."""
        successful = [r for r in self.results if r.get('success', False)]
        failed = [r for r in self.results if not r.get('success', False)]
        
        if len(successful) == 0:
            return {
                'successful_count': 0,
                'failed_count': len(failed),
                'total_files': len(self.results)
            }
        
        # Numeric metrics aligned with ``BATCH_SUMMARY_COLUMNS`` (masses + percents).
        metrics = {
            col: [r[col] for r in successful if r.get(col) is not None]
            for col in BATCH_SUMMARY_COLUMNS[2:]
        }
        
        # Calculate statistics
        stats = {
            'successful_count': len(successful),
            'failed_count': len(failed),
            'total_files': len(self.results),
            'metrics': {}
        }
        
        for metric_name, values in metrics.items():
            if len(values) > 0:
                stats['metrics'][metric_name] = {
                    'mean': float(np.mean(values)),
                    'std': float(np.std(values)),
                    'min': float(np.min(values)),
                    'max': float(np.max(values)),
                    'median': float(np.median(values))
                }

        return stats
    
    @staticmethod
    def regenerate_excel_from_json(json_path: Path, excel_path: Optional[Path] = None) -> bool:
        """
        Regenerate Excel summary from existing JSON batch results.
        
        This ensures the Excel file matches the JSON data exactly, useful when
        the Excel was generated with older code or needs to be updated.
        
        Args:
            json_path: Path to batch_results.json file
            excel_path: Optional path for output Excel file. If None, uses same directory as JSON.
            
        Returns:
            True if successful, False otherwise
        """
        try:
            if not json_path.exists():
                logger.error(f"JSON file not found: {json_path}")
                return False
            
            if excel_path is None:
                excel_path = json_path.parent / "batch_summary.xlsx"
            
            logger.info(f"Regenerating Excel from JSON: {json_path}")
            
            # Read JSON
            with open(json_path, 'r', encoding='utf-8') as f:
                json_data = json.load(f)
            
            results = json_data.get('results', [])
            if len(results) == 0:
                logger.error("No results found in JSON file")
                return False
            
            # Deduplicate by file_name (legacy JSON may contain extra keys; export uses ``BATCH_SUMMARY_COLUMNS`` only).
            rows_in = list(results)
            seen_fn: set[str] = set()
            rows_dedup: List[Dict[str, Any]] = []
            for r in rows_in:
                if not isinstance(r, dict):
                    continue
                fn = r.get("file_name")
                if fn is None:
                    continue
                k = str(fn)
                if k in seen_fn:
                    continue
                seen_fn.add(k)
                rows_dedup.append(r)

            excel_path.parent.mkdir(parents=True, exist_ok=True)
            export_df = BatchAudioAnalyzer._batch_rows_to_summary_dataframe(rows_dedup)
            try:
                from metadata_sanitizer import publication_redaction_enabled, sanitize_dataframe_for_publication

                if publication_redaction_enabled() and not export_df.empty:
                    export_df = sanitize_dataframe_for_publication(export_df)
            except Exception:
                pass
            with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
                export_df.to_excel(writer, sheet_name="Batch Summary", index=False)
                worksheet = writer.sheets["Batch Summary"]
                from openpyxl.utils import get_column_letter
                from openpyxl.styles import Font, PatternFill, Alignment

                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                for col_idx, col_name in enumerate(export_df.columns, 1):
                    max_length = max(
                        int(export_df[col_name].astype(str).map(len).max()) if len(export_df) > 0 else 0,
                        len(str(col_name)),
                    )
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

            logger.info("Batch summary written to %s (columns: %s)", excel_path, ", ".join(BATCH_SUMMARY_COLUMNS))
            logger.info("✓ Excel file regenerated: %s", excel_path)
            return True
            
        except Exception as e:
            logger.error(f"Error regenerating Excel from JSON: {e}", exc_info=True)
            return False
    
    @staticmethod
    def _note_to_midi_number(note_str: str) -> int:
        """
        Convert musical note to MIDI number for chromatic ordering.
        
        Chromatic order: C, C#, D, D#, E, F, F#, G, G#, A, A#, B
        MIDI formula: (octave + 1) * 12 + semitone_offset
        
        Args:
            note_str: Note name (e.g., 'A4', 'C#3', 'Bb2') or filename containing note
            
        Returns:
            MIDI note number (0-127) or 9999 if invalid (sorts to end)
        """
        import re
        import librosa
        
        # Extract note from string (handle both note strings and filenames)
        patterns = [
            r'([A-G][#b]?)(\d+)',  # Standard: A4, C#3, Bb2
            r'([A-G][♯♭]?)(\d+)',  # Unicode sharps/flats
        ]
        
        note_match = None
        for pattern in patterns:
            match = re.search(pattern, note_str)
            if match:
                note_match = match
                break
        
        if not note_match:
            return 9999  # Invalid note, sort to end
        
        note_name = note_match.group(1).replace('♯', '#').replace('♭', 'b')
        octave = int(note_match.group(2))
        
        try:
            # Use librosa to convert note to MIDI number
            full_note = f"{note_name}{octave}"
            midi_num = librosa.note_to_midi(full_note)
            return int(midi_num) if 0 <= midi_num <= 127 else 9999
        except Exception:
            # Fallback: manual calculation
            semitone_map = {
                'C': 0, 'C#': 1, 'Db': 1,
                'D': 2, 'D#': 3, 'Eb': 3,
                'E': 4, 'Fb': 4, 'E#': 5,
                'F': 5, 'F#': 6, 'Gb': 6,
                'G': 7, 'G#': 8, 'Ab': 8,
                'A': 9, 'A#': 10, 'Bb': 10,
                'B': 11, 'Cb': 11, 'B#': 0
            }
            semitone = semitone_map.get(note_name, 0)
            midi_num = (octave + 1) * 12 + semitone
            return int(midi_num) if 0 <= midi_num <= 127 else 9999

    @staticmethod
    def _batch_rows_to_summary_dataframe(rows: List[Dict[str, Any]]) -> pd.DataFrame:
        """Build the single-sheet batch summary: one row per file, fixed columns, chromatic sort."""
        deduped = _dedupe_batch_rows_by_file_name(rows)
        records: List[Dict[str, Any]] = []
        for r in deduped:
            if not isinstance(r, dict):
                continue
            fn = r.get("file_name")
            if fn is None:
                continue
            fns = str(fn).strip().upper()
            if not fns or fns in ("MEAN", "MEDIAN") or "TIER" in fns:
                continue
            pub = _batch_row_public_dict(r)
            if pub.get("note") is None or (isinstance(pub.get("note"), float) and pd.isna(pub.get("note"))):
                pub["note"] = (
                    r.get("note")
                    or r.get("Note")
                    or r.get("note_name")
                    or _extract_note_from_filename(str(fn))
                )
            records.append(pub)
        df = pd.DataFrame.from_records(records, columns=list(BATCH_SUMMARY_COLUMNS))
        if df.empty:
            return df
        df["_midi_sort"] = df["note"].astype(str).map(BatchAudioAnalyzer._note_to_midi_number)
        return df.sort_values("_midi_sort").drop(columns=["_midi_sort"]).reset_index(drop=True)

    def _save_batch_results(self) -> None:
        """Save batch results to files."""
        # Save detailed results JSON
        results_path = self.output_dir / "batch_results.json"
        try:
            from metadata_sanitizer import publication_redaction_enabled, sanitize_metadata_dict

            pub_rows = [_batch_row_public_dict(x) for x in self.results]
            _rows = (
                [sanitize_metadata_dict(dict(x)) for x in pub_rows]
                if publication_redaction_enabled()
                else pub_rows
            )
        except Exception:
            _rows = [_batch_row_public_dict(x) for x in self.results]
        with open(results_path, 'w', encoding='utf-8') as f:
            json.dump({
                'batch_analysis_date': datetime.now().isoformat(),
                'total_files': len(self.audio_files),
                'pipeline_options': {
                    'minimal_spectral_probe': bool(
                        self.analyzer_kwargs.get('minimal_spectral_probe', False)
                    ),
                },
                'results': _rows,
                'summary': self.summary_stats
            }, f, indent=2, default=str)
        
        logger.info(f"Saved batch results: {results_path}")
        
        # Save summary Excel (``BATCH_SUMMARY_COLUMNS`` only; chromatic sort via ``_batch_rows_to_summary_dataframe``).
        if len(self.results) > 0:
            export_df = BatchAudioAnalyzer._batch_rows_to_summary_dataframe(self.results)
            excel_path = self.output_dir / "batch_summary.xlsx"
            try:
                from metadata_sanitizer import publication_redaction_enabled, sanitize_dataframe_for_publication

                if publication_redaction_enabled() and not export_df.empty:
                    export_df = sanitize_dataframe_for_publication(export_df)
            except Exception:
                pass
            try:
                # Try to save as Excel with proper formatting
                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    export_df.to_excel(writer, sheet_name='Batch Summary', index=False)
                    # Auto-adjust column widths and format MEAN/MEDIAN rows
                    worksheet = writer.sheets['Batch Summary']
                    from openpyxl.utils import get_column_letter
                    from openpyxl.styles import Font, PatternFill, Alignment
                    
                    # Find all summary rows (tier rows, MEAN, and MEDIAN)
                    tier_mean_indices = []
                    tier_median_indices = []
                    mean_row_idx = None
                    median_row_idx = None
                    
                    for idx, row in export_df.iterrows():
                        file_name_str = str(row.get('file_name', '')).upper()
                        if file_name_str.startswith('TIER') and file_name_str.endswith('_MEAN'):
                            tier_mean_indices.append(idx + 2)  # +2 because Excel is 1-indexed and has header
                        elif file_name_str.startswith('TIER') and file_name_str.endswith('_MEDIAN'):
                            tier_median_indices.append(idx + 2)
                        elif file_name_str == 'MEAN':
                            mean_row_idx = idx + 2
                        elif file_name_str == 'MEDIAN':
                            median_row_idx = idx + 2
                    
                    # Format header row
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Format tier MEAN rows (lighter blue)
                    tier_mean_fill = PatternFill(start_color="B4D7ED", end_color="B4D7ED", fill_type="solid")
                    tier_mean_font = Font(bold=True, italic=True)
                    for row_idx in tier_mean_indices:
                        for cell in worksheet[row_idx]:
                            cell.fill = tier_mean_fill
                            cell.font = tier_mean_font
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Format tier MEDIAN rows (lighter green)
                    tier_median_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
                    tier_median_font = Font(bold=True, italic=True)
                    for row_idx in tier_median_indices:
                        for cell in worksheet[row_idx]:
                            cell.fill = tier_median_fill
                            cell.font = tier_median_font
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Format overall MEAN row (darker blue)
                    if mean_row_idx:
                        mean_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                        mean_font = Font(bold=True)
                        for cell in worksheet[mean_row_idx]:
                            cell.fill = mean_fill
                            cell.font = mean_font
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Format overall MEDIAN row (darker green)
                    if median_row_idx:
                        median_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                        median_font = Font(bold=True)
                        for cell in worksheet[median_row_idx]:
                            cell.fill = median_fill
                            cell.font = median_font
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Auto-adjust column widths
                    for idx, col in enumerate(export_df.columns, start=1):
                        max_length = max(
                            export_df[col].astype(str).map(len).max() if len(export_df) > 0 else 0,
                            len(str(col))
                        )
                        # Limit max width to 50 characters
                        max_length = min(max_length, 50)
                        column_letter = get_column_letter(idx)
                        worksheet.column_dimensions[column_letter].width = max_length + 2
                
                logger.info("Batch summary (public columns) written to %s", excel_path)
            except ImportError:
                # Fallback to CSV if openpyxl not available
                logger.warning("openpyxl not available, saving as CSV instead. Install with: pip install openpyxl")
                csv_path = self.output_dir / "batch_summary.csv"
                export_df.to_csv(csv_path, index=False)
                logger.info(f"Saved summary CSV (fallback): {csv_path}")
            except Exception as e:
                # Fallback to CSV on any error
                logger.warning(f"Error saving Excel file: {e}, saving as CSV instead")
                csv_path = self.output_dir / "batch_summary.csv"
                export_df.to_csv(csv_path, index=False)
                logger.info(f"Saved summary CSV (fallback): {csv_path}")

        # Save summary statistics
        stats_path = self.output_dir / "batch_statistics.txt"
        with open(stats_path, 'w', encoding='utf-8') as f:
            f.write("BATCH AUDIO ANALYSIS - SUMMARY STATISTICS\n")
            f.write("=" * 60 + "\n\n")
            f.write(f"Analysis Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Total Files: {self.summary_stats['total_files']}\n")
            f.write(f"Successful: {self.summary_stats['successful_count']}\n")
            f.write(f"Failed: {self.summary_stats['failed_count']}\n")
            f.write(
                f"Light Phase 1 (minimal_spectral_probe): "
                f"{'YES' if self.analyzer_kwargs.get('minimal_spectral_probe') else 'NO'}\n"
                f"  -> When YES: per-file folder omits long metrics_summary / spectrum CSVs; "
                f"see batch_weight_probe_pie.png and super_analysis_results.json.\n\n"
            )

            if 'metrics' in self.summary_stats:
                f.write("METRICS SUMMARY\n")
                f.write("-" * 60 + "\n")
                for metric_name, metric_stats in self.summary_stats['metrics'].items():
                    f.write(f"\n{metric_name.replace('_', ' ').title()}:\n")
                    f.write(f"  Mean:   {metric_stats['mean']:.4f}\n")
                    f.write(f"  Std:    {metric_stats['std']:.4f}\n")
                    f.write(f"  Min:    {metric_stats['min']:.4f}\n")
                    f.write(f"  Max:    {metric_stats['max']:.4f}\n")
                    f.write(f"  Median: {metric_stats['median']:.4f}\n")
        
        logger.info(f"Saved statistics: {stats_path}")


def main():
    """Command-line interface for batch analysis."""
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Batch Audio Analyzer - Process multiple audio files',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Use default config (auto-extract weights per file)
  python batch_audio_analyzer.py *.wav
  
  # Use custom config file
  python batch_audio_analyzer.py *.wav --config batch_config.json
  
  # Override config with command-line arguments
  python batch_audio_analyzer.py *.wav --config batch_config.json --harmonic-tolerance 0.03
  
  # Regenerate Excel from existing JSON
  python batch_audio_analyzer.py --regenerate-excel batch_results.json
        """
    )
    parser.add_argument(
        'audio_files',
        nargs='*',
        type=Path,
        help='Audio files to analyze (up to 100 files). Omit if using --regenerate-excel.'
    )
    parser.add_argument(
        '--regenerate-excel',
        type=Path,
        metavar='JSON_PATH',
        help='Regenerate Excel file from existing batch_results.json. Provide path to JSON file.'
    )
    parser.add_argument(
        '--output-dir',
        type=Path,
        default=Path('batch_results'),
        help='Output directory for batch results (default: batch_results/)'
    )
    parser.add_argument(
        '--config',
        type=Path,
        default=None,
        help='Path to batch configuration JSON file (default: batch_config.json if exists)'
    )
    parser.add_argument(
        '--max-workers',
        type=int,
        default=None,
        help='Maximum parallel workers (default: auto)'
    )
    parser.add_argument(
        '--harmonic-tolerance',
        type=float,
        default=None,
        help='Harmonic tolerance (overrides config)'
    )
    parser.add_argument(
        '--use-90-tier',
        action='store_true',
        default=None,
        help='Use 90-tier granular clustering (overrides config)'
    )
    parser.add_argument(
        '--weight-mode',
        choices=['auto_extract', 'fixed'],
        default=None,
        help='Weight mode: auto_extract (per-file) or fixed (from config)'
    )
    parser.add_argument(
        '--harmonic-weight',
        type=float,
        default=None,
        help='Fixed harmonic weight (0.0-1.0, only used with --weight-mode fixed)'
    )
    parser.add_argument(
        '--inharmonic-weight',
        type=float,
        default=None,
        help='Fixed inharmonic weight (0.0-1.0, only used with --weight-mode fixed)'
    )
    
    args = parser.parse_args()
    
    # Handle regenerate Excel option
    if args.regenerate_excel:
        json_path = Path(args.regenerate_excel)
        if not json_path.exists():
            print(f"ERROR: JSON file not found: {json_path}")
            sys.exit(1)
        
        print(f"Regenerating Excel from JSON: {json_path}")
        success = BatchAudioAnalyzer.regenerate_excel_from_json(json_path)
        if success:
            excel_path = json_path.parent / "batch_summary.xlsx"
            print(f"\n✓ SUCCESS! Excel file regenerated: {excel_path}")
            sys.exit(0)
        else:
            print("\n✗ ERROR: Failed to regenerate Excel file")
            sys.exit(1)
    
    # Check file limit
    if len(args.audio_files) > 100:
        print(f"ERROR: Maximum 100 files supported. Got {len(args.audio_files)} files.")
        sys.exit(1)
    
    # Determine config file
    config_file = args.config
    if config_file is None:
        default_config = Path('batch_config.json')
        if default_config.exists():
            config_file = default_config
            print(f"Using default config: {config_file}")
    
    # Prepare analyzer kwargs
    analyzer_kwargs = {}
    if args.harmonic_tolerance is not None:
        analyzer_kwargs['harmonic_tolerance'] = args.harmonic_tolerance
    if args.use_90_tier is not None:
        analyzer_kwargs['use_90_tier'] = args.use_90_tier
    
    # Handle weight mode override
    if args.weight_mode == 'fixed':
        if args.harmonic_weight is not None and args.inharmonic_weight is not None:
            analyzer_kwargs['auto_extract_weights'] = False
            analyzer_kwargs['harmonic_weight'] = args.harmonic_weight
            analyzer_kwargs['inharmonic_weight'] = args.inharmonic_weight
            print(f"Using FIXED weights: harmonic={args.harmonic_weight}, inharmonic={args.inharmonic_weight}")
        else:
            print("WARNING: --weight-mode fixed requires --harmonic-weight and --inharmonic-weight")
    
    # Create batch analyzer
    analyzer = BatchAudioAnalyzer(
        audio_files=args.audio_files,
        output_dir=args.output_dir,
        max_workers=args.max_workers,
        config_file=config_file,
        **analyzer_kwargs
    )
    
    # Run analysis
    results = analyzer.run_batch_analysis()
    
    print(f"\n✓ Batch analysis complete!")
    print(f"  Results saved to: {args.output_dir}")
    print(f"  Successful: {results['summary']['successful_count']}/{len(args.audio_files)}")
    print(f"  Failed: {results['summary']['failed_count']}/{len(args.audio_files)}")


if __name__ == '__main__':
    main()

